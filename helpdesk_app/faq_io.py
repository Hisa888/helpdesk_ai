from __future__ import annotations

import csv
import io
import os
import re
import zipfile
import threading
from datetime import datetime
from pathlib import Path

import pandas as pd

from helpdesk_app.faq_db import (
    is_managed_faq_path,
    load_faq_df_from_db,
    save_faq_df_to_db,
    sync_faq_csv_cache,
    initialize_faq_database as _initialize_faq_database,
)


# FAQ標準スキーマ。
# faq_id / enabled は FAQ 一括メンテナンス（+追加 / 空白更新 / -削除）で使用する管理項目です。
# answer_format は既存のリンク/Markdown/HTML表示機能を残すため保持する。
FAQ_COLUMNS = [
    "faq_id", "question", "answer", "intent", "keywords", "category",
    "answer_format", "enabled", "updated_at", "updated_by", "note",
    # 検索安全制御用（任意）。空欄なら自動推定ルールで補完します。
    "required_keywords", "exclude_keywords", "ambiguity_keywords",
    "prefer_candidate", "auto_answer_allowed",
]
FAQ_SEARCH_COLUMNS = ["question", "answer", "intent", "keywords", "category"]
FAQ_EXPORT_JA_COLUMNS = {
    "faq_id": "FAQ_ID",
    "question": "質問",
    "answer": "回答",
    "intent": "意図",
    "keywords": "キーワード・言い換え",
    "category": "カテゴリ",
    "answer_format": "表示形式",
    "enabled": "有効",
    "updated_at": "更新日",
    "updated_by": "更新者",
    "note": "備考",
    "required_keywords": "必須キーワード",
    "exclude_keywords": "除外キーワード",
    "ambiguity_keywords": "曖昧判定キーワード",
    "prefer_candidate": "候補表示優先",
    "auto_answer_allowed": "自動回答許可",
}
FAQ_IMPORT_OPERATION_COLUMN = "operation"
FAQ_IMPORT_OPERATION_EXPORT_NAME = "操作"


# ===== CSV読み込みを頑丈にする（文字コード/区切り/カラム揺れ対策）=====
def _read_csv_flexible_raw(path: Path) -> pd.DataFrame:
    """DBを参照せず、CSVファイルだけを柔軟に読む内部関数。"""
    raw = path.read_bytes()
    encs = ["utf-8", "utf-8-sig", "cp932", "shift_jis"]
    text = None
    for enc in encs:
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        text = raw.decode("latin1", errors="ignore")

    import csv as _csv

    sample = text[:5000]
    delim = ","
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        delim = dialect.delimiter
    except Exception:
        if sample.count("\t") > sample.count(","):
            delim = "\t"

    try:
        return pd.read_csv(io.StringIO(text), sep=delim, engine="python", on_bad_lines="skip")
    except Exception:
        try:
            return pd.read_csv(io.StringIO(text), sep=delim, engine="python")
        except Exception:
            return pd.DataFrame()


def read_csv_flexible(path: Path) -> pd.DataFrame:
    """CSVをできる限り失敗しないで読む。

    runtime_data/faq.csv については、SQLite(runtime_data/helpdesk.db)を正本として読み、
    既存の検索/UIが壊れないようにCSVはキャッシュとして残す。
    """
    path = Path(path)
    try:
        db_df = load_faq_df_from_db(path)
        if db_df is not None:
            return db_df
    except Exception:
        pass
    return _read_csv_flexible_raw(path)


def initialize_faq_database(faq_path: Path) -> bool:
    """既存faq.csvからruntime_data/helpdesk.dbを初期作成する。"""
    try:
        return _initialize_faq_database(Path(faq_path), _read_csv_flexible_raw)
    except Exception:
        return False


def pick_question_column(cols) -> str | None:
    """質問カラム名の揺れを吸収"""
    cand = [
        "question", "質問", "問い合わせ", "問合せ", "query", "user_question",
        "content", "text",
    ]
    for c in cand:
        if c in cols:
            return c
    lower_map = {str(c).lower(): c for c in cols}
    for c in cand:
        if c.lower() in lower_map:
            return lower_map[c.lower()]
    return None


def _is_false_like(value) -> bool:
    s = str(value or "").strip().lower()
    return s in {"false", "0", "no", "n", "off", "disabled", "disable", "無効", "削除", "停止", "×", "✕"}


def _normalize_enabled(value, default: bool = True) -> str:
    s = str(value or "").strip()
    if not s:
        return "TRUE" if default else "FALSE"
    return "FALSE" if _is_false_like(s) else "TRUE"


def _extract_faq_id_number(value: str) -> int | None:
    m = re.search(r"(\d+)$", str(value or "").strip())
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _assign_missing_faq_ids(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "faq_id" not in out.columns:
        out["faq_id"] = ""

    used = {str(v).strip() for v in out["faq_id"].fillna("").astype(str).tolist() if str(v).strip()}
    max_no = 0
    for fid in used:
        no = _extract_faq_id_number(fid)
        if no is not None:
            max_no = max(max_no, no)

    next_no = max_no + 1
    assigned: list[str] = []
    for value in out["faq_id"].fillna("").astype(str).tolist():
        fid = str(value or "").strip()
        if fid:
            assigned.append(fid)
            continue
        while True:
            candidate = f"FAQ-{next_no:06d}"
            next_no += 1
            if candidate not in used:
                used.add(candidate)
                assigned.append(candidate)
                break
    out["faq_id"] = assigned
    return out


def normalize_faq_columns(
    df: pd.DataFrame,
    *,
    drop_empty_required: bool = True,
    assign_ids: bool = True,
) -> pd.DataFrame:
    """列名揺れを吸収して FAQ_COLUMNS に正規化する。

    旧形式（question/answer/category/answer_format）もそのまま読み込めるようにしつつ、
    intent / keywords を追加して検索精度改善に使える形へ寄せる。

    drop_empty_required=False は、差分アップロードの「-削除」行など、回答が空でも
    いったん行として読ませたい場合に使う。
    """
    from helpdesk_app.modules.faq_answer_renderer import normalize_answer_format

    base_cols = FAQ_COLUMNS
    if df is None or len(df) == 0:
        return pd.DataFrame(columns=base_cols)
    out = df.copy()
    rename_map = {}
    for c in out.columns:
        original = str(c).strip()
        key = original.lower().strip()
        if original in ["FAQ_ID", "FAQID", "FAQ番号", "管理ID", "ID", "id"] or key in ["faq_id", "faqid", "faq id", "id"]:
            rename_map[c] = "faq_id"
        elif original in ["質問", "問い合わせ", "問合せ", "ユーザー質問", "ユーザー言語"] or key in ["question", "query", "user_question"]:
            rename_map[c] = "question"
        elif original in ["回答", "答え", "対応方法"] or key in ["answer", "answer_text", "reply"]:
            rename_map[c] = "answer"
        elif original in ["意図", "意味", "目的", "問い合わせ意図", "インテント"] or key in ["intent", "meaning", "purpose"]:
            rename_map[c] = "intent"
        elif original in ["キーワード", "検索語", "言い換え", "類義語", "同義語", "別表現", "キーワード・言い換え"] or key in ["keywords", "keyword", "synonyms", "aliases", "phrases"]:
            rename_map[c] = "keywords"
        elif original in ["カテゴリ", "カテゴリー", "分類", "区分"] or key in ["category", "group"]:
            rename_map[c] = "category"
        elif original in ["表示形式", "回答表示形式", "回答形式"] or key in ["answer_format", "display_format", "format", "answer_type"]:
            rename_map[c] = "answer_format"
        elif original in ["有効", "有効フラグ", "利用", "利用可", "公開", "enabled"] or key in ["enabled", "is_enabled", "active"]:
            rename_map[c] = "enabled"
        elif original in ["更新日", "更新日時", "最終更新日", "変更日"] or key in ["updated_at", "updated", "modified_at", "last_updated"]:
            rename_map[c] = "updated_at"
        elif original in ["更新者", "担当者", "変更者", "登録者"] or key in ["updated_by", "updater", "modified_by", "owner"]:
            rename_map[c] = "updated_by"
        elif original in ["備考", "メモ", "コメント", "注記"] or key in ["note", "notes", "memo", "comment", "remarks"]:
            rename_map[c] = "note"
        elif original in ["必須キーワード", "必須語", "必須検索語", "必要キーワード"] or key in ["required_keywords", "required", "must_keywords", "must_terms"]:
            rename_map[c] = "required_keywords"
        elif original in ["除外キーワード", "除外語", "除外検索語", "禁止キーワード"] or key in ["exclude_keywords", "excluded_keywords", "exclude", "negative_keywords", "ng_keywords"]:
            rename_map[c] = "exclude_keywords"
        elif original in ["曖昧判定キーワード", "曖昧キーワード", "候補表示キーワード", "確認キーワード"] or key in ["ambiguity_keywords", "ambiguous_keywords", "clarify_keywords", "candidate_keywords"]:
            rename_map[c] = "ambiguity_keywords"
        elif original in ["候補表示優先", "候補優先", "候補表示"] or key in ["prefer_candidate", "candidate_first", "suggest_first"]:
            rename_map[c] = "prefer_candidate"
        elif original in ["自動回答許可", "自動回答", "直答許可"] or key in ["auto_answer_allowed", "allow_auto_answer", "auto_answer"]:
            rename_map[c] = "auto_answer_allowed"
    out = out.rename(columns=rename_map)
    out = out.loc[:, ~out.columns.duplicated()].copy()
    for col in base_cols:
        if col not in out.columns:
            if col == "answer_format":
                out[col] = "markdown"
            elif col == "enabled":
                out[col] = "TRUE"
            else:
                out[col] = ""
    out = out[base_cols].copy()

    for col in ["faq_id", "question", "intent", "keywords", "category", "required_keywords", "exclude_keywords", "ambiguity_keywords"]:
        out[col] = (
            out[col]
            .fillna("")
            .astype(str)
            .str.replace("\n", " ", regex=False)
            .str.replace("\r", " ", regex=False)
            .str.strip()
        )
    out["answer"] = (
        out["answer"]
        .fillna("")
        .astype(str)
        .str.replace("\r\n", "\n", regex=False)
        .str.replace("\r", "\n", regex=False)
        .str.strip()
    )
    out["answer_format"] = out["answer_format"].map(lambda v: normalize_answer_format(v, default="markdown"))
    out["enabled"] = out["enabled"].map(lambda v: _normalize_enabled(v, default=True))
    for col in ["updated_at", "updated_by", "note"]:
        if col in out.columns:
            out[col] = (
                out[col]
                .fillna("")
                .astype(str)
                .str.replace("\r\n", "\n", regex=False)
                .str.replace("\r", "\n", regex=False)
                .str.strip()
            )
    for col in ["prefer_candidate", "auto_answer_allowed"]:
        if col in out.columns:
            out[col] = out[col].fillna("").astype(str).str.strip()

    if drop_empty_required:
        out = out[(out["question"] != "") & (out["answer"] != "")].reset_index(drop=True)
    else:
        out = out.reset_index(drop=True)

    if assign_ids and len(out) > 0:
        out = _assign_missing_faq_ids(out)
    return out

def _xlsx_escape(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', '&quot;')
    )


def _build_minimal_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "FAQ") -> bytes:
    from io import BytesIO

    rows = [list(df.columns)] + df.fillna("").astype(str).values.tolist()
    shared_strings: list[str] = []
    sst_index: dict[str, int] = {}

    def get_sst_idx(value: str) -> int:
        value = str(value)
        if value not in sst_index:
            sst_index[value] = len(shared_strings)
            shared_strings.append(value)
        return sst_index[value]

    sheet_rows = []
    for r_idx, row in enumerate(rows, start=1):
        cells = []
        for c_idx, value in enumerate(row, start=1):
            col = ""
            n = c_idx
            while n:
                n, rem = divmod(n - 1, 26)
                col = chr(65 + rem) + col
            ref = f"{col}{r_idx}"
            s_idx = get_sst_idx(value)
            cells.append(f'<c r="{ref}" t="s"><v>{s_idx}</v></c>')
        sheet_rows.append(f'<row r="{r_idx}">' + "".join(cells) + "</row>")
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetData>' + ''.join(sheet_rows) + '</sheetData>'
        '</worksheet>'
    )
    sst_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(shared_strings)}" uniqueCount="{len(shared_strings)}">'
        + ''.join(f'<si><t xml:space="preserve">{_xlsx_escape(s)}</t></si>' for s in shared_strings)
        + '</sst>'
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets><sheet name="{_xlsx_escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets>'
        '</workbook>'
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        '</Relationships>'
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        '</Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        '</Types>'
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        '</styleSheet>'
    )
    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        '<dc:title>FAQ</dc:title><dc:creator>ChatGPT</dc:creator></cp:coreProperties>'
    )
    app_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        '<Application>Microsoft Excel</Application></Properties>'
    )
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", root_rels)
        zf.writestr("docProps/core.xml", core_xml)
        zf.writestr("docProps/app.xml", app_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        zf.writestr("xl/styles.xml", styles_xml)
        zf.writestr("xl/sharedStrings.xml", sst_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return bio.getvalue()


def faq_df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    has_operation_input = FAQ_IMPORT_OPERATION_COLUMN in getattr(df, "columns", []) or FAQ_IMPORT_OPERATION_EXPORT_NAME in getattr(df, "columns", [])
    clean = normalize_faq_columns(df, drop_empty_required=not has_operation_input)
    export_df = clean.rename(columns=FAQ_EXPORT_JA_COLUMNS)
    if FAQ_IMPORT_OPERATION_COLUMN in getattr(df, "columns", []):
        operation_values = df[FAQ_IMPORT_OPERATION_COLUMN].fillna("").astype(str).tolist()
        operation_values = [_normalize_operation(v) for v in operation_values]
    elif FAQ_IMPORT_OPERATION_EXPORT_NAME in getattr(df, "columns", []):
        operation_values = df[FAQ_IMPORT_OPERATION_EXPORT_NAME].fillna("").astype(str).tolist()
        operation_values = [_normalize_operation(v) for v in operation_values]
    else:
        operation_values = [""] * len(export_df)
    if len(operation_values) != len(export_df):
        operation_values = [""] * len(export_df)
    export_df.insert(0, FAQ_IMPORT_OPERATION_EXPORT_NAME, operation_values)
    return _build_minimal_xlsx_bytes(export_df, sheet_name="FAQ")


def _read_xlsx_bytes(raw: bytes) -> pd.DataFrame:
    """XLSXを安全にDataFrame化する。
    1) pandas + openpyxl を優先
    2) openpyxl の values_only 読み
    3) 最後の手段で XML 手動解析
    """
    from io import BytesIO

    try:
        df = pd.read_excel(BytesIO(raw), engine="openpyxl")
        if df is not None:
            df.columns = [str(c).strip() for c in df.columns]
            return df
    except Exception:
        pass

    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            max_cols = max(len(r) if r is not None else 0 for r in rows)
            norm_rows = []
            for r in rows:
                r = list(r or [])
                r += [""] * (max_cols - len(r))
                norm_rows.append(["" if v is None else str(v) for v in r])
            header = [str(x).strip() for x in norm_rows[0]]
            body = norm_rows[1:] if len(norm_rows) > 1 else []
            return pd.DataFrame(body, columns=header)
    except Exception:
        pass

    import xml.etree.ElementTree as ET

    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def _text_without_rph(parent):
        parts = []
        for child in list(parent):
            tag = child.tag.split("}")[-1]
            if tag == "rPh":
                continue
            if tag == "t":
                parts.append(child.text or "")
            else:
                for tnode in child.findall(".//a:t", ns):
                    parts.append(tnode.text or "")
        return "".join(parts)

    with zipfile.ZipFile(BytesIO(raw)) as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", ns):
                shared.append(_text_without_rph(si))

        sheet_path = "xl/worksheets/sheet1.xml"
        if "xl/workbook.xml" in zf.namelist() and "xl/_rels/workbook.xml.rels" in zf.namelist():
            wb = ET.fromstring(zf.read("xl/workbook.xml"))
            rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rid_to_target = {
                rel.attrib.get("Id"): rel.attrib.get("Target")
                for rel in rels.findall("pr:Relationship", ns)
            }
            first_sheet = wb.find(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/"
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
            )
            if first_sheet is not None:
                rid = first_sheet.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                target = rid_to_target.get(rid)
                if target:
                    if not target.startswith("worksheets/"):
                        target = target.split("xl/")[-1]
                    sheet_path = "xl/" + target

        sheet = ET.fromstring(zf.read(sheet_path))
        rows = []
        for row in sheet.findall("a:sheetData/a:row", ns):
            row_map = {}
            max_col = 0
            for c in row.findall("a:c", ns):
                ref = c.attrib.get("r", "")
                col_letters = "".join(ch for ch in ref if ch.isalpha())
                col_idx = 0
                for ch in col_letters:
                    col_idx = col_idx * 26 + (ord(ch.upper()) - 64)
                max_col = max(max_col, col_idx)

                t = c.attrib.get("t")
                value = ""
                if t == "s":
                    v = c.find("a:v", ns)
                    if v is not None and (v.text or "").isdigit():
                        si_idx = int(v.text)
                        value = shared[si_idx] if 0 <= si_idx < len(shared) else ""
                elif t == "inlineStr":
                    is_el = c.find("a:is", ns)
                    if is_el is not None:
                        value = _text_without_rph(is_el)
                else:
                    v = c.find("a:v", ns)
                    value = v.text if v is not None and v.text is not None else ""

                if col_idx > 0:
                    row_map[col_idx] = value

            if max_col:
                rows.append([row_map.get(i, "") for i in range(1, max_col + 1)])

    if not rows:
        return pd.DataFrame()
    max_cols = max(len(r) for r in rows)
    rows = [r + [""] * (max_cols - len(r)) for r in rows]
    header = [str(x).strip() for x in rows[0]]
    body = rows[1:] if len(rows) > 1 else []
    return pd.DataFrame(body, columns=header)


def _strip_excel_phonetic_artifacts(text: str) -> str:
    """Excel由来のふりがな混入っぽい末尾カタカナを保守的に除去する。"""
    s = "" if text is None else str(text)
    s = s.replace("\u3000", " ").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not s:
        return s

    s = re.sub(r"([。．.!?！？]\s*)([ァ-ヶーｦ-ﾟ]{1,20})$", r"\1", s)

    m = re.match(r"^(.*[一-龥々〆ヵヶ])([ァ-ヶー]{2,20})$", s)
    if m:
        prefix, _suffix = m.groups()
        if not re.search(r"[ァ-ヶー]$", prefix):
            s = prefix

    return s.strip()


def _read_faq_raw_uploaded_file(file_name: str, raw: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        tmp = Path("/tmp/_faq_upload.csv")
        tmp.write_bytes(raw)
        return read_csv_flexible(tmp)
    return _read_xlsx_bytes(raw)


def read_faq_uploaded_file(file_name: str, raw: bytes) -> pd.DataFrame:
    """FAQアップロード読込。旧来の全置換用。"""
    df = _read_faq_raw_uploaded_file(file_name, raw)

    df = normalize_faq_columns(df)
    for col in ["question", "answer", "intent", "keywords", "category"]:
        if col in df.columns:
            df[col] = (
                df[col]
                .fillna("")
                .astype(str)
                .map(_strip_excel_phonetic_artifacts)
                .str.replace("\u3000", " ", regex=False)
                .str.replace("\r\n", "\n", regex=False)
                .str.replace("\r", "\n", regex=False)
                .str.strip()
            )
    return df


def _find_operation_column(cols) -> str | None:
    candidates = {"操作", "処理", "operation", "op", "action", "mode"}
    lower_candidates = {c.lower() for c in candidates}
    for c in cols:
        raw = str(c).strip()
        if raw in candidates or raw.lower() in lower_candidates:
            return c
    return None


def _has_faq_id_column(cols) -> bool:
    names = {"FAQ_ID", "FAQID", "FAQ番号", "管理ID", "ID", "id", "faq_id", "faqid", "faq id"}
    lowers = {n.lower() for n in names}
    return any(str(c).strip() in names or str(c).strip().lower() in lowers for c in cols)


def _normalize_operation(value) -> str:
    s = str(value or "").strip()
    low = s.lower()
    if s in {"+", "＋"} or low in {"add", "insert", "new", "create"} or s in {"追加", "新規", "登録"}:
        return "+"
    if s in {"-", "－", "−", "ー"} or low in {"delete", "remove", "disable", "disabled"} or s in {"削除", "無効", "除外"}:
        return "-"
    return ""


def read_faq_operation_uploaded_file(file_name: str, raw: bytes) -> pd.DataFrame:
    """FAQ差分アップロード読込。

    操作列の正式仕様:
      +   追加
      -   削除
      空白 更新
    """
    raw_df = _read_faq_raw_uploaded_file(file_name, raw)
    op_col = _find_operation_column(raw_df.columns)
    has_operation_spec = bool(op_col is not None or _has_faq_id_column(raw_df.columns))
    op_values = raw_df[op_col].fillna("").astype(str).tolist() if op_col is not None else [""] * len(raw_df)

    clean = normalize_faq_columns(raw_df, drop_empty_required=False, assign_ids=False)
    if len(op_values) != len(clean):
        op_values = [""] * len(clean)
    clean.insert(0, FAQ_IMPORT_OPERATION_COLUMN, [_normalize_operation(v) for v in op_values])
    for col in ["question", "answer", "intent", "keywords", "category"]:
        if col in clean.columns:
            clean[col] = (
                clean[col]
                .fillna("")
                .astype(str)
                .map(_strip_excel_phonetic_artifacts)
                .str.replace("\u3000", " ", regex=False)
                .str.replace("\r\n", "\n", regex=False)
                .str.replace("\r", "\n", regex=False)
                .str.strip()
            )
    clean.attrs["has_operation_spec"] = has_operation_spec
    return clean


def _faq_question_key(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _allocate_next_faq_id(used_ids: set[str], start_no: int) -> tuple[str, int]:
    next_no = max(1, start_no)
    while True:
        candidate = f"FAQ-{next_no:06d}"
        next_no += 1
        if candidate not in used_ids:
            used_ids.add(candidate)
            return candidate, next_no


def _current_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _default_admin_name() -> str:
    return str(os.environ.get("ADMIN_DISPLAY_NAME") or "管理者").strip() or "管理者"


def _apply_audit_fields(record: dict, *, updated_at: str, updated_by: str, note: str | None = None) -> dict:
    record = dict(record)
    record["updated_at"] = updated_at
    record["updated_by"] = updated_by
    if note is not None:
        record["note"] = str(note or "").strip()
    return record


def apply_faq_upload_operations(
    current_df: pd.DataFrame,
    incoming_df: pd.DataFrame,
    *,
    updated_by: str | None = None,
    updated_at: str | None = None,
) -> tuple[pd.DataFrame, dict]:
    """+追加 / 空白更新 / -削除 の差分反映結果を作る。保存は呼び出し側で行う。

    高速化ポイント:
    - 以前はアップロード1行ごとに既存FAQ全件のID/質問マップを作り直していたため、
      FAQ件数×アップロード行数で急激に遅くなっていた。
    - 現在はID/質問マップを最初に1回だけ作り、追加/更新/削除のたびに差分更新する。
    - 既存機能（+追加、空白更新、-削除、更新日/更新者自動設定）は維持する。
    """
    actor = str(updated_by or _default_admin_name()).strip() or "管理者"
    stamp = str(updated_at or _current_timestamp()).strip() or _current_timestamp()
    current = normalize_faq_columns(current_df, drop_empty_required=True, assign_ids=True)
    raw_incoming = incoming_df.copy() if incoming_df is not None else pd.DataFrame()
    if FAQ_IMPORT_OPERATION_COLUMN not in raw_incoming.columns:
        raw_incoming.insert(0, FAQ_IMPORT_OPERATION_COLUMN, "")
    op_values = [_normalize_operation(v) for v in raw_incoming[FAQ_IMPORT_OPERATION_COLUMN].fillna("").astype(str).tolist()]
    incoming = normalize_faq_columns(raw_incoming, drop_empty_required=False, assign_ids=False)
    if len(op_values) != len(incoming):
        op_values = [""] * len(incoming)
    incoming.insert(0, FAQ_IMPORT_OPERATION_COLUMN, op_values)
    incoming = incoming[[FAQ_IMPORT_OPERATION_COLUMN] + FAQ_COLUMNS]

    records = current.to_dict("records")
    used_ids = {str(r.get("faq_id", "")).strip() for r in records if str(r.get("faq_id", "")).strip()}
    max_no = 0
    for fid0 in used_ids:
        no = _extract_faq_id_number(fid0)
        if no is not None:
            max_no = max(max_no, no)
    next_no = max_no + 1

    summary = {"added": 0, "updated": 0, "deleted": 0, "errors": [], "details": [], "updated_at": stamp, "updated_by": actor}

    def build_maps():
        id_map = {}
        q_map = {}
        for idx, rec in enumerate(records):
            fid0 = str(rec.get("faq_id", "")).strip()
            if fid0:
                id_map[fid0] = idx
            q0 = _faq_question_key(rec.get("question", ""))
            if q0 and q0 not in q_map:
                q_map[q0] = idx
        return id_map, q_map

    id_map, q_map = build_maps()
    deleted_indices: set[int] = set()

    def _compact_records_if_needed(force: bool = False) -> None:
        """削除済みNoneが増えたら配列を詰め直し、indexマップを再構築する。"""
        nonlocal records, id_map, q_map, deleted_indices
        if not deleted_indices:
            return
        if not force and len(deleted_indices) < 100:
            return
        records = [r for r in records if r is not None]
        id_map, q_map = build_maps()
        deleted_indices = set()

    def _set_record(idx: int, rec: dict) -> None:
        records[idx] = rec
        fid0 = str(rec.get("faq_id", "")).strip()
        q0 = _faq_question_key(rec.get("question", ""))
        if fid0:
            id_map[fid0] = idx
        if q0:
            q_map[q0] = idx

    def _append_record(rec: dict) -> None:
        idx = len(records)
        records.append(rec)
        fid0 = str(rec.get("faq_id", "")).strip()
        q0 = _faq_question_key(rec.get("question", ""))
        if fid0:
            id_map[fid0] = idx
        if q0:
            q_map[q0] = idx

    def _delete_record(idx: int) -> dict:
        rec = records[idx]
        records[idx] = None
        deleted_indices.add(idx)
        fid0 = str(rec.get("faq_id", "")).strip()
        q0 = _faq_question_key(rec.get("question", ""))
        if fid0 and id_map.get(fid0) == idx:
            id_map.pop(fid0, None)
        if q0 and q_map.get(q0) == idx:
            q_map.pop(q0, None)
        return rec

    def row_to_record(row, *, faq_id: str) -> dict | None:
        candidate = {col: row.get(col, "") for col in FAQ_COLUMNS}
        candidate["faq_id"] = faq_id
        if not str(candidate.get("enabled", "")).strip():
            candidate["enabled"] = "TRUE"
        clean = normalize_faq_columns(pd.DataFrame([candidate]), drop_empty_required=True, assign_ids=False)
        if len(clean) == 0:
            return None
        rec = clean.iloc[0].to_dict()
        return _apply_audit_fields(
            rec,
            updated_at=stamp,
            updated_by=actor,
            note=row.get("note", ""),
        )

    for pos, row in incoming.iterrows():
        excel_row = int(pos) + 2
        op = _normalize_operation(row.get(FAQ_IMPORT_OPERATION_COLUMN, ""))
        fid = str(row.get("faq_id", "")).strip()
        question = _faq_question_key(row.get("question", ""))

        if op == "+":
            if not question or not str(row.get("answer", "")).strip():
                summary["errors"].append(f"{excel_row}行目: 追加（+）は質問と回答が必須です。")
                continue

            original_fid = fid
            fid, next_no = _allocate_next_faq_id(used_ids, next_no)
            rec = row_to_record(row, faq_id=fid)
            if rec is None:
                summary["errors"].append(f"{excel_row}行目: 追加（+）の内容が不正です。")
                continue
            _append_record(rec)
            summary["added"] += 1
            detail = {"行": excel_row, "処理": "追加", "FAQ_ID": fid, "質問": rec.get("question", ""), "更新日": stamp, "更新者": actor, "備考": row.get("note", "")}
            if original_fid and original_fid != fid:
                detail["元FAQ_ID"] = original_fid
                auto_note = "+追加のため新しいFAQ_IDを自動採番"
                if str(detail.get("備考", "")).strip():
                    detail["備考"] = f"{detail['備考']} / {auto_note}"
                else:
                    detail["備考"] = auto_note
            summary["details"].append(detail)
            continue

        if op == "-":
            target_idx = id_map.get(fid) if fid else q_map.get(question)
            if target_idx is None or target_idx in deleted_indices or records[target_idx] is None:
                summary["errors"].append(f"{excel_row}行目: 削除（-）対象が見つかりません。FAQ_IDまたは既存と同じ質問を指定してください。")
                continue
            removed = _delete_record(target_idx)
            summary["deleted"] += 1
            summary["details"].append({"行": excel_row, "処理": "削除", "FAQ_ID": removed.get("faq_id", ""), "質問": removed.get("question", ""), "更新日": stamp, "更新者": actor, "備考": row.get("note", "")})
            _compact_records_if_needed()
            continue

        target_idx = id_map.get(fid) if fid else q_map.get(question)
        if target_idx is None or target_idx in deleted_indices or records[target_idx] is None:
            summary["errors"].append(f"{excel_row}行目: 更新対象が見つかりません。既存FAQ_IDを指定してください。新規追加する場合は操作に + を入力してください。")
            continue
        target_fid = str(records[target_idx].get("faq_id", "")).strip()
        old_question = _faq_question_key(records[target_idx].get("question", ""))
        rec = row_to_record(row, faq_id=target_fid)
        if rec is None:
            summary["errors"].append(f"{excel_row}行目: 更新は質問と回答が必須です。")
            continue
        if old_question and q_map.get(old_question) == target_idx:
            q_map.pop(old_question, None)
        _set_record(target_idx, rec)
        summary["updated"] += 1
        summary["details"].append({"行": excel_row, "処理": "更新", "FAQ_ID": target_fid, "質問": rec.get("question", ""), "更新日": stamp, "更新者": actor, "備考": row.get("note", "")})

    _compact_records_if_needed(force=True)
    result = normalize_faq_columns(pd.DataFrame(records), drop_empty_required=True, assign_ids=True)
    return result, summary

def append_faq_import_history(faq_path: Path, summary: dict, *, persist_callback=None) -> int:
    """FAQ差分反映の履歴を runtime_data/faq_import_history.csv に追記する。"""
    details = list((summary or {}).get("details") or [])
    if not details:
        return 0
    hist_path = faq_path.parent / "faq_import_history.csv"
    hist_path.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    for d in details:
        rows.append({
            "日時": d.get("更新日") or (summary or {}).get("updated_at") or _current_timestamp(),
            "更新者": d.get("更新者") or (summary or {}).get("updated_by") or _default_admin_name(),
            "操作": d.get("処理", ""),
            "FAQ_ID": d.get("FAQ_ID", ""),
            "質問": d.get("質問", ""),
            "結果": "成功",
            "備考": d.get("備考", ""),
        })
    new_df = pd.DataFrame(rows, columns=["日時", "更新者", "操作", "FAQ_ID", "質問", "結果", "備考"])
    if hist_path.exists():
        try:
            old_df = read_csv_flexible(hist_path)
            new_df = pd.concat([old_df, new_df], ignore_index=True)
        except Exception:
            pass
    new_df.to_csv(hist_path, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_ALL)
    if callable(persist_callback):
        try:
            persist_callback()
        except TypeError:
            persist_callback(hist_path)
    return len(rows)


def save_faq_csv_full(faq_path: Path, df: pd.DataFrame, persist_callback=None, *, already_normalized: bool = False) -> int:
    """FAQ全体を保存する。

    高速化ポイント:
    - 管理画面側で正規化済みの場合は normalize_faq_columns を二重実行しない。
    - SQLiteを正本にしつつ、CSVは互換用キャッシュとして同期する。
    - persist_callback は設定側で非同期化できるため、GitHub永続化ONでも画面を待たせにくい。
    """
    clean = df.copy() if already_normalized else normalize_faq_columns(df)
    faq_path.parent.mkdir(parents=True, exist_ok=True)
    if is_managed_faq_path(faq_path):
        save_faq_df_to_db(faq_path, clean)
        # 既存のダウンロード/外部永続化互換のためCSVも残す。
        # メモリ爆速版では、DB保存後のCSV書き出し/GitHub送信をバックグラウンド化し、
        # 画面操作と次回検索を待たせない。
        sync_csv = str(os.environ.get("HELP_DESK_FAQ_SYNC_CSV", "1")).strip().lower() not in {"0", "false", "no", "off"}
        async_csv = str(os.environ.get("HELP_DESK_FAQ_ASYNC_CSV_SYNC", "1")).strip().lower() not in {"0", "false", "no", "off"}

        def _sync_and_persist():
            if sync_csv:
                sync_faq_csv_cache(faq_path, clean.copy())
            if callable(persist_callback):
                try:
                    persist_callback()
                except Exception:
                    pass

        if async_csv:
            try:
                threading.Thread(target=_sync_and_persist, daemon=True).start()
            except Exception:
                _sync_and_persist()
        else:
            _sync_and_persist()
    else:
        clean.to_csv(faq_path, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_ALL)
        if callable(persist_callback):
            persist_callback()
    return len(clean)
