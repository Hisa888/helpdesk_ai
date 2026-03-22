from pathlib import Path
from datetime import datetime, timedelta
import base64
import csv
import io
import json
import os
import re
import threading
import zipfile

import pandas as pd
import requests
import streamlit as st
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

def read_csv_flexible(path: Path) -> pd.DataFrame:
    """CSVをできる限り失敗しないで読む（UTF-8/UTF-8-SIG/CP932等 + delimiter推定）。"""
    # まずは生bytesを読む
    raw = path.read_bytes()
    # 文字コード候補
    encs = ["utf-8", "utf-8-sig", "cp932", "shift_jis"]
    last_err = None
    text = None
    for enc in encs:
        try:
            text = raw.decode(enc)
            break
        except Exception as e:
            last_err = e
            continue
    if text is None:
        # 最終手段：latin1で無理やり
        text = raw.decode("latin1", errors="ignore")

    # delimiter推定（csv.Snifferが失敗する場合もあるのでガード）
    import csv as _csv
    sample = text[:5000]
    delim = ","
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        delim = dialect.delimiter
    except Exception:
        # 行にタブが多ければタブ、それ以外はカンマ
        if sample.count("\t") > sample.count(","):
            delim = "\t"

    # pandasで読む（on_bad_linesはpandas2系で有効）
    try:
        return pd.read_csv(io.StringIO(text), sep=delim, engine="python", on_bad_lines="skip")
    except Exception:
        try:
            return pd.read_csv(io.StringIO(text), sep=delim, engine="python")
        except Exception:
            # それでもダメなら空
            return pd.DataFrame()

def pick_question_column(cols) -> str | None:
    """質問カラム名の揺れを吸収"""
    cand = [
        "question", "質問", "問い合わせ", "問合せ", "query", "user_question",
        "content", "text"
    ]
    for c in cand:
        if c in cols:
            return c
    # 大文字小文字無視
    lower_map = {str(c).lower(): c for c in cols}
    for c in cand:
        if c.lower() in lower_map:
            return lower_map[c.lower()]
    return None

def extract_json_array(text: str) -> str | None:
    """LLM出力から最初のJSON配列（[...]）を抜き出す。"""
    if not text:
        return None
    s = str(text).strip()
    # ```json ... ``` を除去
    s = re.sub(r"^```(?:json)?\s*|\s*```$", "", s.strip(), flags=re.IGNORECASE)
    # 先頭から最初の [ ... ] を抽出（DOTALL）
    m = re.search(r"\[[\s\S]*\]", s)
    return m.group(0).strip() if m else None

def normalize_faq_columns(df: pd.DataFrame) -> pd.DataFrame:
    """列名揺れを吸収して question/answer/category に正規化する。"""
    if df is None or len(df) == 0:
        return pd.DataFrame(columns=["question", "answer", "category"])
    out = df.copy()
    rename_map = {}
    for c in out.columns:
        original = str(c).strip()
        key = original.lower()
        if original in ["質問", "問い合わせ", "問合せ"] or key in ["question", "query"]:
            rename_map[c] = "question"
        elif original in ["回答"] or key in ["answer", "answer_text", "reply"]:
            rename_map[c] = "answer"
        elif original in ["カテゴリ", "分類"] or key in ["category"]:
            rename_map[c] = "category"
    out = out.rename(columns=rename_map)
    for col in ["question", "answer", "category"]:
        if col not in out.columns:
            out[col] = ""
    out = out[["question", "answer", "category"]].copy()
    for col in ["question", "answer", "category"]:
        if col in out.columns:
            out[col] = out[col].fillna("").astype(str).str.replace("\n", " ").str.replace("\r", " ")

    out = out[(out["question"] != "") & (out["answer"] != "")].reset_index(drop=True)

    return out

def _xlsx_escape(text: str) -> str:
    return (str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', '&quot;'))

def _build_minimal_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "FAQ") -> bytes:
    from io import BytesIO
    rows = [list(df.columns)] + df.fillna("").astype(str).values.tolist()
    shared_strings = []
    sst_index = {}
    def get_sst_idx(value: str) -> int:
        value = str(value)
        if value not in sst_index:
            sst_index[value] = len(shared_strings)
            shared_strings.append(value)
        return sst_index[value]
    sheet_rows = []
    for r_idx, row in enumerate(rows, start=1):
        cells = []
        for c_idx, value in enumerate(row, start=1):
            col = ""
            n = c_idx
            while n:
                n, rem = divmod(n - 1, 26)
                col = chr(65 + rem) + col
            ref = f"{col}{r_idx}"
            s_idx = get_sst_idx(value)
            cells.append(f'<c r="{ref}" t="s"><v>{s_idx}</v></c>')
        sheet_rows.append(f'<row r="{r_idx}">' + "".join(cells) + '</row>')
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetData>' + ''.join(sheet_rows) + '</sheetData>'
        '</worksheet>'
    )
    sst_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(shared_strings)}" uniqueCount="{len(shared_strings)}">'
        + ''.join(f'<si><t xml:space="preserve">{_xlsx_escape(s)}</t></si>' for s in shared_strings) +
        '</sst>'
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets><sheet name="{_xlsx_escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets>'
        '</workbook>'
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        '</Relationships>'
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        '</Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        '</Types>'
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        '</styleSheet>'
    )
    core_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
                'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
                'xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
                '<dc:title>FAQ</dc:title><dc:creator>ChatGPT</dc:creator></cp:coreProperties>')
    app_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
               '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
               'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
               '<Application>Microsoft Excel</Application></Properties>')
    bio = BytesIO()
    with zipfile.ZipFile(bio, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', content_types)
        zf.writestr('_rels/.rels', root_rels)
        zf.writestr('docProps/core.xml', core_xml)
        zf.writestr('docProps/app.xml', app_xml)
        zf.writestr('xl/workbook.xml', workbook_xml)
        zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels)
        zf.writestr('xl/styles.xml', styles_xml)
        zf.writestr('xl/sharedStrings.xml', sst_xml)
        zf.writestr('xl/worksheets/sheet1.xml', sheet_xml)
    return bio.getvalue()

def faq_df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    export_df = normalize_faq_columns(df).rename(columns={"question": "質問", "answer": "回答", "category": "カテゴリ"})
    return _build_minimal_xlsx_bytes(export_df, sheet_name="FAQ")

def _read_xlsx_bytes(raw: bytes) -> pd.DataFrame:
    """XLSXを安全にDataFrame化する。
    1) pandas + openpyxl を優先
    2) openpyxl の values_only 読み
    3) 最後の手段で XML 手動解析
    """
    from io import BytesIO

    # 1) まずは pandas + openpyxl
    try:
        df = pd.read_excel(BytesIO(raw), engine="openpyxl")
        if df is not None:
            df.columns = [str(c).strip() for c in df.columns]
            return df
    except Exception:
        pass

    # 2) openpyxl の values_only 読み
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            max_cols = max(len(r) if r is not None else 0 for r in rows)
            norm_rows = []
            for r in rows:
                r = list(r or [])
                r += [""] * (max_cols - len(r))
                norm_rows.append(["" if v is None else str(v) for v in r])
            header = [str(x).strip() for x in norm_rows[0]]
            body = norm_rows[1:] if len(norm_rows) > 1 else []
            return pd.DataFrame(body, columns=header)
    except Exception:
        pass

    # 3) フォールバック: XML手動解析（ふりがな rPh は無視）
    import xml.etree.ElementTree as ET
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def _text_without_rph(parent):
        parts = []
        for child in list(parent):
            tag = child.tag.split("}")[-1]
            if tag == "rPh":
                continue
            if tag == "t":
                parts.append(child.text or "")
            else:
                for tnode in child.findall(".//a:t", ns):
                    parts.append(tnode.text or "")
        return "".join(parts)

    with zipfile.ZipFile(BytesIO(raw)) as zf:
        shared = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", ns):
                shared.append(_text_without_rph(si))

        sheet_path = "xl/worksheets/sheet1.xml"
        if "xl/workbook.xml" in zf.namelist() and "xl/_rels/workbook.xml.rels" in zf.namelist():
            wb = ET.fromstring(zf.read("xl/workbook.xml"))
            rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rid_to_target = {
                rel.attrib.get("Id"): rel.attrib.get("Target")
                for rel in rels.findall("pr:Relationship", ns)
            }
            first_sheet = wb.find(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/"
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
            )
            if first_sheet is not None:
                rid = first_sheet.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                target = rid_to_target.get(rid)
                if target:
                    if not target.startswith("worksheets/"):
                        target = target.split("xl/")[-1]
                    sheet_path = "xl/" + target

        sheet = ET.fromstring(zf.read(sheet_path))
        rows = []
        for row in sheet.findall("a:sheetData/a:row", ns):
            row_map = {}
            max_col = 0
            for c in row.findall("a:c", ns):
                ref = c.attrib.get("r", "")
                col_letters = "".join(ch for ch in ref if ch.isalpha())
                col_idx = 0
                for ch in col_letters:
                    col_idx = col_idx * 26 + (ord(ch.upper()) - 64)
                max_col = max(max_col, col_idx)

                t = c.attrib.get("t")
                value = ""
                if t == "s":
                    v = c.find("a:v", ns)
                    if v is not None and (v.text or "").isdigit():
                        si_idx = int(v.text)
                        value = shared[si_idx] if 0 <= si_idx < len(shared) else ""
                elif t == "inlineStr":
                    is_el = c.find("a:is", ns)
                    if is_el is not None:
                        value = _text_without_rph(is_el)
                else:
                    v = c.find("a:v", ns)
                    value = v.text if v is not None and v.text is not None else ""

                if col_idx > 0:
                    row_map[col_idx] = value

            if max_col:
                rows.append([row_map.get(i, "") for i in range(1, max_col + 1)])

    if not rows:
        return pd.DataFrame()
    max_cols = max(len(r) for r in rows)
    rows = [r + [""] * (max_cols - len(r)) for r in rows]
    header = [str(x).strip() for x in rows[0]]
    body = rows[1:] if len(rows) > 1 else []
    return pd.DataFrame(body, columns=header)

def _strip_excel_phonetic_artifacts(text: str) -> str:
    """Excel由来のふりがな混入っぽい末尾カタカナを保守的に除去する。"""
    s = "" if text is None else str(text)
    s = s.replace("\u3000", " ").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not s:
        return s

    # 句読点の直後にだけ付いた読み仮名を除去
    s = re.sub(r"([。．.!?！？]\s*)([ァ-ヶーｦ-ﾟ]{1,20})$", r"\1", s)

    # 漢字を含む語の末尾にだけ付いたカタカナ読みを除去
    m = re.match(r"^(.*[一-龥々〆ヵヶ])([ァ-ヶー]{2,20})$", s)
    if m:
        prefix, suffix = m.groups()
        if not re.search(r"[ァ-ヶー]$", prefix):
            s = prefix

    return s.strip()

def read_faq_uploaded_file(file_name: str, raw: bytes) -> pd.DataFrame:
    """FAQアップロード読込。
    まずCSV/XLSXを安全にDataFrame化し、列名を正規化する。
    文字化けやExcel由来のふりがな混入を抑えるため、読み込み後に文字列を明示的に整形する。
    """
    suffix = Path(file_name).suffix.lower()
    if suffix == '.csv':
        tmp = Path('/tmp/_faq_upload.csv')
        tmp.write_bytes(raw)
        df = read_csv_flexible(tmp)
    else:
        df = _read_xlsx_bytes(raw)

    df = normalize_faq_columns(df)
    for col in ["question", "answer", "category"]:
        if col in df.columns:
            df[col] = (
                df[col]
                .fillna("")
                .astype(str)
                .map(_strip_excel_phonetic_artifacts)
                .str.replace("\u3000", " ", regex=False)
                .str.replace("\r\n", "\n", regex=False)
                .str.replace("\r", "\n", regex=False)
                .str.strip()
            )
    return df

def save_faq_csv_full(faq_path: Path, df: pd.DataFrame) -> int:
    clean = normalize_faq_columns(df)
    faq_path.parent.mkdir(parents=True, exist_ok=True)
    clean.to_csv(faq_path, index=False, encoding='utf-8-sig', quoting=csv.QUOTE_ALL)
    if faq_path == FAQ_PATH:
        persist_faq_now()
    return len(clean)
