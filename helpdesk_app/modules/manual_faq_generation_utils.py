from __future__ import annotations

import io
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from helpdesk_app.modules.admin_faq_generation_utils import extract_json_array
from helpdesk_app.modules.document_rag_extractors import extract_sections_from_uploaded_file, normalize_doc_text

SUPPORTED_MANUAL_FAQ_EXTENSIONS = ("pdf", "docx", "xlsx", "xlsm", "txt", "md")
FAQ_COLUMNS = ["question", "answer", "intent", "keywords", "category", "answer_format", "source"]


def _normalize_generation_limit(n_items: Any) -> int:
    """FAQ生成件数の上限を正規化する。

    0以下は「上限なし」として扱う。
    これにより、Excelから直接抽出できるFAQ候補が20件で止まらない。
    """
    try:
        n = int(n_items)
    except Exception:
        return 0
    return max(0, n)


def _limit_reached(count: int, n_items: Any) -> bool:
    limit = _normalize_generation_limit(n_items)
    return limit > 0 and count >= limit


def _limit_df(df: pd.DataFrame, n_items: Any) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame(columns=FAQ_COLUMNS)
    limit = _normalize_generation_limit(n_items)
    out = df.reset_index(drop=True)
    if limit > 0:
        out = out.head(limit).reset_index(drop=True)
    return out


def _llm_target_count(n_items: Any, *, fallback: int = 50) -> int:
    """LLMへ依頼する件数。

    UIの0件指定は「無制限」だが、LLMには無限件数を依頼できないため、
    直接抽出できないPDF/Word/Textでは実用的な初期目安を使う。
    Excelのように直接FAQ化できるものは全件を返す。
    """
    limit = _normalize_generation_limit(n_items)
    return limit if limit > 0 else fallback

QUESTION_COL_CANDIDATES = {"question", "q", "質問", "問い合わせ", "問合せ", "faq_question", "faq質問", "質問文"}
ANSWER_COL_CANDIDATES = {"answer", "a", "回答", "答え", "対応", "対処", "faq_answer", "faq回答", "回答文"}
CATEGORY_COL_CANDIDATES = {"category", "カテゴリ", "カテゴリー", "分類", "項目", "section", "セクション", "大項目", "中項目", "種別", "対象"}
CONTENT_COL_CANDIDATES = {"content", "本文", "内容", "手順", "説明", "detail", "details", "manual", "マニュアル", "備考", "注意", "対応内容", "対処方法"}

_BOX_CHARS_RE = re.compile(r"[\u25A0-\u25FF\u2580-\u259F\uE000-\uF8FF]+")
_CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_SPACE_RE = re.compile(r"[ \t　]+")


def _basic_clean_text(text: Any) -> str:
    """セル・LLM出力・PDF抽出結果をFAQ向けに安全化する。"""
    if text is None:
        return ""
    s = str(text)
    if s.lower().strip() in {"nan", "none", "null"}:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\r", "\n")
    s = _CONTROL_RE.sub("", s)
    # PDF抽出やフォント化けで出る黒四角・私用領域文字を除去
    s = _BOX_CHARS_RE.sub("", s)
    s = s.replace("□", "").replace("■", "")
    s = normalize_doc_text(s)
    s = _SPACE_RE.sub(" ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def _clean_cell(value: Any) -> str:
    return _basic_clean_text(value)


def _normalize_header(value: Any) -> str:
    return _clean_cell(value).lower().replace(" ", "").replace("　", "").replace("_", "").replace("-", "")


def _find_col(columns: list[Any], candidates: set[str]) -> str | None:
    normalized_candidates = {_normalize_header(c) for c in candidates}
    for col in columns:
        if _normalize_header(col) in normalized_candidates:
            return col
    return None


def _safe_category(text: str) -> str:
    c = _clean_cell(text)
    if not c:
        return "マニュアル"
    # 長すぎるカテゴリは扱いにくいため短縮
    return c[:30]


def _auto_intent(question: str, category: str = "") -> str:
    q = _clean_cell(question).rstrip("？?")
    cat = _safe_category(category) if category else "マニュアル"
    if q:
        return f"{cat}について、{q}内容を解決したい"
    return f"{cat}について必要な手順を確認したい"


def _auto_keywords(question: str, category: str = "", content: str = "") -> str:
    text = f"{category} {question} {content}"
    keywords: list[str] = []
    synonym_sets = [
        ("VPN", ["つながらない", "接続できない", "在宅", "社外", "認証エラー", "ワンタイムパスワード"]),
        ("パスワード", ["ログインできない", "サインインできない", "忘れた", "再設定", "リセット", "アカウントロック"]),
        ("Outlook", ["メール", "受信できない", "送信できない", "見れない", "共有メールボックス"]),
        ("プリンタ", ["印刷できない", "プリンター", "紙詰まり", "出力できない"]),
        ("PC", ["パソコン", "起動しない", "重い", "遅い", "固まる"]),
        ("申請", ["依頼", "承認", "手続き", "アカウント発行", "権限"]),
    ]
    for trigger, words in synonym_sets:
        if trigger.lower() in text.lower() or any(w in text for w in words):
            keywords.extend([trigger, *words])
    for token in re.findall(r"[A-Za-z0-9]+|[ぁ-んァ-ヶ一-龥]{2,}", _clean_cell(text)):
        if token not in keywords and len(keywords) < 18:
            keywords.append(token)
    return ", ".join(dict.fromkeys(keywords))


def _answer_from_content(content: str) -> str:
    text = _clean_cell(content)
    if not text:
        return "資料に記載された内容を確認してください。"
    if len(text) <= 18 and not text.endswith("。"):
        return f"{text}。詳細は資料を確認してください。"
    return text


def _question_from_category_content(category: str, content: str) -> str:
    c = _clean_cell(category) or "この内容"
    body = _clean_cell(content)
    target = f"{c}\n{body}"
    if re.search(r"VPN|エラー\s*691|691", target, flags=re.I):
        return "VPNに接続できない場合はどうすればよいですか？"
    if "共有メール" in target or re.search(r"outlook", target, flags=re.I):
        return "Outlookで共有メールボックスを利用する方法を教えてください。"
    if "ロック" in target:
        return "アカウントがロックされた場合はどうすればよいですか？"
    if "パスワード" in target:
        return "パスワードを忘れた場合はどうすればよいですか？"
    if "重い" in target or "遅い" in target:
        return "PCが重い場合はどうすればよいですか？"
    if re.search(r"申請|依頼|手続|承認", target):
        return f"{c}の申請方法を教えてください。"
    return f"{c}について教えてください。"


def _rows_from_openpyxl(file_bytes: bytes) -> list[dict[str, Any]]:
    """Excelを全シート・全行セル単位で安全に読み取る。"""
    rows: list[dict[str, Any]] = []
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    known_headers = {
        *{_normalize_header(x) for x in QUESTION_COL_CANDIDATES},
        *{_normalize_header(x) for x in ANSWER_COL_CANDIDATES},
        *{_normalize_header(x) for x in CATEGORY_COL_CANDIDATES},
        *{_normalize_header(x) for x in CONTENT_COL_CANDIDATES},
    }

    for ws in wb.worksheets:
        matrix: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            values = [_clean_cell(v) for v in row]
            if any(values):
                matrix.append(values)
        if not matrix:
            continue

        max_cols = max(len(r) for r in matrix)
        matrix = [r + [""] * (max_cols - len(r)) for r in matrix]

        # 先頭10行からヘッダーらしい行を探す。見つからない場合も1行目をヘッダーとして扱う。
        header_idx = 0
        best_score = -1
        for i, row in enumerate(matrix[:10]):
            non_empty = [v for v in row if v]
            header_hits = sum(1 for v in non_empty if _normalize_header(v) in known_headers)
            # 既知ヘッダーを強く優先。列名っぽい短い語も少し加点。
            short_words = sum(1 for v in non_empty if 0 < len(v) <= 20)
            score = header_hits * 100 + short_words + len(non_empty)
            if score > best_score:
                best_score = score
                header_idx = i

        headers = [h if h else f"列{idx + 1}" for idx, h in enumerate(matrix[header_idx])]
        data_rows = matrix[header_idx + 1 :] if header_idx + 1 < len(matrix) else []

        # データ行がない場合は、表ではなく文章だけのシートとして扱う。
        if not data_rows:
            for r_no, row in enumerate(matrix, start=1):
                item = {f"列{i+1}": row[i] for i in range(len(row))}
                if any(_clean_cell(v) for v in item.values()):
                    rows.append({"__sheet__": ws.title, "__row__": str(r_no), **item})
            continue

        for r_no, row in enumerate(data_rows, start=header_idx + 2):
            item = {headers[i]: row[i] for i in range(len(headers))}
            if any(_clean_cell(v) for v in item.values()):
                rows.append({"__sheet__": ws.title, "__row__": str(r_no), **item})
    return rows


def _manual_text_from_rows(rows: list[dict[str, Any]], *, filename: str) -> tuple[str, list[dict[str, str]]]:
    blocks_by_sheet: dict[str, list[str]] = {}
    for row in rows:
        sheet = str(row.get("__sheet__", "sheet"))
        row_no = str(row.get("__row__", ""))
        parts: list[str] = []
        for col, val in row.items():
            if str(col).startswith("__"):
                continue
            text = _clean_cell(val)
            if text:
                parts.append(f"{col}: {text}")
        if parts:
            prefix = f"row {row_no}: " if row_no else ""
            blocks_by_sheet.setdefault(sheet, []).append(prefix + " / ".join(parts))

    sections: list[dict[str, str]] = []
    text_blocks: list[str] = []
    for sheet, lines in blocks_by_sheet.items():
        body = "\n".join(lines)
        block = f"[Excel資料] {filename} / sheet {sheet}\n{body}"
        text_blocks.append(block)
        sections.append({
            "source_name": filename,
            "source_type": "xlsx",
            "location": f"sheet {sheet}",
            "text": block,
        })
    return "\n\n".join(text_blocks), sections


def _row_to_content(row: dict[str, Any], *, skip_cols: set[str]) -> str:
    parts: list[str] = []
    for col, val in row.items():
        if str(col).startswith("__") or col in skip_cols:
            continue
        text = _clean_cell(val)
        if text:
            parts.append(f"{col}: {text}")
    return " / ".join(parts)


def extract_excel_auto(*, file_bytes: bytes, filename: str) -> tuple[pd.DataFrame, list[dict[str, str]], str]:
    """Excelを自動判定してFAQ候補とLLM用本文を作る。

    - question/answer型: そのままFAQ候補化
    - 項目/内容型: 行単位でFAQ候補化
    - 表形式: 全列を文章化してFAQ候補化
    """
    try:
        rows = _rows_from_openpyxl(file_bytes)
    except Exception:
        return pd.DataFrame(columns=FAQ_COLUMNS), [], ""

    text, sections = _manual_text_from_rows(rows, filename=filename)
    if not rows:
        return pd.DataFrame(columns=FAQ_COLUMNS), sections, text

    all_columns: list[Any] = []
    for row in rows:
        for c in row.keys():
            if not str(c).startswith("__") and c not in all_columns:
                all_columns.append(c)

    q_col = _find_col(all_columns, QUESTION_COL_CANDIDATES)
    a_col = _find_col(all_columns, ANSWER_COL_CANDIDATES)
    cat_col = _find_col(all_columns, CATEGORY_COL_CANDIDATES)
    content_col = _find_col(all_columns, CONTENT_COL_CANDIDATES)

    direct_rows: list[dict[str, str]] = []
    for row in rows:
        sheet = _clean_cell(row.get("__sheet__", "sheet"))
        row_no = _clean_cell(row.get("__row__", ""))
        source = f"{filename} / {sheet}" + (f" row {row_no}" if row_no else "")

        if q_col and a_col:
            question = _clean_cell(row.get(q_col, ""))
            answer = _clean_cell(row.get(a_col, ""))
            if question and answer:
                direct_rows.append({
                    "question": question,
                    "answer": _answer_from_content(answer),
                    "intent": _auto_intent(question, _safe_category(row.get(cat_col, "Excel")) if cat_col else "Excel"),
                    "keywords": _auto_keywords(question, _safe_category(row.get(cat_col, "Excel")) if cat_col else "Excel", answer),
                    "category": _safe_category(row.get(cat_col, "Excel")) if cat_col else "Excel",
                    "answer_format": "markdown",
                    "source": source,
                })
            continue

        category = _safe_category(row.get(cat_col, "Excel")) if cat_col else "Excel"
        content = _clean_cell(row.get(content_col, "")) if content_col else ""
        if not content:
            # どの列名でも、行の全内容をFAQ化できるようにする。
            skip = {cat_col} if cat_col else set()
            content = _row_to_content(row, skip_cols=skip)

        if content:
            direct_rows.append({
                "question": _question_from_category_content(category, content),
                "answer": _answer_from_content(content),
                "intent": _auto_intent(_question_from_category_content(category, content), category or "Excel"),
                "keywords": _auto_keywords(_question_from_category_content(category, content), category or "Excel", content),
                "category": category or "Excel",
                "answer_format": "markdown",
                "source": source,
            })

    return normalize_manual_faq_df(pd.DataFrame(direct_rows, columns=FAQ_COLUMNS)), sections, text


def collect_manual_source_text(*, uploaded_files: list[Any] | None = None, wiki_text: str = "", max_chars: int = 22000) -> tuple[str, list[dict[str, str]], pd.DataFrame]:
    """PDF/Word/Excel/TXT/MD/Wikiを自動読込し、LLM投入用本文と直接FAQ候補を返す。"""
    uploaded_files = uploaded_files or []
    sections: list[dict[str, str]] = []
    direct_faq_frames: list[pd.DataFrame] = []

    for uf in uploaded_files:
        filename = str(getattr(uf, "name", "document")).strip() or "document"
        ext = Path(filename).suffix.lower().lstrip(".")
        try:
            file_bytes = uf.getvalue() if hasattr(uf, "getvalue") else uf.read()
        except Exception:
            file_bytes = b""
        if not file_bytes:
            continue

        if ext in {"xlsx", "xlsm"}:
            direct_df, xlsx_sections, _xlsx_text = extract_excel_auto(file_bytes=file_bytes, filename=filename)
            if len(direct_df) > 0:
                direct_faq_frames.append(direct_df)
            sections.extend(xlsx_sections)
            continue

        # PDF / Word / txt / md は既存抽出器を使う。UploadedFileのポインタ位置対策でseekする。
        try:
            if hasattr(uf, "seek"):
                uf.seek(0)
            extracted = extract_sections_from_uploaded_file(uf)
        except Exception:
            extracted = []

        for sec in extracted:
            body = _basic_clean_text(sec.get("text", ""))
            if body:
                sec["text"] = body
                sections.append(sec)

    wiki_clean = _basic_clean_text(wiki_text or "")
    if wiki_clean:
        sections.append({
            "source_name": "wiki_input",
            "source_type": "wiki",
            "location": "pasted text",
            "text": wiki_clean,
        })

    blocks: list[str] = []
    total = 0
    for sec in sections:
        body = _basic_clean_text(sec.get("text", ""))
        if not body:
            continue
        block = (
            f"[資料名] {sec.get('source_name', 'document')}\n"
            f"[種類] {sec.get('source_type', 'text')}\n"
            f"[場所] {sec.get('location', 'document')}\n"
            f"[本文]\n{body}"
        ).strip()
        if total + len(block) > max_chars:
            remain = max_chars - total
            if remain > 300:
                blocks.append(block[:remain])
            break
        blocks.append(block)
        total += len(block)

    direct_df = pd.concat(direct_faq_frames, ignore_index=True) if direct_faq_frames else pd.DataFrame(columns=FAQ_COLUMNS)
    direct_df = normalize_manual_faq_df(direct_df)
    return "\n\n".join(blocks), sections, direct_df


def normalize_manual_faq_df(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame(columns=FAQ_COLUMNS)
    out = df.copy().fillna("")
    for col in FAQ_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    out = out[FAQ_COLUMNS]
    for col in FAQ_COLUMNS:
        out[col] = out[col].astype(str).map(_basic_clean_text)
    out = out[(out["question"].str.len() > 0) & (out["answer"].str.len() > 0)]
    out["answer_format"] = out["answer_format"].replace("", "markdown")
    # 同じ質問は1件にまとめる。answerの違いだけで重複表示されるのを防ぐ。
    out = out.drop_duplicates(subset=["question"]).reset_index(drop=True)
    return out


def _make_question_from_text(category: str, text: str) -> str:
    """PDF/Word/TXTの本文から、社員が聞きそうな質問文を作る。"""
    return _question_from_category_content(category or "マニュアル", text)


def _extract_title_from_block(block: str) -> str:
    for line in (block or "").split("\n"):
        line = _basic_clean_text(line)
        if not line:
            continue
        if line.startswith("[資料名]") or line.startswith("[種類]") or line.startswith("[場所]") or line.startswith("[本文]"):
            continue
        # 見出し・短い行をカテゴリ候補にする
        if len(line) <= 40:
            return line
        return line[:30]
    return "マニュアル"


def _split_source_blocks(source_text: str) -> list[tuple[str, str]]:
    """資料ブロック単位に分解する。PDFはページ、Wordは見出し/表ごとに入りやすい。"""
    clean = _basic_clean_text(source_text or "")
    if not clean:
        return []
    raw_blocks = [b.strip() for b in re.split(r"\n\n(?=\[資料名\])", clean) if b.strip()]
    blocks: list[tuple[str, str]] = []
    for b in raw_blocks:
        location = "マニュアル"
        body = b
        m = re.search(r"\[場所\]\s*(.+)", b)
        if m:
            location = _basic_clean_text(m.group(1)) or "マニュアル"
        m_body = re.search(r"\[本文\]\s*(.*)", b, flags=re.S)
        if m_body:
            body = _basic_clean_text(m_body.group(1))
        if body:
            blocks.append((location, body))
    if not blocks:
        blocks.append(("マニュアル", clean))
    return blocks


def _paragraph_faq_candidates(source_text: str, n_items: int = 8) -> pd.DataFrame:
    """LLMが使えない場合でもPDF/Word/TXT本文からセクション単位のFAQを作る。"""
    rows: list[dict[str, str]] = []
    for location, body in _split_source_blocks(source_text):
        # 見出しごとに分割。Word/PDFの章立てにもある程度対応。
        parts = [p.strip() for p in re.split(r"\n{2,}|(?=^#{1,3}\s)|(?=^■)|(?=^【.+?】)", body, flags=re.M) if p.strip()]
        if not parts:
            parts = [body]
        for part in parts:
            part = _basic_clean_text(part)
            if len(part) < 8:
                continue
            title = _extract_title_from_block(part)
            q = _make_question_from_text(title or location, part)
            rows.append({
                "question": q,
                "answer": _answer_from_content(part),
                "intent": _auto_intent(q, _safe_category(title or location)),
                "keywords": _auto_keywords(q, _safe_category(title or location), part),
                "category": _safe_category(title or location),
                "answer_format": "markdown",
                "source": location,
            })
            if _limit_reached(len(rows), n_items):
                break
        if _limit_reached(len(rows), n_items):
            break
    return _limit_df(normalize_manual_faq_df(pd.DataFrame(rows, columns=FAQ_COLUMNS)), n_items)


def generate_heuristic_faq_candidates(*, source_text: str, n_items: int = 8) -> pd.DataFrame:
    """LLMが使えない/失敗した場合の簡易FAQ生成。

    PDF/Word/TXT/MDでは、まずキーワードに基づく定番FAQを作り、足りない分を
    セクション・段落単位のFAQで補完する。
    n_items が 0 の場合は、20件上限をかけずに抽出できる範囲を返す。
    """
    clean = _basic_clean_text(source_text or "")
    rows: list[dict[str, str]] = []
    if not clean:
        return pd.DataFrame(columns=FAQ_COLUMNS)

    patterns = [
        ("VPN", r"VPN|エラー\s*691|691|接続", "VPNに接続できない場合はどうすればよいですか？"),
        ("Outlook", r"Outlook|共有メール|メールボックス", "Outlookで共有メールボックスを利用する方法を教えてください。"),
        ("認証", r"アカウント.*ロック|ロック", "アカウントがロックされた場合はどうすればよいですか？"),
        ("認証", r"パスワード.*忘|再発行|初期化", "パスワードを忘れた場合はどうすればよいですか？"),
        ("PC", r"PC.*重い|重い|遅い|不要なアプリ|再起動", "PCが重い場合はどうすればよいですか？"),
        ("申請", r"申請|依頼|承認|手続", "申請や依頼を行う場合はどうすればよいですか？"),
    ]
    sentences = [s.strip() for s in re.split(r"(?<=[。！？\?])\s*|\n+", clean) if s.strip()]
    for category, pat, question in patterns:
        hits = [s for s in sentences if re.search(pat, s, flags=re.I)]
        if hits:
            rows.append({
                "question": question,
                "answer": _answer_from_content(" ".join(hits[:4])),
                "intent": _auto_intent(question, category),
                "keywords": _auto_keywords(question, category, " ".join(hits[:4])),
                "category": category,
                "answer_format": "markdown",
                "source": "自動抽出",
            })
        if _limit_reached(len(rows), n_items):
            break

    keyword_df = normalize_manual_faq_df(pd.DataFrame(rows, columns=FAQ_COLUMNS))
    if _limit_reached(len(keyword_df), n_items):
        return _limit_df(keyword_df, n_items)

    paragraph_df = _paragraph_faq_candidates(clean, n_items=n_items)
    frames = [df for df in [keyword_df, paragraph_df] if isinstance(df, pd.DataFrame) and len(df) > 0]
    if frames:
        return _limit_df(normalize_manual_faq_df(pd.concat(frames, ignore_index=True)), n_items)

    return pd.DataFrame(columns=FAQ_COLUMNS)


def generate_manual_faq_candidates(*, source_text: str, n_items: int, llm_chat, direct_candidates: pd.DataFrame | None = None) -> pd.DataFrame:
    clean = _basic_clean_text(source_text or "")
    direct_df = normalize_manual_faq_df(direct_candidates)
    limit = _normalize_generation_limit(n_items)

    # Excelなどから question/answer を直接抽出できる場合は、それを最優先にする。
    # n_items=0 の場合は全件返す。これで「20件しかFAQ化できない」問題を解消する。
    if len(direct_df) > 0 and (limit <= 0 or len(direct_df) >= limit):
        return _limit_df(direct_df, n_items)

    llm_df = pd.DataFrame(columns=FAQ_COLUMNS)
    if clean:
        target_count = _llm_target_count(n_items, fallback=50)
        target_text = f"{target_count} 件" if limit > 0 else "可能な範囲でできるだけ多く"
        prompt = f"""あなたは社内情シス向けFAQの作成者です。
以下のマニュアル・資料本文を読み、社員向けに使いやすいFAQ案を {target_text} 作成してください。

【目的】
- ドキュメント読み込み → 文章化 → 内容理解 → Q&A生成 を行う
- 社員が問い合わせ前に自己解決できるFAQにする
- 管理者が確認してFAQへ反映しやすい形にする

【重要ルール】
- 日本語のみ
- 出力は必ずJSON配列のみ
- 各要素は question / answer / intent / keywords / category / source を持つ
- question は社員が実際に入力しそうな自然な質問にする
- answer は本文に基づいて短く結論→手順の順で書く。必要なら箇条書きにする
- intent は質問の意味・目的を1文で書く
- keywords は言い換え・表記ゆれをカンマ区切りで書く
- Excel表の場合は、列名と値の関係を読み取ってFAQにする
- マニュアル本文にない内容を断定しない
- source は資料名、シート名、ページ等を短く書く
- code block は使わない

【JSON形式】
[
  {{"question":"VPNに接続できない場合はどうすればよいですか？", "answer":"...", "intent":"社外からVPNへ接続できない問題を解決したい", "keywords":"VPN, つながらない, 接続できない, 在宅, 認証エラー", "category":"VPN", "source":"vpn_manual.pdf page 1"}},
  {{"question":"共有メールボックスを追加する方法を教えてください。", "answer":"...", "intent":"Outlookで共有メールボックスを追加したい", "keywords":"Outlook, 共有メール, メールボックス, 追加", "category":"Outlook", "source":"outlook.docx document"}}
]

【資料本文】
{clean}
"""
        try:
            out = llm_chat([
                {"role": "system", "content": "あなたは社内情シスFAQの作成者です。出力はJSON配列のみです。"},
                {"role": "user", "content": prompt},
            ])
            out_text = out if isinstance(out, str) else str(out)
            json_text = extract_json_array(out_text) or out_text.strip()
            data = json.loads(json_text)
            if isinstance(data, list):
                rows: list[dict[str, str]] = []
                for item in data:
                    if not isinstance(item, dict):
                        continue
                    category = _clean_cell(item.get("category", "")) or "マニュアル"
                    question = _clean_cell(item.get("question", ""))
                    answer = _clean_cell(item.get("answer", ""))
                    intent = _clean_cell(item.get("intent", "")) or _auto_intent(question, category)
                    keywords = _clean_cell(item.get("keywords", "")) or _auto_keywords(question, category, answer)
                    source = _clean_cell(item.get("source", "")) or "自動生成"
                    if question and answer:
                        rows.append({
                            "question": question,
                            "answer": answer,
                            "intent": intent,
                            "keywords": keywords,
                            "category": category,
                            "answer_format": "markdown",
                            "source": source,
                        })
                llm_df = normalize_manual_faq_df(pd.DataFrame(rows, columns=FAQ_COLUMNS))
        except Exception:
            llm_df = pd.DataFrame(columns=FAQ_COLUMNS)

    if len(llm_df) == 0 and clean:
        llm_df = generate_heuristic_faq_candidates(source_text=clean, n_items=n_items)

    # 直接抽出FAQを優先し、足りない分をLLM/簡易生成で補完する。
    frames = [df for df in [direct_df, llm_df] if isinstance(df, pd.DataFrame) and len(df) > 0]
    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=FAQ_COLUMNS)
    combined = normalize_manual_faq_df(combined)
    return _limit_df(combined, n_items)
