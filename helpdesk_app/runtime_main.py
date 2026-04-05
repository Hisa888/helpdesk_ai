def run_app():
    import streamlit as st
    from pathlib import Path
    import pandas as pd
    import io
    import re
    import json
    import requests
    from datetime import datetime, timedelta
    import streamlit.components.v1 as components
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    try:
        from sentence_transformers import SentenceTransformer
        SENTENCE_TRANSFORMERS_AVAILABLE = True
    except Exception:
        SentenceTransformer = None
        SENTENCE_TRANSFORMERS_AVAILABLE = False

    import os
    import re
    import uuid
    import csv
    import io
    import zipfile
    import base64
    import threading

    # ===== CSV読み込みを頑丈にする（文字コード/区切り/カラム揺れ対策）=====
    def read_csv_flexible(path: Path) -> pd.DataFrame:
        """CSVをできる限り失敗しないで読む（UTF-8/UTF-8-SIG/CP932等 + delimiter推定）。"""
        # まずは生bytesを読む
        raw = path.read_bytes()
        # 文字コード候補
        encs = ["utf-8", "utf-8-sig", "cp932", "shift_jis"]
        last_err = None
        text = None
        for enc in encs:
            try:
                text = raw.decode(enc)
                break
            except Exception as e:
                last_err = e
                continue
        if text is None:
            # 最終手段：latin1で無理やり
            text = raw.decode("latin1", errors="ignore")

        # delimiter推定（csv.Snifferが失敗する場合もあるのでガード）
        import csv as _csv
        sample = text[:5000]
        delim = ","
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
            delim = dialect.delimiter
        except Exception:
            # 行にタブが多ければタブ、それ以外はカンマ
            if sample.count("\t") > sample.count(","):
                delim = "\t"

        # pandasで読む（on_bad_linesはpandas2系で有効）
        try:
            return pd.read_csv(io.StringIO(text), sep=delim, engine="python", on_bad_lines="skip")
        except Exception:
            try:
                return pd.read_csv(io.StringIO(text), sep=delim, engine="python")
            except Exception:
                # それでもダメなら空
                return pd.DataFrame()

    def pick_question_column(cols) -> str | None:
        """質問カラム名の揺れを吸収"""
        cand = [
            "question", "質問", "問い合わせ", "問合せ", "query", "user_question",
            "content", "text"
        ]
        for c in cand:
            if c in cols:
                return c
        # 大文字小文字無視
        lower_map = {str(c).lower(): c for c in cols}
        for c in cand:
            if c.lower() in lower_map:
                return lower_map[c.lower()]
        return None

    def extract_json_array(text: str) -> str | None:
        """LLM出力から最初のJSON配列（[...]）を抜き出す。"""
        if not text:
            return None
        s = str(text).strip()
        # ```json ... ``` を除去
        s = re.sub(r"^```(?:json)?\s*|\s*```$", "", s.strip(), flags=re.IGNORECASE)
        # 先頭から最初の [ ... ] を抽出（DOTALL）
        m = re.search(r"\[[\s\S]*\]", s)
        return m.group(0).strip() if m else None



    def normalize_faq_columns(df: pd.DataFrame) -> pd.DataFrame:
        """列名揺れを吸収して question/answer/category に正規化する。"""
        if df is None or len(df) == 0:
            return pd.DataFrame(columns=["question", "answer", "category"])
        out = df.copy()
        rename_map = {}
        for c in out.columns:
            original = str(c).strip()
            key = original.lower()
            if original in ["質問", "問い合わせ", "問合せ"] or key in ["question", "query"]:
                rename_map[c] = "question"
            elif original in ["回答"] or key in ["answer", "answer_text", "reply"]:
                rename_map[c] = "answer"
            elif original in ["カテゴリ", "分類"] or key in ["category"]:
                rename_map[c] = "category"
        out = out.rename(columns=rename_map)
        for col in ["question", "answer", "category"]:
            if col not in out.columns:
                out[col] = ""
        out = out[["question", "answer", "category"]].copy()
        for col in ["question", "answer", "category"]:
            if col in out.columns:
                out[col] = out[col].fillna("").astype(str).str.replace("\n", " ").str.replace("\r", " ")

        out = out[(out["question"] != "") & (out["answer"] != "")].reset_index(drop=True)

        return out

    def _xlsx_escape(text: str) -> str:
        return (str(text)
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace('"', '&quot;'))


    def _build_minimal_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "FAQ") -> bytes:
        from io import BytesIO
        rows = [list(df.columns)] + df.fillna("").astype(str).values.tolist()
        shared_strings = []
        sst_index = {}
        def get_sst_idx(value: str) -> int:
            value = str(value)
            if value not in sst_index:
                sst_index[value] = len(shared_strings)
                shared_strings.append(value)
            return sst_index[value]
        sheet_rows = []
        for r_idx, row in enumerate(rows, start=1):
            cells = []
            for c_idx, value in enumerate(row, start=1):
                col = ""
                n = c_idx
                while n:
                    n, rem = divmod(n - 1, 26)
                    col = chr(65 + rem) + col
                ref = f"{col}{r_idx}"
                s_idx = get_sst_idx(value)
                cells.append(f'<c r="{ref}" t="s"><v>{s_idx}</v></c>')
            sheet_rows.append(f'<row r="{r_idx}">' + "".join(cells) + '</row>')
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData>' + ''.join(sheet_rows) + '</sheetData>'
            '</worksheet>'
        )
        sst_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(shared_strings)}" uniqueCount="{len(shared_strings)}">'
            + ''.join(f'<si><t xml:space="preserve">{_xlsx_escape(s)}</t></si>' for s in shared_strings) +
            '</sst>'
        )
        workbook_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<sheets><sheet name="{_xlsx_escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets>'
            '</workbook>'
        )
        workbook_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
            '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
            '</Relationships>'
        )
        root_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
            '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
            '</Relationships>'
        )
        content_types = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
            '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
            '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
            '</Types>'
        )
        styles_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
            '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
            '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
            '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
            '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
            '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
            '</styleSheet>'
        )
        core_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
                    'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
                    'xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
                    '<dc:title>FAQ</dc:title><dc:creator>ChatGPT</dc:creator></cp:coreProperties>')
        app_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
                   'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
                   '<Application>Microsoft Excel</Application></Properties>')
        bio = BytesIO()
        with zipfile.ZipFile(bio, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('[Content_Types].xml', content_types)
            zf.writestr('_rels/.rels', root_rels)
            zf.writestr('docProps/core.xml', core_xml)
            zf.writestr('docProps/app.xml', app_xml)
            zf.writestr('xl/workbook.xml', workbook_xml)
            zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels)
            zf.writestr('xl/styles.xml', styles_xml)
            zf.writestr('xl/sharedStrings.xml', sst_xml)
            zf.writestr('xl/worksheets/sheet1.xml', sheet_xml)
        return bio.getvalue()


    def faq_df_to_excel_bytes(df: pd.DataFrame) -> bytes:
        export_df = normalize_faq_columns(df).rename(columns={"question": "質問", "answer": "回答", "category": "カテゴリ"})
        return _build_minimal_xlsx_bytes(export_df, sheet_name="FAQ")


    def _read_xlsx_bytes(raw: bytes) -> pd.DataFrame:
        """XLSXを安全にDataFrame化する。
        1) pandas + openpyxl を優先
        2) openpyxl の values_only 読み
        3) 最後の手段で XML 手動解析
        """
        from io import BytesIO

        # 1) まずは pandas + openpyxl
        try:
            df = pd.read_excel(BytesIO(raw), engine="openpyxl")
            if df is not None:
                df.columns = [str(c).strip() for c in df.columns]
                return df
        except Exception:
            pass

        # 2) openpyxl の values_only 読み
        try:
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                max_cols = max(len(r) if r is not None else 0 for r in rows)
                norm_rows = []
                for r in rows:
                    r = list(r or [])
                    r += [""] * (max_cols - len(r))
                    norm_rows.append(["" if v is None else str(v) for v in r])
                header = [str(x).strip() for x in norm_rows[0]]
                body = norm_rows[1:] if len(norm_rows) > 1 else []
                return pd.DataFrame(body, columns=header)
        except Exception:
            pass

        # 3) フォールバック: XML手動解析（ふりがな rPh は無視）
        import xml.etree.ElementTree as ET
        ns = {
            "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
        }

        def _text_without_rph(parent):
            parts = []
            for child in list(parent):
                tag = child.tag.split("}")[-1]
                if tag == "rPh":
                    continue
                if tag == "t":
                    parts.append(child.text or "")
                else:
                    for tnode in child.findall(".//a:t", ns):
                        parts.append(tnode.text or "")
            return "".join(parts)

        with zipfile.ZipFile(BytesIO(raw)) as zf:
            shared = []
            if "xl/sharedStrings.xml" in zf.namelist():
                root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for si in root.findall("a:si", ns):
                    shared.append(_text_without_rph(si))

            sheet_path = "xl/worksheets/sheet1.xml"
            if "xl/workbook.xml" in zf.namelist() and "xl/_rels/workbook.xml.rels" in zf.namelist():
                wb = ET.fromstring(zf.read("xl/workbook.xml"))
                rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                rid_to_target = {
                    rel.attrib.get("Id"): rel.attrib.get("Target")
                    for rel in rels.findall("pr:Relationship", ns)
                }
                first_sheet = wb.find(
                    "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/"
                    "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
                )
                if first_sheet is not None:
                    rid = first_sheet.attrib.get(
                        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                    )
                    target = rid_to_target.get(rid)
                    if target:
                        if not target.startswith("worksheets/"):
                            target = target.split("xl/")[-1]
                        sheet_path = "xl/" + target

            sheet = ET.fromstring(zf.read(sheet_path))
            rows = []
            for row in sheet.findall("a:sheetData/a:row", ns):
                row_map = {}
                max_col = 0
                for c in row.findall("a:c", ns):
                    ref = c.attrib.get("r", "")
                    col_letters = "".join(ch for ch in ref if ch.isalpha())
                    col_idx = 0
                    for ch in col_letters:
                        col_idx = col_idx * 26 + (ord(ch.upper()) - 64)
                    max_col = max(max_col, col_idx)

                    t = c.attrib.get("t")
                    value = ""
                    if t == "s":
                        v = c.find("a:v", ns)
                        if v is not None and (v.text or "").isdigit():
                            si_idx = int(v.text)
                            value = shared[si_idx] if 0 <= si_idx < len(shared) else ""
                    elif t == "inlineStr":
                        is_el = c.find("a:is", ns)
                        if is_el is not None:
                            value = _text_without_rph(is_el)
                    else:
                        v = c.find("a:v", ns)
                        value = v.text if v is not None and v.text is not None else ""

                    if col_idx > 0:
                        row_map[col_idx] = value

                if max_col:
                    rows.append([row_map.get(i, "") for i in range(1, max_col + 1)])

        if not rows:
            return pd.DataFrame()
        max_cols = max(len(r) for r in rows)
        rows = [r + [""] * (max_cols - len(r)) for r in rows]
        header = [str(x).strip() for x in rows[0]]
        body = rows[1:] if len(rows) > 1 else []
        return pd.DataFrame(body, columns=header)


    def _strip_excel_phonetic_artifacts(text: str) -> str:
        """Excel由来のふりがな混入っぽい末尾カタカナを保守的に除去する。"""
        s = "" if text is None else str(text)
        s = s.replace("\u3000", " ").replace("\r\n", "\n").replace("\r", "\n").strip()
        if not s:
            return s

        # 句読点の直後にだけ付いた読み仮名を除去
        s = re.sub(r"([。．.!?！？]\s*)([ァ-ヶーｦ-ﾟ]{1,20})$", r"\1", s)

        # 漢字を含む語の末尾にだけ付いたカタカナ読みを除去
        m = re.match(r"^(.*[一-龥々〆ヵヶ])([ァ-ヶー]{2,20})$", s)
        if m:
            prefix, suffix = m.groups()
            if not re.search(r"[ァ-ヶー]$", prefix):
                s = prefix

        return s.strip()


    def read_faq_uploaded_file(file_name: str, raw: bytes) -> pd.DataFrame:
        """FAQアップロード読込。
        まずCSV/XLSXを安全にDataFrame化し、列名を正規化する。
        文字化けやExcel由来のふりがな混入を抑えるため、読み込み後に文字列を明示的に整形する。
        """
        suffix = Path(file_name).suffix.lower()
        if suffix == '.csv':
            tmp = Path('/tmp/_faq_upload.csv')
            tmp.write_bytes(raw)
            df = read_csv_flexible(tmp)
        else:
            df = _read_xlsx_bytes(raw)

        df = normalize_faq_columns(df)
        for col in ["question", "answer", "category"]:
            if col in df.columns:
                df[col] = (
                    df[col]
                    .fillna("")
                    .astype(str)
                    .map(_strip_excel_phonetic_artifacts)
                    .str.replace("\u3000", " ", regex=False)
                    .str.replace("\r\n", "\n", regex=False)
                    .str.replace("\r", "\n", regex=False)
                    .str.strip()
                )
        return df


    def save_faq_csv_full(faq_path: Path, df: pd.DataFrame) -> int:
        clean = normalize_faq_columns(df)
        faq_path.parent.mkdir(parents=True, exist_ok=True)
        clean.to_csv(faq_path, index=False, encoding='utf-8-sig', quoting=csv.QUOTE_ALL)
        if faq_path == FAQ_PATH:
            persist_faq_now()
        return len(clean)

    from helpdesk_app.modules.pdf_runtime import (
        REPORTLAB_AVAILABLE,
        generate_effect_report_pdf,
        generate_ops_manual_pdf,
        generate_sales_proposal_pdf,
        generate_sales_proposal_pdf_v25,
    )

    from helpdesk_app.modules.settings_and_persistence import create_runtime_context
    from helpdesk_app.modules.ui_helpers import render_match_bar
    from helpdesk_app.admin_menu_complete import render_admin_complete_tools
    from helpdesk_app.modules.main_view_runtime import (
        ensure_admin_session_state,
        render_admin_login_sidebar,
        render_admin_tools_if_logged_in,
        render_public_sidebar,
        render_sales_kpi_sections,
    )
    from helpdesk_app.modules.faq_index_runtime import create_faq_index_runtime
    from helpdesk_app.modules.search_runtime import create_search_runtime

    settings_ctx = create_runtime_context(
        st=st,
        requests=requests,
        base_llm_chat=lambda messages: "",
        root_dir=Path("."),
    )

    FAQ_PATH = settings_ctx.FAQ_PATH
    LOG_DIR = settings_ctx.LOG_DIR
    COMPANY_NAME = settings_ctx.COMPANY_NAME
    LOGO_PATH = settings_ctx.LOGO_PATH
    CONTACT_URL = settings_ctx.CONTACT_URL
    CONTACT_EMAIL = settings_ctx.CONTACT_EMAIL
    LLM_SETTINGS = settings_ctx.LLM_SETTINGS
    SEARCH_SETTINGS = settings_ctx.SEARCH_SETTINGS
    DEFAULT_SEARCH_THRESHOLD = settings_ctx.DEFAULT_SEARCH_THRESHOLD
    DEFAULT_SUGGEST_THRESHOLD = settings_ctx.DEFAULT_SUGGEST_THRESHOLD
    default_ui_theme_settings = settings_ctx.default_ui_theme_settings
    default_ui_layout_settings = settings_ctx.default_ui_layout_settings
    sanitize_ui_theme_settings = settings_ctx.sanitize_ui_theme_settings
    sanitize_ui_layout_settings = settings_ctx.sanitize_ui_layout_settings
    default_llm_settings = settings_ctx.default_llm_settings
    sanitize_llm_settings = settings_ctx.sanitize_llm_settings
    current_llm_settings = settings_ctx.current_llm_settings
    save_llm_settings = settings_ctx.save_llm_settings
    default_search_settings = settings_ctx.default_search_settings
    _sanitize_search_settings = settings_ctx._sanitize_search_settings
    current_search_settings = settings_ctx.current_search_settings
    save_search_settings = settings_ctx.save_search_settings
    current_search_threshold = settings_ctx.current_search_threshold
    current_suggest_threshold = settings_ctx.current_suggest_threshold
    current_ui_theme_settings = settings_ctx.current_ui_theme_settings
    current_ui_layout_settings = settings_ctx.current_ui_layout_settings
    save_ui_theme_settings = settings_ctx.save_ui_theme_settings
    save_ui_layout_settings = settings_ctx.save_ui_layout_settings
    build_contact_link = settings_ctx.build_contact_link
    persist_faq_now = settings_ctx.persist_faq_now
    persist_log_now = settings_ctx.persist_log_now
    llm_chat = settings_ctx.llm_chat

    def _csv_bytes_as_utf8_sig(data) -> bytes:
        import pandas as pd

        if data is None:
            return pd.DataFrame().to_csv(index=False).encode("utf-8-sig")

        if isinstance(data, pd.DataFrame):
            df = data
        elif isinstance(data, list):
            if len(data) == 0:
                df = pd.DataFrame()
            elif isinstance(data[0], dict):
                df = pd.DataFrame(data)
            else:
                df = pd.DataFrame({"value": data})
        elif isinstance(data, dict):
            df = pd.DataFrame([data])
        else:
            df = pd.DataFrame({"value": [str(data)]})

        return df.to_csv(index=False).encode("utf-8-sig")

    def list_log_files() -> list[Path]:
        try:
            return sorted(LOG_DIR.glob("nohit_*.csv"), reverse=True)
        except Exception:
            return []

    def make_logs_zip(files) -> bytes:
        bio = io.BytesIO()
        with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in files or []:
                try:
                    zf.writestr(Path(p).name, Path(p).read_bytes())
                except Exception:
                    continue
        return bio.getvalue()

    def count_nohit_logs(days: int = 7):
        files = list_log_files()
        if not files:
            return 0, 0, 0
        today_str = datetime.now().strftime("%Y%m%d")
        today = datetime.now().date()
        recent_days = {(today - timedelta(days=i)).strftime("%Y%m%d") for i in range(days)}
        today_count = recent_count = total_count = 0
        for p in files:
            m = re.match(r"nohit_(\d{8})\.csv$", Path(p).name)
            day = m.group(1) if m else ""
            try:
                cnt = int(len(read_csv_flexible(Path(p))))
            except Exception:
                cnt = 0
            total_count += cnt
            if day == today_str:
                today_count += cnt
            if day in recent_days:
                recent_count += cnt
        return today_count, recent_count, total_count

    def read_interactions(days: int = 7) -> pd.DataFrame:
        frames = []
        for i in range(days):
            d = (datetime.now() - timedelta(days=i)).strftime("%Y%m%d")
            p = LOG_DIR / f"interactions_{d}.csv"
            if p.exists():
                try:
                    frames.append(read_csv_flexible(p))
                except Exception:
                    pass
        if not frames:
            return pd.DataFrame(columns=["timestamp", "question", "matched", "best_score", "category"])
        df_all = pd.concat(frames, ignore_index=True)
        if "matched" in df_all.columns:
            df_all["matched"] = pd.to_numeric(df_all["matched"], errors="coerce").fillna(0).astype(int)
        else:
            df_all["matched"] = 0
        if "best_score" in df_all.columns:
            df_all["best_score"] = pd.to_numeric(df_all["best_score"], errors="coerce").fillna(0.0)
        else:
            df_all["best_score"] = 0.0
        if "category" not in df_all.columns:
            df_all["category"] = ""
        return df_all

    def format_minutes_to_hours(minutes: float) -> str:
        try:
            m = float(minutes)
        except Exception:
            m = 0.0
        return f"{int(round(m))}分" if m < 60 else f"{m/60.0:.1f}時間"

    def _faq_cache_token() -> str:
        try:
            if FAQ_PATH.exists():
                stat = FAQ_PATH.stat()
                return f"{FAQ_PATH}:{stat.st_mtime_ns}:{stat.st_size}"
        except Exception:
            pass
        return str(FAQ_PATH)

    def check_password(pwd: str) -> bool:
        expected = str(st.secrets.get("ADMIN_PASSWORD", os.environ.get("ADMIN_PASSWORD", "admin")))
        return str(pwd or "") == expected

    TOP_K = 3
    st.set_page_config(page_title="情シス問い合わせAI", layout="wide")
    # Render/Streamlit初回起動時の白画面対策: 最初に軽い描画を出す
    _startup_status = st.empty()
    _startup_status.caption("🚀 情シス問い合わせAI を起動しています…")

    # ===== プロっぽい見た目（CSS）=====
    st.markdown(
        """
    <style>
    :root {
      --bg-soft: #f8fafc;
      --border: #e2e8f0;
      --text-main: #0f172a;
      --text-sub: #475569;
      --brand: #0ea5e9;
      --brand-2: #22c55e;
      --shadow: 0 10px 30px rgba(15, 23, 42, 0.08);
    }

    .block-container {padding-top: 2rem !important; padding-bottom: 9rem !important; max-width: 1180px;}
    h1, h2, h3 {line-height: 1.25 !important;}

    [data-testid="stAppViewContainer"] {
      background: radial-gradient(circle at top left, #f0f9ff 0%, #ffffff 32%, #f8fafc 100%);
    }
    [data-testid="stSidebar"] {
      background: linear-gradient(180deg, #0f172a 0%, #111827 100%);
    }
    [data-testid="stSidebar"] * {color: #e5eef8 !important;}
    [data-testid="stSidebar"] .stAlert * {color: inherit !important;}
    [data-testid="stSidebar"] [data-testid="stExpander"] {
      border: 1px solid rgba(255,255,255,0.08);
      border-radius: 16px;
      background: rgba(255,255,255,0.04);
    }

    .hero-shell {
      background: linear-gradient(135deg, #0f172a 0%, #0ea5e9 52%, #22c55e 100%);
      padding: 26px 28px;
      border-radius: 24px;
      box-shadow: 0 18px 40px rgba(14, 165, 233, 0.18);
      color: #fff;
      margin-bottom: 18px;
      position: relative;
      overflow: hidden;
    }
    .hero-shell::after {
      content: "";
      position: absolute;
      right: -60px;
      top: -60px;
      width: 220px;
      height: 220px;
      background: rgba(255,255,255,0.10);
      border-radius: 999px;
      filter: blur(4px);
    }
    .hero {position: relative; z-index: 1;}
    .hero h1 {font-size: 38px; margin: 0 0 8px 0; letter-spacing: -0.03em;}
    .hero p {margin: 0; font-size: 15px; opacity: .96; max-width: 780px;}
    .badges {margin-top: 14px; display:flex; gap:8px; flex-wrap:wrap;}
    .badge {background: rgba(255,255,255,0.16); border: 1px solid rgba(255,255,255,0.22); padding: 7px 11px; border-radius: 999px; font-size: 12px; backdrop-filter: blur(6px);}
    .cta-row {display:flex; gap:10px; flex-wrap:wrap; margin-top:14px;}
    .cta {background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.20); padding: 10px 12px; border-radius: 14px; font-size: 13px; backdrop-filter: blur(6px);}

    .topbar-card {
      background: rgba(255,255,255,0.88);
      border: 1px solid var(--border);
      border-radius: 18px;
      padding: 14px 16px;
      box-shadow: var(--shadow);
      margin-bottom: 14px;
    }
    .brand-row {display:flex; align-items:center; gap:12px;}
    .brand-title {font-size: 1.15rem; font-weight: 800; color: var(--text-main); margin:0;}
    .brand-sub {font-size: .88rem; color: var(--text-sub); margin-top:2px;}

    .glass-card, .card {
      background: rgba(255,255,255,0.88);
      border: 1px solid var(--border);
      border-radius: 18px;
      padding: 16px 18px;
      box-shadow: var(--shadow);
    }
    .card h3 {margin: 0 0 8px 0; font-size: 16px;}
    .small {font-size: 12px; color:#6b7280;}
    .section-title {font-size:20px; font-weight:800; margin: 8px 0 12px 0; color: var(--text-main);}
    .section-caption {font-size: 13px; color: var(--text-sub); margin-top: -2px; margin-bottom: 12px;}
    .query-panel {margin: 18px 0 10px 0;}
    .query-panel .eyebrow {font-size: 12px; color: #0369a1; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase;}
    .query-panel h3 {margin: 5px 0 6px 0; font-size: 22px; color: var(--text-main);}
    .query-panel p {margin: 0; color: var(--text-sub); font-size: 14px;}
    .hero-eyebrow {font-size: 12px; font-weight: 800; letter-spacing: .08em; text-transform: uppercase; opacity: .9; margin-bottom: 8px;}
    .hero-consult {margin-top: 18px;}
    .hero-consult a {display:inline-block; background:#ffffff; color:#0f172a; text-decoration:none; padding:10px 16px; border-radius:12px; font-weight:800;}

    .kpi-grid {display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 12px; margin: 14px 0 18px 0;}
    @media (max-width: 1100px){ .kpi-grid {grid-template-columns: repeat(2, minmax(0, 1fr));} }
    .kpi {background: rgba(255,255,255,0.92); border:1px solid var(--border); border-radius:20px; padding:16px 16px; box-shadow: var(--shadow);}
    .kpi .label {font-size:12px; color:#64748b; margin-bottom:8px; font-weight:600;}
    .kpi .value {font-size:30px; font-weight:800; letter-spacing:-0.03em; color:var(--text-main); margin:0;}
    .kpi .sub {font-size:12px; color:#64748b; margin-top:6px;}

    .refbox {border-left: 4px solid #0ea5e9; background: linear-gradient(180deg, #f8fbff 0%, #f8fafc 100%); padding: 12px 14px; border-radius: 14px; border:1px solid #dbeafe;}
    .answerbox {border-left: 4px solid #22c55e; background: linear-gradient(180deg, #f0fdf4 0%, #ffffff 100%); padding: 14px 16px; border-radius: 16px; line-height: 1.72; border: 1px solid #bbf7d0; box-shadow: 0 8px 20px rgba(34,197,94,0.08);}

    [data-testid="stExpander"] {border: 1px solid var(--border); border-radius: 16px; background: rgba(255,255,255,0.85);}
    [data-testid="stChatMessage"] {background: transparent;}
    [data-testid="stChatInput"] {
      background: rgba(255,255,255,0.96);
      border: 1px solid var(--border);
      border-radius: 18px;
      box-shadow: 0 10px 24px rgba(15,23,42,0.08);
    }

    .stButton > button, .stDownloadButton > button, .stLinkButton a {
      border-radius: 14px !important;
      font-weight: 700 !important;
    }
    .stButton > button[kind="primary"], .stDownloadButton > button[kind="primary"] {
      background: linear-gradient(135deg, #0ea5e9 0%, #22c55e 100%) !important;
      border: 0 !important;
      color: #fff !important;
    }

    div[data-testid="column"] .stButton > button {width: 100%; min-height: 54px;}
    @media (max-width: 768px) {
      .block-container {padding-top: 1.2rem !important; padding-bottom: 8rem !important;}
      .hero-shell {padding: 20px 18px; border-radius: 20px;}
      .hero h1 {font-size: 30px;}
      .kpi-grid {grid-template-columns: 1fr;}
    }
    </style>
    """,
        unsafe_allow_html=True,
    )

    # ===== ユーザー変更可能UI（配色 / レイアウト / ドラッグリサイズ）=====
    ui_theme = current_ui_theme_settings()
    ui_layout = current_ui_layout_settings()
    st.session_state["ui_theme_settings"] = ui_theme
    st.session_state["ui_layout_settings"] = ui_layout

    st.markdown(f"""
    <style>
    :root {{
      --user-sidebar-width: {int(ui_layout['sidebar_width'])}px;
      --user-main-max-width: {int(ui_layout['main_max_width'])}px;
      --user-main-padding-top: {int(ui_layout['main_padding_top'])}px;
      --user-main-padding-bottom: {int(ui_layout['main_padding_bottom'])}px;
      --user-card-radius: {int(ui_layout['card_radius'])}px;
      --user-card-shadow: 0 10px {int(ui_layout['card_shadow_blur'])}px rgba(15, 23, 42, {float(ui_layout['card_shadow_alpha']):.2f});
      --user-sidebar-bg-start: {ui_theme['sidebar_bg_start']};
      --user-sidebar-bg-end: {ui_theme['sidebar_bg_end']};
      --user-sidebar-text: {ui_theme['sidebar_text']};
      --user-sidebar-text-muted: {ui_theme['sidebar_text_muted']};
      --user-sidebar-panel-bg: {ui_theme['sidebar_panel_bg']};
      --user-sidebar-panel-border: {ui_theme['sidebar_panel_border']};
      --user-button-bg: {ui_theme['button_bg']};
      --user-button-text: {ui_theme['button_text']};
      --user-button-border: {ui_theme['button_border']};
      --user-button-hover-bg: {ui_theme['button_hover_bg']};
      --user-button-hover-text: {ui_theme['button_hover_text']};
      --user-button-disabled-bg: {ui_theme['button_disabled_bg']};
      --user-button-disabled-text: {ui_theme['button_disabled_text']};
      --user-main-bg-start: {ui_theme['main_bg_start']};
      --user-main-bg-mid: {ui_theme['main_bg_mid']};
      --user-main-bg-end: {ui_theme['main_bg_end']};
      --user-card-bg: {ui_theme['card_bg']};
      --user-card-border: {ui_theme['card_border']};
      --user-resizer-line: {ui_theme['resizer_line']};
      --user-resizer-knob: {ui_theme['resizer_knob']};
    }}
    [data-testid="stSidebar"] {{
      min-width: var(--user-sidebar-width) !important;
      max-width: var(--user-sidebar-width) !important;
      width: var(--user-sidebar-width) !important;
      background: linear-gradient(180deg, var(--user-sidebar-bg-start) 0%, var(--user-sidebar-bg-end) 100%) !important;
    }}
    [data-testid="stSidebar"] * {{color: var(--user-sidebar-text) !important;}}
    [data-testid="stSidebar"] small,
    [data-testid="stSidebar"] .small,
    [data-testid="stSidebar"] .stCaption,
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {{color: var(--user-sidebar-text-muted) !important;}}
    [data-testid="stSidebar"] [data-testid="stExpander"] {{
      background: var(--user-sidebar-panel-bg) !important;
      border-color: var(--user-sidebar-panel-border) !important;
      border-radius: 16px !important;
    }}
    [data-testid="stSidebar"] .stButton > button,
    [data-testid="stSidebar"] .stDownloadButton > button,
    [data-testid="stSidebar"] .stLinkButton a {{
      background: var(--user-button-bg) !important;
      color: var(--user-button-text) !important;
      border: 1px solid var(--user-button-border) !important;
    }}
    [data-testid="stSidebar"] .stButton > button:hover,
    [data-testid="stSidebar"] .stDownloadButton > button:hover,
    [data-testid="stSidebar"] .stLinkButton a:hover {{
      background: var(--user-button-hover-bg) !important;
      color: var(--user-button-hover-text) !important;
      border-color: var(--user-button-hover-bg) !important;
    }}
    [data-testid="stSidebar"] .stButton > button:disabled,
    [data-testid="stSidebar"] .stDownloadButton > button:disabled {{
      background: var(--user-button-disabled-bg) !important;
      color: var(--user-button-disabled-text) !important;
      opacity: 1 !important;
    }}
    [data-testid="stAppViewContainer"] {{
      background: radial-gradient(circle at top left, var(--user-main-bg-start) 0%, var(--user-main-bg-mid) 32%, var(--user-main-bg-end) 100%) !important;
    }}
    .block-container {{
      max-width: var(--user-main-max-width) !important;
      padding-top: var(--user-main-padding-top) !important;
      padding-bottom: var(--user-main-padding-bottom) !important;
    }}
    .topbar-card, .glass-card, .card, .kpi, [data-testid="stExpander"] {{
      border-radius: var(--user-card-radius) !important;
      box-shadow: var(--user-card-shadow) !important;
    }}
    .topbar-card, .glass-card, .card, .kpi {{
      background: var(--user-card-bg) !important;
      border-color: var(--user-card-border) !important;
    }}
    #oai-sidebar-resizer {{
      display: none !important;
    }}
    #oai-sidebar-resizer::after {{
      display: none !important;
      content: "";
    }}
    #oai-main-resizer {{
      display: none !important;
    }}
    #oai-main-resizer:hover, #oai-sidebar-resizer:hover {{opacity: 0 !important; filter: none !important;}}
    </style>
    """, unsafe_allow_html=True)

    components.html(f"""
    <script>
    (function() {{
      const doc = window.parent.document;
      const root = doc.documentElement;
      const storage = window.parent.localStorage || window.localStorage;
      const sidebarKey = 'oai_sidebar_width';
      const mainKey = 'oai_main_max_width';
      const clamp = (n, min, max) => Math.max(min, Math.min(max, n));
      const defaults = {{ sidebar: {int(ui_layout['sidebar_width'])}, main: {int(ui_layout['main_max_width'])} }};

      const applyStored = () => {{
        const sw = parseInt(storage.getItem(sidebarKey) || String(defaults.sidebar), 10);
        const mw = parseInt(storage.getItem(mainKey) || String(defaults.main), 10);
        if (!Number.isNaN(sw)) root.style.setProperty('--user-sidebar-width', clamp(sw, 240, 620) + 'px');
        if (!Number.isNaN(mw)) root.style.setProperty('--user-main-max-width', clamp(mw, 760, 2000) + 'px');
      }};

      const ensureBar = (id, title) => {{
        let el = doc.getElementById(id);
        if (!el) {{
          el = doc.createElement('div');
          el.id = id;
          el.title = title;
          doc.body.appendChild(el);
        }}
        return el;
      }};

      const oldSidebarBar = doc.getElementById('oai-sidebar-resizer');
      const oldMainBar = doc.getElementById('oai-main-resizer');
      if (oldSidebarBar) oldSidebarBar.remove();
      if (oldMainBar) oldMainBar.remove();
      applyStored();
      return;

      let drag = null;
      const onDown = (e) => {{
        if (e.target && e.target.id === 'oai-sidebar-resizer') {{
          drag = 'sidebar';
          e.preventDefault();
        }} else if (e.target && e.target.id === 'oai-main-resizer') {{
          drag = 'main';
          e.preventDefault();
        }}
      }};
      const onMove = (e) => {{
        if (!drag) return;
        if (drag === 'sidebar') {{
          const val = clamp(e.clientX, 240, 620);
          root.style.setProperty('--user-sidebar-width', val + 'px');
          storage.setItem(sidebarKey, String(val));
        }} else if (drag === 'main') {{
          const val = clamp(e.clientX - 80, 760, 2000);
          root.style.setProperty('--user-main-max-width', val + 'px');
          storage.setItem(mainKey, String(val));
        }}
      }};
      const onUp = () => {{ drag = null; }};
      const onDouble = (e) => {{
        if (e.target && e.target.id === 'oai-sidebar-resizer') {{
          storage.removeItem(sidebarKey);
          root.style.setProperty('--user-sidebar-width', defaults.sidebar + 'px');
        }} else if (e.target && e.target.id === 'oai-main-resizer') {{
          storage.removeItem(mainKey);
          root.style.setProperty('--user-main-max-width', defaults.main + 'px');
        }}
      }};

      doc.removeEventListener('mousedown', onDown);
      doc.removeEventListener('mousemove', onMove);
      doc.removeEventListener('mouseup', onUp);
      doc.removeEventListener('dblclick', onDouble);
      doc.addEventListener('mousedown', onDown);
      doc.addEventListener('mousemove', onMove);
      doc.addEventListener('mouseup', onUp);
      doc.addEventListener('dblclick', onDouble);

      setTimeout(applyStored, 50);
    }})();
    </script>
    """, height=0, width=0)

    # ===== 会社名 / ロゴ（左上）=====
    contact_link = build_contact_link()
    logo_path_obj = Path(LOGO_PATH)
    st.markdown('<div class="topbar-card">', unsafe_allow_html=True)
    col_logo, col_name, col_btn = st.columns([1, 7, 2])
    with col_logo:
        if LOGO_PATH and logo_path_obj.exists():
            st.image(str(logo_path_obj), width=54)
        else:
            st.markdown("### 🏢")
    with col_name:
        st.markdown(f'<div class="brand-title">{COMPANY_NAME}</div>', unsafe_allow_html=True)
        st.markdown('<div class="brand-sub">社内問い合わせの自己解決率を高める、情シス向けAIヘルプデスク</div>', unsafe_allow_html=True)
    with col_btn:
        if contact_link:
            st.link_button("📩 導入相談", contact_link, width="stretch")
        else:
            st.button("📩 導入相談（リンク未設定）", disabled=True, width="stretch")
    st.markdown('</div>', unsafe_allow_html=True)

    # ===== ヒーローヘッダー =====
    st.markdown(f"""
    <div class="hero-shell">
    <div class="hero">
    <div class="hero-eyebrow">導入デモ / 情シス問い合わせAI</div>
    <h1>情シス問い合わせを削減するAI</h1>
    <p>FAQを根拠付きで即回答し、未解決は問い合わせ導線へつなぎ、ナレッジを継続的に蓄積します。営業デモ、管理者運用、効果レポートまで1画面で見せられる営業仕様です。</p>
    <div class="cta-row">
    <span class="cta">✔ FAQで即回答</span>
    <span class="cta">✔ 未解決は問い合わせ誘導</span>
    <span class="cta">✔ ログ可視化と効果測定</span>
    </div>
    <div class="badges">
    <span class="badge">✅ FAQ参照（根拠表示）</span>
    <span class="badge">⚡ 高速回答</span>
    <span class="badge">📝 ログ / 該当なし蓄積</span>
    <span class="badge">🔐 管理者でFAQ育成</span>
    <span class="badge">📊 KPI・導入効果可視化</span>
    <span class="badge">📄 提案・操作資料DL</span>
    </div>
    {"<div class='hero-consult'><a href='" + contact_link + "' target='_blank'>📩 導入相談はこちら</a></div>" if contact_link else ""}
    </div>
    </div>
    """, unsafe_allow_html=True)

    try:
        _startup_status.empty()
    except Exception:
        pass

    # ==== サイドバー ========
    # ===== KPI（営業デモ + 直近7日）=====

    render_sales_kpi_sections(read_interactions=read_interactions)
    render_public_sidebar(
        contact_link=contact_link,
        count_nohit_logs=count_nohit_logs,
        read_interactions=read_interactions,
        list_log_files=list_log_files,
        make_logs_zip=make_logs_zip,
        csv_bytes_as_utf8_sig=_csv_bytes_as_utf8_sig,
        format_minutes_to_hours=format_minutes_to_hours,
    )

    # ======================
    # FAQロード（落ちやすい箇所を全てガード）
    # ======================

    FULLWIDTH_TRANS = str.maketrans({
        "０": "0", "１": "1", "２": "2", "３": "3", "４": "4",
        "５": "5", "６": "6", "７": "7", "８": "8", "９": "9",
        "ａ": "a", "ｂ": "b", "ｃ": "c", "ｄ": "d", "ｅ": "e", "ｆ": "f", "ｇ": "g",
        "ｈ": "h", "ｉ": "i", "ｊ": "j", "ｋ": "k", "ｌ": "l", "ｍ": "m", "ｎ": "n",
        "ｏ": "o", "ｐ": "p", "ｑ": "q", "ｒ": "r", "ｓ": "s", "ｔ": "t", "ｕ": "u",
        "ｖ": "v", "ｗ": "w", "ｘ": "x", "ｙ": "y", "ｚ": "z",
        "Ａ": "a", "Ｂ": "b", "Ｃ": "c", "Ｄ": "d", "Ｅ": "e", "Ｆ": "f", "Ｇ": "g",
        "Ｈ": "h", "Ｉ": "i", "Ｊ": "j", "Ｋ": "k", "Ｌ": "l", "Ｍ": "m", "Ｎ": "n",
        "Ｏ": "o", "Ｐ": "p", "Ｑ": "q", "Ｒ": "r", "Ｓ": "s", "Ｔ": "t", "Ｕ": "u",
        "Ｖ": "v", "Ｗ": "w", "Ｘ": "x", "Ｙ": "y", "Ｚ": "z",
    })

    CANONICAL_PATTERNS = [
        (r"デスクトップパソコン|デスクトップｐｃ|desktop\s*pc", "デスクトップpc"),
        (r"ノートパソコン|ラップトップ", "ノートpc"),
        (r"パーソナルコンピュータ|パソコン|ピーシー|ｐｃ|pc端末", "pc"),
        (r"コンピューター", "コンピュータ"),
        (r"無線lan|wi-?fi|wifi|ワイファイ", "wifi"),
        (r"ｖｐｎ|ぶいぴーえぬ|vpn接続", "vpn"),
        (r"サインイン", "ログイン"),
        (r"サインアウト", "ログアウト"),
        (r"パスコード|passcode", "パスワード"),
        (r"pw", "パスワード"),
        (r"パスワードを忘れました|パスワードを忘れた|パスワード忘れた|パスワードがわからない|password forgotten|forgot password", "パスワード リセット"),
        (r"パスワード再発行|パスワード初期化", "パスワード リセット"),
        (r"立ち上がらない|起ち上がらない|立ちあがらない|起ちあがらない", "起動しない"),
        (r"電源がつかない|電源が付かない|電源がはいらない", "電源が入らない"),
        (r"ログイン出来ない|ログインできません", "ログインできない"),
        (r"接続できない|接続できません|接続出来ない|接続出来ません|接続しない|接続されない|つながりません|繋がりません|繋がらない|繋げない|つなげない", "つながらない"),
        (r"利用できない|使用できない", "使えない"),
        (r"開けない", "起動しない"),
        (r"印字できない|プリントできない", "印刷できない"),
        (r"メール送れない", "メールが送信できない"),
        (r"メール受け取れない", "メールが受信できない"),
        (r"認証に失敗|認証エラー", "認証できない"),
        (r"ロックされた|凍結された", "ロックされた"),
    ]

    CONCEPT_ALIASES = {
        "vpn": ["vpn", "リモートアクセス", "社外接続"],
        "network": ["ネットワーク", "wifi", "lan", "通信", "internet", "インターネット"],
        "login": ["ログイン", "サインイン", "認証", "アカウント"],
        "password": ["パスワード", "password", "pw", "リセット", "再設定", "初期化"],
        "boot": ["起動", "立ち上が", "立上", "電源", "シャットダウン", "再起動"],
        "mail": ["メール", "outlook", "受信", "送信"],
        "print": ["印刷", "プリンタ", "printer", "print"],
        "lock": ["ロック", "凍結", "無効", "停止"],
        "error": ["エラー", "失敗", "不具合", "異常", "障害"],
        "cannot": ["できない", "できません", "使えない", "つながらない", "入らない", "起動しない"],
        "pc_device": ["pc", "パソコン", "ノートpc", "デスクトップpc", "端末", "windows"],
        "office_app": ["excel", "word", "powerpoint", "access", "onenote", "office"],
        "browser_app": ["chrome", "edge", "firefox", "ブラウザ", "web", "sso"],
        "mail_app": ["outlook", "exchange", "共有メール", "メーリングリスト", "メールアプリ"],
    }



    FASTLANE_INTENT_RULES = [
        {
            "name": "password_reset",
            "query_any": ["パスワード", "password", "pw", "認証情報"],
            "query_any2": ["忘れ", "わから", "不明", "リセット", "再設定", "変更", "思い出せ", "失念"],
            "faq_any": ["パスワード", "password", "pw", "リセット", "再設定", "初期化"],
        },
        {
            "name": "account_lock",
            "query_any": ["アカウント", "ログイン", "認証"],
            "query_any2": ["ロック", "凍結", "無効", "停止"],
            "faq_any": ["ロック", "凍結", "無効", "停止", "アカウント"],
        },
        {
            "name": "vpn_connect",
            "query_any": ["vpn", "リモートアクセス", "社外接続"],
            "query_any2": ["つながらない", "接続", "入らない", "失敗", "できない"],
            "faq_any": ["vpn", "リモートアクセス", "接続"],
        },
    ]


    search_ctx = create_search_runtime(
        st=st,
        pd=pd,
        Path=Path,
        FAQ_PATH=FAQ_PATH,
        TfidfVectorizer=TfidfVectorizer,
        cosine_similarity=cosine_similarity,
        normalize_faq_columns=normalize_faq_columns,
        read_csv_flexible=read_csv_flexible,
        current_search_settings=current_search_settings,
        llm_chat=llm_chat,
        _faq_cache_token=_faq_cache_token,
        SENTENCE_TRANSFORMERS_AVAILABLE=SENTENCE_TRANSFORMERS_AVAILABLE,
        SentenceTransformer=SentenceTransformer,
        FULLWIDTH_TRANS=FULLWIDTH_TRANS,
        CANONICAL_PATTERNS=CANONICAL_PATTERNS,
        CONCEPT_ALIASES=CONCEPT_ALIASES,
        FASTLANE_INTENT_RULES=FASTLANE_INTENT_RULES,
        create_faq_index_runtime=create_faq_index_runtime,
    )

    faq_index_ctx = search_ctx.faq_index_ctx
    normalize_search_text = search_ctx.normalize_search_text
    extract_search_tokens = search_ctx.extract_search_tokens
    extract_concepts = search_ctx.extract_concepts
    load_faq_df = search_ctx.load_faq_df
    prepare_faq_dataframe = search_ctx.prepare_faq_dataframe
    _contains_any = search_ctx._contains_any
    _faq_row_matches_words = search_ctx._faq_row_matches_words
    _is_fastlane_query_text = search_ctx._is_fastlane_query_text
    _score_fast_candidate = search_ctx._score_fast_candidate
    _domain_penalty = search_ctx._domain_penalty
    try_ultrafast_answer = search_ctx.try_ultrafast_answer
    retrieve_faq_cached = search_ctx.retrieve_faq_cached
    retrieve_faq = search_ctx.retrieve_faq
    _fastlane_direct_answer = search_ctx._fastlane_direct_answer
    llm_answer_cached = search_ctx.llm_answer_cached
    build_prompt = search_ctx.build_prompt

    WORD_VECTORIZER = faq_index_ctx.WORD_VECTORIZER
    CHAR_VECTORIZER = faq_index_ctx.CHAR_VECTORIZER
    SENTENCE_MODEL_NAME = faq_index_ctx.SENTENCE_MODEL_NAME
    _load_sentence_transformer_model = faq_index_ctx._load_sentence_transformer_model
    _build_sentence_embeddings = faq_index_ctx._build_sentence_embeddings
    _get_sentence_embeddings_cached = faq_index_ctx._get_sentence_embeddings_cached
    _search_with_sentence_transformers = faq_index_ctx._search_with_sentence_transformers
    load_faq_index = faq_index_ctx.load_faq_index
    get_faq_index_state = faq_index_ctx.get_faq_index_state
    reset_faq_index_runtime = faq_index_ctx.reset_faq_index_runtime
    ensure_faq_index_loaded = faq_index_ctx.ensure_faq_index_loaded
    get_faq_index_cached = faq_index_ctx.get_faq_index_cached
    _build_fast_lookup_maps = faq_index_ctx._build_fast_lookup_maps

    def nohit_template():
        return """
    FAQに該当がありませんでした。

    情シスへお問い合わせの際は、以下の情報を添えてください：

    ・何ができないか（具体的な操作）
    ・エラー画面のスクリーンショット
    ・発生時刻
    ・利用場所（社内 / 社外）
    ・ネットワーク（Wi-Fi / VPN）
    ・端末（Windows / Mac）
    ・影響範囲（自分のみ / 他の人も）

    ※これらを共有いただくと対応が早くなります。
    ※このAIは御社の運用に合わせてカスタマイズ可能です。
    """.strip()



    def _ensure_nohit_schema(path: Path):
        """既存nohit CSVが旧形式（timestamp,questionのみ）でも、新スキーマに移行する。"""
        cols = ["timestamp", "question", "device", "location", "network", "error_text", "impact", "channel"]
        if not path.exists():
            return cols

        try:
            # 先頭行だけ見る（軽量）
            with path.open("r", encoding="utf-8", errors="ignore") as f:
                header = f.readline().strip()
            header_cols = [h.strip() for h in header.split(",")] if header else []
        except Exception:
            header_cols = []

        # すでに新スキーマなら何もしない
        if set(cols).issubset(set(header_cols)):
            return header_cols

        # 移行：既存を読み取り→新ヘッダで書き直し
        try:
            old_df = read_csv_flexible(path)
            if old_df is None:
                old_df = pd.DataFrame()
        except Exception:
            old_df = pd.DataFrame()

        if len(old_df) == 0:
            # 空なら新ヘッダで作り直す
            with path.open("w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(cols)
            return cols

        qcol = pick_question_column(old_df.columns) or ("question" if "question" in old_df.columns else None)
        tcol = "timestamp" if "timestamp" in old_df.columns else None

        rows = []
        for _, r in old_df.iterrows():
            ts = str(r.get(tcol, "")).strip() if tcol else ""
            q = str(r.get(qcol, "")).strip() if qcol else ""
            if not q:
                continue
            rows.append([ts, q, "", "", "", "", "", ""])

        with path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(cols)
            w.writerows(rows)

        return cols


    def log_nohit(question: str, extra: dict | None = None) -> str:
        """該当なしログを追記して、記録したtimestamp（秒）を返す。"""
        if not question:
            return ""
        extra = extra or {}
        day = datetime.now().strftime("%Y%m%d")
        path = LOG_DIR / f"nohit_{day}.csv"
        cols = _ensure_nohit_schema(path)

        ts = datetime.now().isoformat(timespec="seconds")
        row = {
            "timestamp": ts,
            "question": question,
            "device": extra.get("device", ""),
            "location": extra.get("location", ""),
            "network": extra.get("network", ""),
            "error_text": extra.get("error_text", ""),
            "impact": extra.get("impact", ""),
            "channel": extra.get("channel", "web"),
        }

        try:
            is_new = not path.exists()
            with path.open("a", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                if is_new:
                    w.writerow(cols)
                w.writerow([row.get(c, "") for c in cols])
            persist_log_now(path)
        except Exception:
            pass
        return ts


    def update_nohit_record(day: str, timestamp: str, question: str, extra: dict) -> bool:
        """同じday/timestamp/question の行があれば更新。無ければ追記。"""
        if not day or not timestamp or not question:
            return False
        path = LOG_DIR / f"nohit_{day}.csv"
        cols = _ensure_nohit_schema(path)

        try:
            df_log = read_csv_flexible(path)
            if df_log is None:
                df_log = pd.DataFrame(columns=cols)
        except Exception:
            df_log = pd.DataFrame(columns=cols)

        # 必須列を揃える
        for c in cols:
            if c not in df_log.columns:
                df_log[c] = ""

        # 既存行を更新（最初の一致）
        mask = (df_log["timestamp"].astype(str) == str(timestamp)) & (df_log["question"].astype(str) == str(question))
        idxs = df_log.index[mask].tolist()
        if idxs:
            i = idxs[0]
            for k, v in (extra or {}).items():
                if k in df_log.columns:
                    df_log.at[i, k] = v
            df_log.at[i, "channel"] = extra.get("channel", df_log.at[i, "channel"] or "web")
        else:
            # 無ければ追記
            row = {c: "" for c in cols}
            row["timestamp"] = timestamp
            row["question"] = question
            for k, v in (extra or {}).items():
                if k in row:
                    row[k] = v
            if not row.get("channel"):
                row["channel"] = "web"
            df_log = pd.concat([df_log, pd.DataFrame([row])], ignore_index=True)

        # UTF-8で書き戻す（Excel対応ならutf-8-sigでもOK）
        try:
            with path.open("w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(cols)
                for _, r in df_log[cols].iterrows():
                    w.writerow([str(r.get(c, "")) for c in cols])
            persist_log_now(path)
            return True
        except Exception:
            return False


    def seed_nohit_questions(n: int = 20) -> int:
        """本番前のデモ用：情シス定番のnohit質問を今日のログに追加する。"""
        seeds = [
            "VPNにつながらない", "Outlookの送受信ができない", "Teamsにログインできない", "パスワードを忘れた",
            "アカウントがロックされた", "共有フォルダにアクセスできない", "プリンタが印刷できない", "Wi-Fiが頻繁に切れる",
            "PCが重い", "PCが固まる", "Excelが起動しない", "Excelがフリーズする", "OneDriveが同期しない",
            "メール添付ファイルが開けない", "二段階認証が通らない", "カメラが映らない", "マイクが認識されない",
            "ソフトのインストール申請方法が分からない", "Windows更新が終わらない", "画面が真っ黒になる",
        ]
        added = 0
        for q in seeds[:n]:
            ts = log_nohit(q, {"channel": "seed"})
            if ts:
                added += 1
        return added



    def log_interaction(question: str, matched: bool, best_score: float, category: str):
        """全ての質問をログ化（削減時間の見える化用）: logs/interactions_YYYYMMDD.csv"""
        if not question:
            return
        day = datetime.now().strftime("%Y%m%d")
        path = LOG_DIR / f"interactions_{day}.csv"
        try:
            is_new = not path.exists()
            with path.open("a", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                if is_new:
                    w.writerow(["timestamp", "question", "matched", "best_score", "category"])
                w.writerow([datetime.now().isoformat(timespec="seconds"), question, int(bool(matched)), float(best_score), category or ""])
            persist_log_now(path)
        except Exception:
            pass




    import json

    def normalize_question(q: str) -> str:
        q = (q or "").strip().lower()
        q = re.sub(r"\s+", " ", q)
        # 日本語も含めて記号類をざっくり除去
        q = re.sub(r"[\u3000\s\t\r\n]+", " ", q)
        q = re.sub(r"[!！?？。、,.，:：;；\-_=+~`'\"()（）\[\]{}<>＜＞/\\|@#%^&*]", "", q)
        return q.strip()



    def load_nohit_questions_from_logs(files, max_questions: int = 100) -> list[str]:
        """nohit_*.csv から質問を収集（新しいログから優先）。文字コード/カラム揺れに強く読む。"""
        questions: list[str] = []
        seen: set[str] = set()
        for p in files:
            try:
                _df = read_csv_flexible(Path(p))
                if _df is None or len(_df) == 0:
                    continue

                qcol = pick_question_column(_df.columns)
                if not qcol:
                    continue

                for q in _df[qcol].fillna("").astype(str).tolist():
                    nq = normalize_question(q)
                    if not nq:
                        continue
                    if nq in seen:
                        continue
                    seen.add(nq)
                    questions.append(q.strip())
                    if len(questions) >= max_questions:
                        return questions
            except Exception:
                continue
        return questions



    def generate_faq_candidates(nohit_questions: list[str], n_items: int = 8) -> pd.DataFrame:
        """該当なしログからFAQ案を生成してDataFrameで返す（category/question/answer）。"""
        if not nohit_questions:
            return pd.DataFrame(columns=["category", "question", "answer"])

        # 入力が長すぎると落ちるので上限
        max_in = min(len(nohit_questions), 80)
        sample = nohit_questions[:max_in]
        examples = "\n".join([f"- {q}" for q in sample])

        prompt = f"""あなたは社内情シスのベテラン担当です。
    以下は『FAQに該当なし』として蓄積された、社員からの問い合わせ例です。

    【目的】
    この問い合わせ例から、社内で使えるFAQ（Q&A）を {n_items} 件作成してください。

    【要件】
    - 日本語のみ
    - 1件ごとに category / question / answer を作る
    - answer は手順を箇条書きで（3〜7行）
    - 個人情報や会社固有の秘密情報は作らない
    - できるだけ汎用的（どの会社でも通用）に
    - 出力は必ずJSONのみ（前後に説明文を入れない）
    - コードブロック ``` は使わない

    【出力JSON形式】
    [
      {{"category":"VPN", "question":"...", "answer":"- ...\n- ..."}},
      ...
    ]

    【問い合わせ例】
    {examples}
    """

        out = llm_chat(
            [
                {"role": "system", "content": "あなたは情シスのFAQ作成者です。出力はJSONのみ。"},
                {"role": "user", "content": prompt},
            ]
        )

        out_text = out if isinstance(out, str) else str(out)

        json_text = extract_json_array(out_text) or out_text.strip()

        try:
            data = json.loads(json_text)
            if not isinstance(data, list):
                raise ValueError("JSON is not a list")
        except Exception:
            return pd.DataFrame(columns=["category", "question", "answer"])

        rows = []
        for item in data:
            if not isinstance(item, dict):
                continue
            cat = str(item.get("category", "")).strip()
            q = str(item.get("question", "")).strip()
            a = str(item.get("answer", "")).strip()
            if not q or not a:
                continue
            rows.append({"category": cat, "question": q, "answer": a})

        return pd.DataFrame(rows, columns=["category", "question", "answer"])

    def append_faq_csv(faq_path: Path, new_df: pd.DataFrame) -> int:
        """faq.csv に追記。重複（question）をざっくり除外して追記件数を返す。"""
        if new_df is None or len(new_df) == 0:
            return 0

        # 必須列を揃える
        for col in ["question", "answer", "category"]:
            if col not in new_df.columns:
                new_df[col] = ""

        new_df = new_df[["question", "answer", "category"]].copy()
        new_df["question"] = new_df["question"].fillna("").astype(str).str.strip()
        new_df["answer"] = new_df["answer"].fillna("").astype(str).str.strip()
        new_df["category"] = new_df["category"].fillna("").astype(str).str.strip()
        new_df = new_df[(new_df["question"] != "") & (new_df["answer"] != "")]
        if len(new_df) == 0:
            return 0

        # 既存読み込み
        if faq_path.exists():
            try:
                exist = normalize_faq_columns(read_csv_flexible(faq_path))
            except Exception:
                exist = pd.DataFrame(columns=["question", "answer", "category"])
        else:
            exist = pd.DataFrame(columns=["question", "answer", "category"])

        exist_q = set(normalize_question(x) for x in exist.get("question", pd.Series(dtype=str)).fillna("").astype(str).tolist())

        rows = []
        for _, r in new_df.iterrows():
            nq = normalize_question(str(r.get("question", "")))
            if not nq:
                continue
            if nq in exist_q:
                continue
            exist_q.add(nq)
            rows.append([r["question"], r["answer"], r.get("category", "")])

        if not rows:
            return 0

        is_new = not faq_path.exists()
        with faq_path.open("a", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            if is_new:
                w.writerow(["question", "answer", "category"])
            w.writerows(rows)

        persist_faq_now()
        return len(rows)



    def generate_slack_bot_zip_bytes():
        """Slack Bot 完全版コード一式をZIPで返す。既存アプリ本体とは分離し、Render等に別サービスとして配置する想定。"""
        import textwrap

        render_base = "https://your-render-url.onrender.com"

        slack_bot_py = textwrap.dedent("""    import os
        import hmac
        import hashlib
        import time
        import json
        from typing import Any

        import requests
        from flask import Flask, request, jsonify

        app = Flask(__name__)

        HELP_DESK_ASK_URL = os.getenv("HELPDESK_ASK_URL", "https://your-streamlit-or-api-url/ask")
        SLACK_SIGNING_SECRET = os.getenv("SLACK_SIGNING_SECRET", "")
        REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "30"))


        def verify_slack_request(req) -> bool:
            if not SLACK_SIGNING_SECRET:
                return True

            timestamp = req.headers.get("X-Slack-Request-Timestamp", "")
            slack_signature = req.headers.get("X-Slack-Signature", "")
            if not timestamp or not slack_signature:
                return False

            try:
                if abs(time.time() - int(timestamp)) > 60 * 5:
                    return False
            except Exception:
                return False

            body = req.get_data(as_text=True)
            basestring = f"v0:{timestamp}:{body}"
            my_signature = "v0=" + hmac.new(
                SLACK_SIGNING_SECRET.encode("utf-8"),
                basestring.encode("utf-8"),
                hashlib.sha256,
            ).hexdigest()

            return hmac.compare_digest(my_signature, slack_signature)


        def ask_helpdesk(question: str, user_name: str = "", channel_name: str = "") -> str:
            payload: dict[str, Any] = {
                "question": question,
                "source": "slack",
                "user_name": user_name,
                "channel_name": channel_name,
            }

            r = requests.post(
                HELP_DESK_ASK_URL,
                json=payload,
                timeout=REQUEST_TIMEOUT,
            )
            r.raise_for_status()

            data = r.json()
            answer = data.get("answer") or data.get("message") or "回答を取得できませんでした。"
            return str(answer)


        @app.get("/")
        def health() -> tuple[str, int]:
            return "ok", 200


        @app.post("/slack/command")
        def slack_command():
            if not verify_slack_request(request):
                return jsonify({"text": "署名検証に失敗しました。"}), 403

            question = (request.form.get("text") or "").strip()
            user_name = request.form.get("user_name", "")
            channel_name = request.form.get("channel_name", "")

            if not question:
                return jsonify({
                    "response_type": "ephemeral",
                    "text": "質問文を入れてください。例: /helpdesk VPNがつながらない",
                })

            try:
                answer = ask_helpdesk(question, user_name=user_name, channel_name=channel_name)
                return jsonify({
                    "response_type": "in_channel",
                    "text": f"*質問:* {question}\n*回答:* {answer}",
                })
            except Exception as e:
                return jsonify({
                    "response_type": "ephemeral",
                    "text": f"問い合わせAIへの接続に失敗しました: {e}",
                }), 500


        @app.post("/slack/events")
        def slack_events():
            if not verify_slack_request(request):
                return jsonify({"error": "invalid signature"}), 403

            data = request.get_json(silent=True) or {}

            if data.get("type") == "url_verification":
                return jsonify({"challenge": data.get("challenge", "")})

            event = data.get("event", {})
            if event.get("type") == "app_mention":
                text = event.get("text", "")
                return jsonify({"ok": True, "note": f"mention received: {text}"})

            return jsonify({"ok": True})


        if __name__ == "__main__":
            port = int(os.getenv("PORT", "3000"))
            app.run(host="0.0.0.0", port=port)
        """).strip() + "\n"

        requirements_txt = textwrap.dedent("""    flask==3.0.3
        requests==2.32.3
        gunicorn==22.0.0
        """).strip() + "\n"

        render_yaml = textwrap.dedent("""    services:
          - type: web
            name: slack-helpdesk-bot
            env: python
            plan: free
            buildCommand: pip install -r requirements.txt
            startCommand: gunicorn slack_bot:app
            autoDeploy: true
        """).strip() + "\n"

        env_example = textwrap.dedent(f"""    HELPDESK_ASK_URL={render_base}/ask
        SLACK_SIGNING_SECRET=your_signing_secret
        REQUEST_TIMEOUT=30
        """).strip() + "\n"

        readme_md = textwrap.dedent("""    # Slack Helpdesk Bot 完全版

        このZIPは、既存の Streamlit アプリ本体とは別サービスとして Render に配置する想定です。

        ## 含まれるファイル
        - `slack_bot.py` : Slack Slash Command / Events 受信用 Flask アプリ
        - `requirements.txt` : 必要ライブラリ
        - `render.yaml` : Render デプロイ設定例
        - `.env.example` : 環境変数の雛形

        ## Slack 側設定
        1. Slack API で App を作成
        2. Slash Commands に `/helpdesk` を追加
        3. Request URL に `https://あなたのRenderURL/slack/command` を設定
        4. Event Subscriptions を使う場合は `https://あなたのRenderURL/slack/events` を設定
        5. Signing Secret を Render の環境変数 `SLACK_SIGNING_SECRET` に設定

        ## Render 側設定
        1. このZIPを GitHub リポジトリに配置
        2. Render で New + → Web Service
        3. Build Command: `pip install -r requirements.txt`
        4. Start Command: `gunicorn slack_bot:app`
        5. `HELPDESK_ASK_URL` に既存問い合わせAIの API URL を設定

        ## 重要
        既存の Streamlit アプリに `/ask` API が無い場合は、別途 API 追加が必要です。
        まずは Slack Bot コード一式を先に配布し、後から API 側を接続しても構いません。
        """).strip() + "\n"

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("slack_bot.py", slack_bot_py)
            zf.writestr("requirements.txt", requirements_txt)
            zf.writestr("render.yaml", render_yaml)
            zf.writestr(".env.example", env_example)
            zf.writestr("README.md", readme_md)

        buf.seek(0)
        return buf.getvalue()


    # ======================
    # 管理者ログイン
    # ======================
    ensure_admin_session_state()
    render_admin_login_sidebar(check_password=check_password)
    render_admin_tools_if_logged_in(
        render_admin_complete_tools=render_admin_complete_tools,
        read_interactions=read_interactions,
        count_nohit_logs=count_nohit_logs,
        list_log_files=list_log_files,
        make_logs_zip=make_logs_zip,
        load_nohit_questions_from_logs=load_nohit_questions_from_logs,
        generate_faq_candidates=generate_faq_candidates,
        append_faq_csv=append_faq_csv,
        seed_nohit_questions=seed_nohit_questions,
        faq_path=FAQ_PATH,
    )

    if st.session_state.is_admin:
        with st.expander("🎯 検索精度設定", expanded=False):
            current_cfg = current_search_settings()
            st.caption("管理者が検索の厳しさを分かりやすく調整できます。まずは上の基本設定だけ触れば十分です。")
            st.info(
                f"現在値：自動回答 {current_cfg['answer_threshold']:.2f} / 候補表示 {current_cfg['suggest_threshold']:.2f} / "
                f"単語重視 {int(current_cfg['word_weight'] * 100)}% / 文字重視 {int(current_cfg['char_weight'] * 100)}%"
            )

            st.markdown("#### ① 基本設定（通常はここだけ）")
            admin_answer_threshold = st.slider(
                "自動回答しきい値",
                min_value=0.10,
                max_value=1.20,
                value=float(current_cfg["answer_threshold"]),
                step=0.01,
                key="admin_answer_threshold_slider",
                help="この値以上ならそのまま回答します。高いほど慎重、低いほど積極的です。",
            )
            max_suggest_value = max(0.05, round(admin_answer_threshold - 0.05, 2))
            suggest_default = min(float(current_cfg["suggest_threshold"]), max_suggest_value)
            admin_suggest_threshold = st.slider(
                "候補表示しきい値",
                min_value=0.05,
                max_value=max_suggest_value,
                value=suggest_default,
                step=0.01,
                key="admin_suggest_threshold_slider",
                help="この値以上かつ自動回答未満なら『近いFAQ候補』として表示します。",
            )

            search_balance = st.radio(
                "検索バランス",
                options=["バランス型", "単語重視", "表記ゆれ重視"],
                index=0 if abs(float(current_cfg["word_weight"]) - 0.54) < 0.03 else (1 if float(current_cfg["word_weight"]) >= 0.60 else 2),
                horizontal=True,
                help="単語重視は意味の近い語句に強く、表記ゆれ重視は細かな言い回し違いに強くなります。",
            )
            if search_balance == "単語重視":
                word_weight, char_weight = 0.65, 0.35
            elif search_balance == "表記ゆれ重視":
                word_weight, char_weight = 0.40, 0.60
            else:
                word_weight, char_weight = 0.54, 0.46

            answer_gap = round(admin_answer_threshold - admin_suggest_threshold, 2)
            if answer_gap >= 0.18:
                st.success("判定差は広めです。誤回答を抑えやすい設定です。")
            elif answer_gap >= 0.10:
                st.info("判定差は標準です。迷ったときは候補表示へ回しやすい設定です。")
            else:
                st.warning("判定差が狭めです。自動回答と候補表示の境目が近くなります。")

            st.markdown("""- 自動回答以上: 通常回答
- 候補表示以上: 近いFAQ候補を表示
- 候補表示未満: 該当なしとして追加情報フォームへ""")

            with st.expander("🔧 詳細設定（上級者向け）", expanded=False):
                st.caption("より細かく精度を触りたい場合だけ使ってください。未設定なら基本設定のままでも十分です。")

                c_adv1, c_adv2 = st.columns(2)
                with c_adv1:
                    exact_bonus = st.slider("完全一致ボーナス", 0.00, 0.80, float(current_cfg["exact_bonus"]), 0.01, key="search_exact_bonus")
                    contains_bonus = st.slider("部分一致ボーナス", 0.00, 0.60, float(current_cfg["contains_bonus"]), 0.01, key="search_contains_bonus")
                    token_bonus_max = st.slider("単語一致ボーナス上限", 0.00, 0.80, float(current_cfg["token_bonus_max"]), 0.01, key="search_token_bonus")
                    concept_bonus_max = st.slider("概念一致ボーナス上限", 0.00, 0.80, float(current_cfg["concept_bonus_max"]), 0.01, key="search_concept_bonus")
                    prefix_bonus = st.slider("書き出し一致ボーナス", 0.00, 0.30, float(current_cfg["prefix_bonus"]), 0.01, key="search_prefix_bonus")
                    top_k = st.slider("候補として保持する件数", 1, 5, int(current_cfg["top_k"]), 1, key="search_top_k")
                with c_adv2:
                    semantic_enabled = st.checkbox("意味検索を使う", value=bool(current_cfg["semantic_enabled"]), key="search_semantic_enabled")
                    semantic_skip_fastlane = st.checkbox("頻出問い合わせでは意味検索を省略", value=bool(current_cfg["semantic_skip_fastlane"]), key="search_semantic_skip_fastlane")
                    semantic_boost = st.slider("意味検索の補正強さ", 0.00, 0.80, float(current_cfg["semantic_boost"]), 0.01, key="search_semantic_boost")
                    semantic_candidate_count = st.slider("意味検索をかける候補数", 1, 20, int(current_cfg["semantic_candidate_count"]), 1, key="search_semantic_candidate_count")
                    semantic_min_query_len = st.slider("意味検索を始める最小文字数", 1, 50, int(current_cfg["semantic_min_query_len"]), 1, key="search_semantic_min_query_len")
                    semantic_trigger_min = st.slider("意味検索を始める下限スコア", 0.00, 1.20, float(current_cfg["semantic_trigger_min"]), 0.01, key="search_semantic_trigger_min")
                    semantic_trigger_max = st.slider("意味検索を始める上限スコア", max(semantic_trigger_min, 0.00), 1.50, float(max(current_cfg["semantic_trigger_max"], semantic_trigger_min)), 0.01, key="search_semantic_trigger_max")

            col_th1, col_th2 = st.columns(2)
            with col_th1:
                if st.button("💾 検索設定を保存", width="stretch"):
                    ok, settings = save_search_settings(
                        admin_answer_threshold,
                        admin_suggest_threshold,
                        extra_settings={
                            "word_weight": word_weight,
                            "char_weight": char_weight,
                            "exact_bonus": exact_bonus,
                            "contains_bonus": contains_bonus,
                            "token_bonus_max": token_bonus_max,
                            "concept_bonus_max": concept_bonus_max,
                            "prefix_bonus": prefix_bonus,
                            "semantic_enabled": semantic_enabled,
                            "semantic_skip_fastlane": semantic_skip_fastlane,
                            "semantic_boost": semantic_boost,
                            "semantic_candidate_count": semantic_candidate_count,
                            "semantic_min_query_len": semantic_min_query_len,
                            "semantic_trigger_min": semantic_trigger_min,
                            "semantic_trigger_max": semantic_trigger_max,
                            "top_k": top_k,
                        },
                    )
                    st.session_state.search_threshold = settings["answer_threshold"]
                    st.session_state.suggest_threshold = settings["suggest_threshold"]
                    st.session_state.search_settings = settings
                    if ok:
                        st.success(
                            f"保存しました。自動回答={settings['answer_threshold']:.2f} / 候補表示={settings['suggest_threshold']:.2f} / "
                            f"単語重視={int(settings['word_weight'] * 100)}%"
                        )
                    else:
                        st.warning("ローカル保存またはGitHub保存に失敗した可能性があります。設定値自体はこのセッションに反映しています。")
                    st.rerun()
            with col_th2:
                if st.button("↩ 初期値に戻す", width="stretch"):
                    ok, settings = save_search_settings(extra_settings=default_search_settings())
                    st.session_state.search_threshold = settings["answer_threshold"]
                    st.session_state.suggest_threshold = settings["suggest_threshold"]
                    st.session_state.search_settings = settings
                    if ok:
                        st.success("検索設定を初期値に戻しました。")
                    else:
                        st.warning("初期値に戻しましたが、外部保存に失敗した可能性があります。")
                    st.rerun()


            # ===== FAQ自動生成（該当なしログ → FAQ案）=====
        with st.expander("🧠 LLM切替設定", expanded=False):
            current_llm = current_llm_settings()
            provider = st.radio(
                "利用するLLM",
                options=["groq", "ollama"],
                index=0 if current_llm["provider"] == "groq" else 1,
                format_func=lambda x: "Groq（クラウド・高速）" if x == "groq" else "Ollama（ローカル・社内完結）",
                key="llm_provider_radio",
            )
            groq_model = st.text_input("Groqモデル名", value=current_llm["groq_model"], key="groq_model_input")
            ollama_model = st.text_input("Ollamaモデル名", value=current_llm["ollama_model"], key="ollama_model_input")
            ollama_base_url = st.text_input("Ollama URL", value=current_llm["ollama_base_url"], key="ollama_base_url_input")

            col_llm1, col_llm2 = st.columns(2)
            with col_llm1:
                if st.button("💾 LLM設定を保存", width="stretch", key="save_llm_settings"):
                    ok, saved = save_llm_settings({
                        "provider": provider,
                        "groq_model": groq_model,
                        "ollama_model": ollama_model,
                        "ollama_base_url": ollama_base_url,
                    })
                    st.session_state["llm_settings"] = saved
                    if ok:
                        st.success("LLM設定を保存しました。")
                    else:
                        st.warning("LLM設定は反映済みですが、保存に失敗した可能性があります。")
                    st.rerun()
            with col_llm2:
                if st.button("↩ LLM設定を初期値に戻す", width="stretch", key="reset_llm_settings"):
                    default_llm = default_llm_settings()
                    save_llm_settings(default_llm)
                    st.session_state["llm_settings"] = default_llm
                    st.rerun()

            provider_label = "Groq（クラウド・高速）" if provider == "groq" else "Ollama（ローカル・社内完結）"
            st.caption(f"現在の利用先: {provider_label}")
            if provider == "ollama":
                st.info("Ollamaはローカル実行です。初回はモデルを事前に pull してください。")

        with st.expander("🎨 UI配色設定", expanded=False):
            current_theme = current_ui_theme_settings()

            c1, c2 = st.columns(2)
            with c1:
                sidebar_bg_start = st.color_picker("左メニュー背景（開始色）", current_theme["sidebar_bg_start"], key="ui_sidebar_bg_start")
                sidebar_text = st.color_picker("左メニュー文字色", current_theme["sidebar_text"], key="ui_sidebar_text")
                sidebar_text_muted = st.color_picker("左メニュー補助文字色", current_theme["sidebar_text_muted"], key="ui_sidebar_text_muted")
                button_bg = st.color_picker("ボタン背景色", current_theme["button_bg"], key="ui_button_bg")
                button_text = st.color_picker("ボタン文字色", current_theme["button_text"], key="ui_button_text")
                button_border = st.color_picker("ボタン枠線色", current_theme["button_border"], key="ui_button_border")
                button_hover_bg = st.color_picker("ボタンホバー背景", current_theme["button_hover_bg"], key="ui_button_hover_bg")
                button_hover_text = st.color_picker("ボタンホバー文字", current_theme["button_hover_text"], key="ui_button_hover_text")
                button_disabled_bg = st.color_picker("無効ボタン背景色", current_theme["button_disabled_bg"], key="ui_button_disabled_bg")
                button_disabled_text = st.color_picker("無効ボタン文字色", current_theme["button_disabled_text"], key="ui_button_disabled_text")
            with c2:
                sidebar_bg_end = st.color_picker("左メニュー背景（終了色）", current_theme["sidebar_bg_end"], key="ui_sidebar_bg_end")
                main_bg_start = st.color_picker("メイン背景（開始色）", current_theme["main_bg_start"], key="ui_main_bg_start")
                main_bg_mid = st.color_picker("メイン背景（中央色）", current_theme["main_bg_mid"], key="ui_main_bg_mid")
                main_bg_end = st.color_picker("メイン背景（終了色）", current_theme["main_bg_end"], key="ui_main_bg_end")
                card_border = st.color_picker("カード枠線色", current_theme["card_border"], key="ui_card_border")
                resizer_knob = st.color_picker("ドラッグつまみ色", current_theme["resizer_knob"], key="ui_resizer_knob")

            sidebar_panel_bg = st.text_input("左メニューパネル背景（hex または rgba）", value=current_theme["sidebar_panel_bg"], key="ui_sidebar_panel_bg")
            sidebar_panel_border = st.text_input("左メニューパネル枠線（hex または rgba）", value=current_theme["sidebar_panel_border"], key="ui_sidebar_panel_border")
            card_bg = st.text_input("カード背景色（hex または rgba）", value=current_theme["card_bg"], key="ui_card_bg")
            resizer_line = st.text_input("ドラッグライン色（hex または rgba）", value=current_theme["resizer_line"], key="ui_resizer_line")

            live_theme = sanitize_ui_theme_settings({
                "sidebar_bg_start": sidebar_bg_start,
                "sidebar_bg_end": sidebar_bg_end,
                "sidebar_text": sidebar_text,
                "sidebar_text_muted": sidebar_text_muted,
                "sidebar_panel_bg": sidebar_panel_bg,
                "sidebar_panel_border": sidebar_panel_border,
                "button_bg": button_bg,
                "button_text": button_text,
                "button_border": button_border,
                "button_hover_bg": button_hover_bg,
                "button_hover_text": button_hover_text,
                "button_disabled_bg": button_disabled_bg,
                "button_disabled_text": button_disabled_text,
                "main_bg_start": main_bg_start,
                "main_bg_mid": main_bg_mid,
                "main_bg_end": main_bg_end,
                "card_bg": card_bg,
                "card_border": card_border,
                "resizer_line": resizer_line,
                "resizer_knob": resizer_knob,
            })
            st.session_state["ui_theme_settings"] = live_theme

            col_ui1, col_ui2 = st.columns(2)
            with col_ui1:
                if st.button("💾 UI配色を保存", width="stretch", key="save_ui_theme"):
                    ok, _ = save_ui_theme_settings(live_theme)
                    st.success("UI配色を保存しました。" if ok else "UI配色は反映済みですが、保存に失敗した可能性があります。")
            with col_ui2:
                if st.button("↩ UI配色を初期値に戻す", width="stretch", key="reset_ui_theme"):
                    default_theme = default_ui_theme_settings()
                    save_ui_theme_settings(default_theme)
                    for k, v in default_theme.items():
                        st.session_state[f"ui_{k}"] = v
                    st.session_state["ui_theme_settings"] = default_theme
                    st.rerun()

        with st.expander("📐 UIレイアウト設定", expanded=False):
            current_layout = current_ui_layout_settings()
            sidebar_width = st.slider("左メニュー幅", 240, 620, int(current_layout["sidebar_width"]), key="ui_layout_sidebar_width")
            main_max_width = st.slider("メイン画面の最大幅", 760, 2000, int(current_layout["main_max_width"]), step=10, key="ui_layout_main_max_width")
            main_padding_top = st.slider("上余白", 4, 96, int(current_layout["main_padding_top"]), key="ui_layout_main_padding_top")
            main_padding_bottom = st.slider("下余白", 72, 280, int(current_layout["main_padding_bottom"]), key="ui_layout_main_padding_bottom")
            card_radius = st.slider("フレーム角丸", 8, 40, int(current_layout["card_radius"]), key="ui_layout_card_radius")
            card_shadow_blur = st.slider("フレーム影のぼかし", 0, 80, int(current_layout["card_shadow_blur"]), key="ui_layout_card_shadow_blur")
            card_shadow_alpha_pct = st.slider("フレーム影の濃さ", 0, 40, int(round(float(current_layout["card_shadow_alpha"]) * 100)), key="ui_layout_card_shadow_alpha")

            live_layout = sanitize_ui_layout_settings({
                "sidebar_width": sidebar_width,
                "main_max_width": main_max_width,
                "main_padding_top": main_padding_top,
                "main_padding_bottom": main_padding_bottom,
                "card_radius": card_radius,
                "card_shadow_blur": card_shadow_blur,
                "card_shadow_alpha": card_shadow_alpha_pct,
            })
            st.session_state["ui_layout_settings"] = live_layout

            components.html(f"""
            <script>
            (function() {{
                const doc = window.parent.document;
                const root = doc.documentElement;
                root.style.setProperty('--user-sidebar-width', '{live_layout['sidebar_width']}px');
                root.style.setProperty('--user-main-max-width', '{live_layout['main_max_width']}px');
                root.style.setProperty('--user-main-padding-top', '{live_layout['main_padding_top']}px');
                root.style.setProperty('--user-main-padding-bottom', '{live_layout['main_padding_bottom']}px');
                root.style.setProperty('--user-card-radius', '{live_layout['card_radius']}px');
                root.style.setProperty('--user-card-shadow', '0 10px {live_layout['card_shadow_blur']}px rgba(15, 23, 42, {live_layout['card_shadow_alpha']:.2f})');
                window.localStorage.setItem('oai_sidebar_width', '{live_layout['sidebar_width']}');
                window.localStorage.setItem('oai_main_max_width', '{live_layout['main_max_width']}');
            }})();
            </script>
            """, height=0, width=0)

            col_layout1, col_layout2 = st.columns(2)
            with col_layout1:
                if st.button("💾 UIレイアウトを保存", width="stretch", key="save_ui_layout"):
                    ok, _ = save_ui_layout_settings(live_layout)
                    st.success("UIレイアウトを保存しました。" if ok else "UIレイアウトは反映済みですが、保存に失敗した可能性があります。")
            with col_layout2:
                if st.button("↩ UIレイアウトを初期値に戻す", width="stretch", key="reset_ui_layout"):
                    default_layout = default_ui_layout_settings()
                    save_ui_layout_settings(default_layout)
                    st.session_state["ui_layout_settings"] = default_layout
                    window_local_js = f"""<script>(function(){{const doc=window.parent.document;const root=doc.documentElement;root.style.setProperty('--user-sidebar-width','{default_layout['sidebar_width']}px');root.style.setProperty('--user-main-max-width','{default_layout['main_max_width']}px');localStorage.removeItem('oai_sidebar_width');localStorage.removeItem('oai_main_max_width');}})();</script>"""
                    components.html(window_local_js, height=0, width=0)
                    st.rerun()

        with st.expander("📂 FAQ管理（Excelダウンロード / アップロード）", expanded=False):
            st.caption("管理者は FAQ を Excel(.xlsx) で一括入出力できます。500件以上でもまとめて置き換え可能です。推奨列名は『質問 / 回答 / カテゴリ』です。")

            if st.session_state.get("faq_replace_result"):
                st.success(st.session_state["faq_replace_result"])
                st.session_state.pop("faq_replace_result", None)

            current_faq_df = normalize_faq_columns(read_csv_flexible(FAQ_PATH)) if FAQ_PATH.exists() else pd.DataFrame(columns=["question", "answer", "category"])
            excel_bytes = faq_df_to_excel_bytes(current_faq_df)
            st.download_button(
                "⬇ 現在のFAQをExcelでダウンロード",
                data=excel_bytes,
                file_name="faq.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
            st.caption(f"現在登録中のFAQ件数: {len(current_faq_df)} 件")

            uploaded_faq = st.file_uploader(
                "FAQファイルをアップロード",
                type=["xlsx", "xls", "csv"],
                key="faq_excel_uploader_admin",
                help="Excel(.xlsx) 推奨。質問 / 回答 / カテゴリ、または question / answer / category に対応。類義語を含む代表質問（例: パソコンが起動しません / 電源が入りません）を入れると精度が上がります。",
            )

            if uploaded_faq is not None:
                try:
                    incoming_df = read_faq_uploaded_file(uploaded_faq.name, uploaded_faq.getvalue())
                    st.success(f"アップロード確認OK: {len(incoming_df)} 件のFAQを検出しました。")
                    preview_df = incoming_df.rename(columns={"question": "質問", "answer": "回答", "category": "カテゴリ"})
                    st.dataframe(preview_df.head(20), width="stretch", height=420)
                    if len(incoming_df) > 20:
                        st.caption(f"先頭20件を表示中です。保存対象は全 {len(incoming_df)} 件です。")

                    if st.button("📥 この内容でFAQを反映する", type="primary", key="replace_faq_excel_admin", width="stretch"):
                        with st.spinner("FAQを保存しています..."):
                            saved = save_faq_csv_full(FAQ_PATH, incoming_df)
                            reloaded_df = normalize_faq_columns(read_csv_flexible(FAQ_PATH)) if FAQ_PATH.exists() else pd.DataFrame(columns=["question", "answer", "category"])

                            try:
                                load_faq_index.clear()
                                get_faq_index_state.clear()
                                reset_faq_index_runtime()
                            except Exception:
                                pass
                            try:
                                st.cache_resource.clear()
                            except Exception:
                                pass
                            try:
                                st.cache_data.clear()
                            except Exception:
                                pass

                            if int(saved) != int(len(reloaded_df)):
                                st.error(f"保存件数と再読込件数が一致しません。保存: {saved} 件 / 再読込: {len(reloaded_df)} 件")
                            else:
                                msg = f"FAQを {saved} 件反映しました。現在登録中のFAQ件数も {len(reloaded_df)} 件です。"
                                st.session_state["faq_replace_result"] = msg
                                st.success(msg)
                                st.info("FAQの反映が完了しました。再読み込みは不要です。GitHub永続化ONなら自動で外部保存されます。")
                                current_faq_df = reloaded_df
                except Exception as e:
                    st.error(f"FAQファイルの取込でエラー: {e}")

        # =========================
        # 管理者向け資料（PDF）ダウンロード
        # =========================
        with st.expander("📘 管理者向け資料（PDF）", expanded=False):
            if not REPORTLAB_AVAILABLE:
                st.warning("PDF出力には reportlab が必要です。requirements.txt に reportlab を追加してください。")
            else:
                col_a, col_b = st.columns(2)
                with col_a:
                    ops_pdf = generate_ops_manual_pdf()
                    st.download_button(
                        "📄 操作説明書PDFをダウンロード",
                        data=ops_pdf,
                        file_name="操作説明書_情シス問い合わせAI.pdf",
                        mime="application/pdf",
                        width="stretch",
                    )
                with col_b:
                    proposal_pdf = generate_sales_proposal_pdf()
                    st.download_button(
                        "📑 提案資料PDFをダウンロード",
                        data=proposal_pdf,
                        file_name="提案資料_情シス問い合わせAI.pdf",
                        mime="application/pdf",
                        width="stretch",
                    )
                st.caption("※ どちらもアプリの現状に合わせて自動生成されます（必要に応じて文面はカスタマイズ可能）。")

        with st.expander("📄 効果レポート（PDF）", expanded=False):
            if not REPORTLAB_AVAILABLE:
                st.warning("PDF出力には reportlab が必要です。requirements.txt に reportlab を追加してください。")
            else:
                hourly_cost = st.number_input(
                    "想定人件費（円/時間）",
                    min_value=0,
                    max_value=20000,
                    value=int(st.session_state.get("hourly_cost", 4000)),
                    step=500,
                    key="admin_hourly_cost",
                )

                avg_min_pdf = float(st.session_state.get("avg_min", 5))
                deflect_pdf = float(st.session_state.get("deflect_rate", st.session_state.get("deflect", 0.7)))

                df_month_all = read_interactions(days=60)
                if df_month_all is None or len(df_month_all) == 0:
                    st.caption("今月の利用ログがまだありません。質問すると自動で蓄積します。")
                else:
                    try:
                        ts = pd.to_datetime(df_month_all["timestamp"], errors="coerce")
                        month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                        df_month = df_month_all[ts >= month_start]
                    except Exception:
                        df_month = df_month_all

                    try:
                        pdf_bytes = generate_effect_report_pdf(
                            df=df_month,
                            avg_min=avg_min_pdf,
                            deflect=deflect_pdf,
                            hourly_cost_yen=int(hourly_cost),
                        )
                        st.download_button(
                            "📄 今月の導入効果レポートをダウンロード",
                            data=pdf_bytes,
                            file_name=f"effect_report_{datetime.now().strftime('%Y%m')}.pdf",
                            mime="application/pdf",
                            width="stretch",
                        )
                    except Exception as e:
                        st.error(f"PDF生成でエラー: {e}")

    # ======================
    # セッション初期化
    # ======================
    if "used_hits" not in st.session_state:
        st.session_state.used_hits = []

    if "messages" not in st.session_state:
        st.session_state.messages = []

    if "pending_q" not in st.session_state:
        st.session_state.pending_q = ""

    if "search_threshold" not in st.session_state:
        st.session_state.search_threshold = float(SEARCH_SETTINGS.get("answer_threshold", DEFAULT_SEARCH_THRESHOLD))

    if "suggest_threshold" not in st.session_state:
        st.session_state.suggest_threshold = float(SEARCH_SETTINGS.get("suggest_threshold", DEFAULT_SUGGEST_THRESHOLD))

    if "search_settings" not in st.session_state:
        st.session_state.search_settings = _sanitize_search_settings(SEARCH_SETTINGS)
    if "llm_settings" not in st.session_state:
        st.session_state.llm_settings = sanitize_llm_settings(LLM_SETTINGS)


    def render_used_hits_expander(used_hits, best_score: float, answer_threshold: float, was_nohit: bool = False):
        with st.expander("🔎 回答の根拠を見る", expanded=False):
            if was_nohit or not used_hits:
                st.markdown('<div class="refbox">該当なし（スコアが低いため問い合わせ誘導）</div>', unsafe_allow_html=True)
                return
            st.caption(f"上位FAQ候補の一致度を表示しています（best score: {best_score:.2f} / 自動回答しきい値: {answer_threshold:.2f}）。")
            for i, (row, score) in enumerate(used_hits, 1):
                render_match_bar(score)
                q_html = str(row.get("question", "")).replace("\n", "<br>")
                a_html = str(row.get("answer", "")).replace("\n", "<br>")
                cat = str(row.get("category", ""))
                match_pct = int(max(0.0, min(1.0, float(score))) * 100)
                st.markdown(
                    f"""
    <div class="refbox">
    <b>FAQ{i}</b>（一致度：{match_pct}% / category={cat}）<br>
    <b>Q:</b> {q_html}<br>
    <b>A:</b> {a_html}
    </div>
    """,
                    unsafe_allow_html=True,
                )


    def render_quick_start_buttons():
        st.markdown('<div class="glass-card query-panel"><div class="eyebrow">Quick Start</div><h3>よくある問い合わせをワンクリックで試す</h3><p>デモで見せやすい代表質問を用意しています。クリックするとそのまま送信されます。</p></div>', unsafe_allow_html=True)
        st.markdown("### 💡 おすすめ質問（クリックで送信）")
        c1, c2, c3 = st.columns(3)

        if c1.button("🔐 パスワードを忘れた"):
            st.session_state.pending_q = "パスワードを忘れました"
            st.rerun()

        if c2.button("🧩 アカウントがロックされた"):
            st.session_state.pending_q = "アカウントがロックされました"
            st.rerun()

        if c3.button("🌐 VPNに接続できない"):
            st.session_state.pending_q = "VPNに接続できません"
            st.rerun()


    show_welcome = len(st.session_state.messages) == 0
    if show_welcome:
        st.markdown('<div class="glass-card query-panel"><div class="eyebrow">AI Demo</div><h3>実際の問い合わせをそのまま入力してください</h3><p>例：パスワードを忘れました / VPNにつながらない / ディスプレイが真っ暗です</p></div>', unsafe_allow_html=True)
        render_quick_start_buttons()

    # ======================
    # チャット履歴表示
    # ======================
    for m in st.session_state.messages:
        with st.chat_message(m.get("role", "assistant")):
            st.markdown(m.get("content", ""), unsafe_allow_html=True)


    def build_suggest_answer(user_q: str, hits) -> str:
        if not hits:
            return nohit_template()
        row, score = hits[0]
        q = str(row.get("question", "")).strip()
        a = str(row.get("answer", "")).strip()
        cat = str(row.get("category", "")).strip()
        parts = [
            "入力内容に近いFAQ候補が見つかりました。完全一致ではありませんが、まずはこちらを確認してください。",
        ]
        if q:
            parts.append(f"【候補FAQ】{q}")
        if cat:
            parts.append(f"【カテゴリ】{cat}")
        if a:
            parts.append(f"【回答】\n{a}")
        parts.append("解決しない場合は、下の『追加情報を記録』から状況を補足してください。")
        return "\n\n".join(parts)


    def render_nohit_extra_form(info: dict | None = None, expanded: bool = True):
        """『該当なし』直後に表示する追加情報フォーム（端末/利用場所/ネットワーク等）。"""
        info = info or (st.session_state.get("pending_nohit", {}) or {})
        with st.expander("📝 追加情報を記録（任意）", expanded=expanded):
            st.caption("解決しない場合は、状況を少し補足するとFAQ改善に役立ちます。")
            with st.form("nohit_extra_form", clear_on_submit=False):
                c1, c2, c3 = st.columns(3)
                with c1:
                    device = st.selectbox(
                        "端末",
                        ["", "Windows", "Mac", "iPhone/iPad", "Android", "不明"],
                        index=0,
                        key="nohit_device",
                    )
                with c2:
                    location = st.selectbox(
                        "利用場所",
                        ["", "社内", "社外", "不明"],
                        index=0,
                        key="nohit_location",
                    )
                with c3:
                    network = st.selectbox(
                        "ネットワーク",
                        ["", "Wi-Fi", "有線", "VPN", "モバイル回線", "不明"],
                        index=0,
                        key="nohit_network",
                    )

                impact = st.selectbox(
                    "影響範囲",
                    ["", "自分のみ", "他の人も", "不明"],
                    index=0,
                    key="nohit_impact",
                )
                error_text = st.text_area(
                    "エラー内容（任意）",
                    placeholder="例：0x80190001 / '資格情報が無効です' など",
                    key="nohit_error_text",
                )

                submitted = st.form_submit_button("✅ この内容で記録")
                if submitted:
                    ok = update_nohit_record(
                        day=str(info.get("day", "")),
                        timestamp=str(info.get("timestamp", "")),
                        question=str(info.get("question", "")),
                        extra={
                            "device": device,
                            "location": location,
                            "network": network,
                            "impact": impact,
                            "error_text": error_text,
                            "channel": "web",
                        },
                    )
                    if ok:
                        st.success("追加情報をログに保存しました。ありがとうございます！")
                        # 次回以降はフォームを閉じる（保持は消す）
                        st.session_state["pending_nohit_active"] = False
                    else:
                        st.warning("保存に失敗しました（もう一度お試しください）。")


    # ======================
    # 入力 → 検索 → 回答
    # ======================
    # ===== 該当なし（nohit）の追加情報フォーム =====
    if st.session_state.get("pending_nohit_active"):
        render_nohit_extra_form(expanded=True)

    # 先に chat_input を描画（画面下に固定されます）
    chat_typed = st.chat_input("情シス問い合わせを入力してください")

    # 入力が無いときは pending_q（おすすめボタン等）を使う
    user_q = (chat_typed or st.session_state.get("pending_q", ""))

    # 「おすすめ質問」など pending_q から来たかどうか
    used_pending = (not chat_typed) and bool(st.session_state.get("pending_q", ""))

    if user_q:
        st.session_state.pending_q = ""

        st.session_state.messages.append({"role": "user", "content": user_q})
        with st.chat_message("user"):
            st.markdown(user_q)

        ultrafast = try_ultrafast_answer(user_q)
        if ultrafast:
            hits = ultrafast.get("hits", [])
            best_score = float(ultrafast.get("best_score", 0.0))
        else:
            packed_hits = retrieve_faq_cached(user_q, _faq_cache_token())
            local_df, *_ = ensure_faq_index_loaded()
            hits = []
            for idx, score in packed_hits:
                try:
                    if local_df is not None and int(idx) >= 0:
                        hits.append((local_df.iloc[int(idx)], float(score)))
                except Exception:
                    continue
            best_score = hits[0][1] if hits else 0.0

        if hits:
            render_match_bar(best_score)

        answer_threshold = current_search_threshold()
        suggest_threshold = current_suggest_threshold()

        if ultrafast:
            used_hits = hits[:1]
            answer = str(ultrafast.get("answer", "")).strip() or nohit_template()
            was_nohit = False
            was_suggest = False
        elif best_score < suggest_threshold:
            used_hits = []
            answer = nohit_template()
            ts_nohit = log_nohit(user_q)
            st.session_state["last_nohit"] = {"day": datetime.now().strftime("%Y%m%d"), "timestamp": ts_nohit, "question": user_q}
            was_nohit = True
            was_suggest = False
        elif best_score < answer_threshold:
            used_hits = hits[:1]
            answer = build_suggest_answer(user_q, used_hits)
            was_nohit = False
            was_suggest = True
        else:
            used_hits = hits
            was_nohit = False
            was_suggest = False
            faq_answer = ""
            top_question = ""
            try:
                faq_answer = str(hits[0][0].get("answer", "")).strip()
                top_question = str(hits[0][0].get("question", "")).strip()
            except Exception:
                faq_answer = ""
                top_question = ""

            fastlane_answer = _fastlane_direct_answer(
                user_q=user_q,
                hits=hits,
                best_score=float(best_score),
                answer_threshold=float(answer_threshold),
                suggest_threshold=float(suggest_threshold),
            )

            if fastlane_answer:
                answer = fastlane_answer
            else:
                prompt = build_prompt(user_q, hits)
                cached_answer = llm_answer_cached(user_q, prompt, _faq_cache_token(), top_question)
                if cached_answer:
                    answer = cached_answer
                else:
                    answer = faq_answer if faq_answer else "現在AIの回答機能でエラーが発生しています。しばらくしてから再度お試しください。"

        st.session_state.used_hits = used_hits

        # 利用ログ（削減時間の見える化用）
        top_cat = ""
        if used_hits:
            try:
                top_cat = str(used_hits[0][0].get("category", ""))
            except Exception:
                top_cat = ""
        log_interaction(user_q, matched=(best_score >= answer_threshold), best_score=best_score, category=top_cat)

        with st.chat_message("assistant"):
            answer_html = str(answer).replace("\n", "<br>")
            st.markdown(f'<div class="answerbox">{answer_html}</div>', unsafe_allow_html=True)

            if 'was_nohit' in locals() and was_nohit:
                # 「該当なし」のとき、追加情報フォームを"次のrerunでも"表示できるように保持
                st.session_state["pending_nohit_active"] = True
                st.session_state["pending_nohit"] = st.session_state.get("last_nohit", {})
                st.info("該当なしログに追加しました。必要なら下の『追加情報を記録』で状況を補足できます。")
            elif 'was_suggest' in locals() and was_suggest:
                st.info(f"近いFAQ候補を表示しています（スコア: {best_score:.2f} / 自動回答しきい値: {answer_threshold:.2f}）。")
                st.caption("管理者はサイドバーの『検索精度設定』から判定基準を調整できます。")

        st.session_state.messages.append({"role": "assistant", "content": str(answer)})

        # nohit はフォーム表示のために再描画
        if "was_nohit" in locals() and was_nohit:
            st.rerun()

        # おすすめ質問ボタンから自動送信した場合は、もう一度 rerun して入力欄を確実に表示
        if used_pending:
            st.rerun()

