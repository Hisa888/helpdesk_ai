import streamlit as st

st.set_page_config(page_title="情シス問い合わせAI")

st.write("🚀 情シス問い合わせAI 起動成功")


from pathlib import Path
import pandas as pd
import io
import re
import json
import requests

import os
import re
import uuid
import csv
import io
import zipfile
import base64

# ===== CSV読み込みを頑丈にする（文字コード/区切り/カラム揺れ対策）=====
def read_csv_flexible(path: Path) -> pd.DataFrame:
    """CSVをできる限り失敗しないで読む（UTF-8/UTF-8-SIG/CP932等 + delimiter推定）。"""
    # まずは生bytesを読む
    raw = path.read_bytes()
    # 文字コード候補
    encs = ["utf-8", "utf-8-sig", "cp932", "shift_jis"]
    last_err = None
    text = None
    for enc in encs:
        try:
            text = raw.decode(enc)
            break
        except Exception as e:
            last_err = e
            continue
    if text is None:
        # 最終手段：latin1で無理やり
        text = raw.decode("latin1", errors="ignore")

    # delimiter推定（csv.Snifferが失敗する場合もあるのでガード）
    import csv as _csv
    sample = text[:5000]
    delim = ","
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        delim = dialect.delimiter
    except Exception:
        # 行にタブが多ければタブ、それ以外はカンマ
        if sample.count("\t") > sample.count(","):
            delim = "\t"

    # pandasで読む（on_bad_linesはpandas2系で有効）
    try:
        return pd.read_csv(io.StringIO(text), sep=delim, engine="python", on_bad_lines="skip")
    except Exception:
        try:
            return pd.read_csv(io.StringIO(text), sep=delim, engine="python")
        except Exception:
            # それでもダメなら空
            return pd.DataFrame()

def pick_question_column(cols) -> str | None:
    """質問カラム名の揺れを吸収"""
    cand = [
        "question", "質問", "問い合わせ", "問合せ", "query", "user_question",
        "content", "text"
    ]
    for c in cand:
        if c in cols:
            return c
    # 大文字小文字無視
    lower_map = {str(c).lower(): c for c in cols}
    for c in cand:
        if c.lower() in lower_map:
            return lower_map[c.lower()]
    return None

def extract_json_array(text: str) -> str | None:
    """LLM出力から最初のJSON配列（[...]）を抜き出す。"""
    if not text:
        return None
    s = str(text).strip()
    # ```json ... ``` を除去
    s = re.sub(r"^```(?:json)?\s*|\s*```$", "", s.strip(), flags=re.IGNORECASE)
    # 先頭から最初の [ ... ] を抽出（DOTALL）
    m = re.search(r"\[[\s\S]*\]", s)
    return m.group(0).strip() if m else None



def normalize_faq_columns(df: pd.DataFrame) -> pd.DataFrame:
    """列名揺れを吸収して question/answer/category に正規化する。"""
    if df is None or len(df) == 0:
        return pd.DataFrame(columns=["question", "answer", "category"])
    out = df.copy()
    rename_map = {}
    for c in out.columns:
        original = str(c).strip()
        key = original.lower()
        if original in ["質問", "問い合わせ", "問合せ"] or key in ["question", "query"]:
            rename_map[c] = "question"
        elif original in ["回答"] or key in ["answer", "answer_text", "reply"]:
            rename_map[c] = "answer"
        elif original in ["カテゴリ", "分類"] or key in ["category"]:
            rename_map[c] = "category"
    out = out.rename(columns=rename_map)
    for col in ["question", "answer", "category"]:
        if col not in out.columns:
            out[col] = ""
    out = out[["question", "answer", "category"]].copy()
    for col in ["question", "answer", "category"]:
        if col in out.columns:
            out[col] = out[col].fillna("").astype(str).str.replace("\n", " ").str.replace("\r", " ")

    out = out[(out["question"] != "") & (out["answer"] != "")].reset_index(drop=True)

    return out

def _xlsx_escape(text: str) -> str:
    return (str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', '&quot;'))


def _build_minimal_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "FAQ") -> bytes:
    from io import BytesIO
    rows = [list(df.columns)] + df.fillna("").astype(str).values.tolist()
    shared_strings = []
    sst_index = {}
    def get_sst_idx(value: str) -> int:
        value = str(value)
        if value not in sst_index:
            sst_index[value] = len(shared_strings)
            shared_strings.append(value)
        return sst_index[value]
    sheet_rows = []
    for r_idx, row in enumerate(rows, start=1):
        cells = []
        for c_idx, value in enumerate(row, start=1):
            col = ""
            n = c_idx
            while n:
                n, rem = divmod(n - 1, 26)
                col = chr(65 + rem) + col
            ref = f"{col}{r_idx}"
            s_idx = get_sst_idx(value)
            cells.append(f'<c r="{ref}" t="s"><v>{s_idx}</v></c>')
        sheet_rows.append(f'<row r="{r_idx}">' + "".join(cells) + '</row>')
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetData>' + ''.join(sheet_rows) + '</sheetData>'
        '</worksheet>'
    )
    sst_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(shared_strings)}" uniqueCount="{len(shared_strings)}">'
        + ''.join(f'<si><t xml:space="preserve">{_xlsx_escape(s)}</t></si>' for s in shared_strings) +
        '</sst>'
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets><sheet name="{_xlsx_escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets>'
        '</workbook>'
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        '</Relationships>'
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        '</Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        '</Types>'
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        '</styleSheet>'
    )
    core_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
                'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
                'xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
                '<dc:title>FAQ</dc:title><dc:creator>ChatGPT</dc:creator></cp:coreProperties>')
    app_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
               '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
               'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
               '<Application>Microsoft Excel</Application></Properties>')
    bio = BytesIO()
    with zipfile.ZipFile(bio, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', content_types)
        zf.writestr('_rels/.rels', root_rels)
        zf.writestr('docProps/core.xml', core_xml)
        zf.writestr('docProps/app.xml', app_xml)
        zf.writestr('xl/workbook.xml', workbook_xml)
        zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels)
        zf.writestr('xl/styles.xml', styles_xml)
        zf.writestr('xl/sharedStrings.xml', sst_xml)
        zf.writestr('xl/worksheets/sheet1.xml', sheet_xml)
    return bio.getvalue()


def faq_df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    export_df = normalize_faq_columns(df).rename(columns={"question": "質問", "answer": "回答", "category": "カテゴリ"})
    return _build_minimal_xlsx_bytes(export_df, sheet_name="FAQ")


def _read_xlsx_bytes(raw: bytes) -> pd.DataFrame:
    """XLSXを安全にDataFrame化する。
    1) pandas + openpyxl を優先
    2) openpyxl の values_only 読み
    3) 最後の手段で XML 手動解析
    """
    from io import BytesIO

    # 1) まずは pandas + openpyxl
    try:
        df = pd.read_excel(BytesIO(raw), engine="openpyxl")
        if df is not None:
            df.columns = [str(c).strip() for c in df.columns]
            return df
    except Exception:
        pass

    # 2) openpyxl の values_only 読み
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            max_cols = max(len(r) if r is not None else 0 for r in rows)
            norm_rows = []
            for r in rows:
                r = list(r or [])
                r += [""] * (max_cols - len(r))
                norm_rows.append(["" if v is None else str(v) for v in r])
            header = [str(x).strip() for x in norm_rows[0]]
            body = norm_rows[1:] if len(norm_rows) > 1 else []
            return pd.DataFrame(body, columns=header)
    except Exception:
        pass

    # 3) フォールバック: XML手動解析（ふりがな rPh は無視）
    import xml.etree.ElementTree as ET
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def _text_without_rph(parent):
        parts = []
        for child in list(parent):
            tag = child.tag.split("}")[-1]
            if tag == "rPh":
                continue
            if tag == "t":
                parts.append(child.text or "")
            else:
                for tnode in child.findall(".//a:t", ns):
                    parts.append(tnode.text or "")
        return "".join(parts)

    with zipfile.ZipFile(BytesIO(raw)) as zf:
        shared = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", ns):
                shared.append(_text_without_rph(si))

        sheet_path = "xl/worksheets/sheet1.xml"
        if "xl/workbook.xml" in zf.namelist() and "xl/_rels/workbook.xml.rels" in zf.namelist():
            wb = ET.fromstring(zf.read("xl/workbook.xml"))
            rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rid_to_target = {
                rel.attrib.get("Id"): rel.attrib.get("Target")
                for rel in rels.findall("pr:Relationship", ns)
            }
            first_sheet = wb.find(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/"
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
            )
            if first_sheet is not None:
                rid = first_sheet.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                target = rid_to_target.get(rid)
                if target:
                    if not target.startswith("worksheets/"):
                        target = target.split("xl/")[-1]
                    sheet_path = "xl/" + target

        sheet = ET.fromstring(zf.read(sheet_path))
        rows = []
        for row in sheet.findall("a:sheetData/a:row", ns):
            row_map = {}
            max_col = 0
            for c in row.findall("a:c", ns):
                ref = c.attrib.get("r", "")
                col_letters = "".join(ch for ch in ref if ch.isalpha())
                col_idx = 0
                for ch in col_letters:
                    col_idx = col_idx * 26 + (ord(ch.upper()) - 64)
                max_col = max(max_col, col_idx)

                t = c.attrib.get("t")
                value = ""
                if t == "s":
                    v = c.find("a:v", ns)
                    if v is not None and (v.text or "").isdigit():
                        si_idx = int(v.text)
                        value = shared[si_idx] if 0 <= si_idx < len(shared) else ""
                elif t == "inlineStr":
                    is_el = c.find("a:is", ns)
                    if is_el is not None:
                        value = _text_without_rph(is_el)
                else:
                    v = c.find("a:v", ns)
                    value = v.text if v is not None and v.text is not None else ""

                if col_idx > 0:
                    row_map[col_idx] = value

            if max_col:
                rows.append([row_map.get(i, "") for i in range(1, max_col + 1)])

    if not rows:
        return pd.DataFrame()
    max_cols = max(len(r) for r in rows)
    rows = [r + [""] * (max_cols - len(r)) for r in rows]
    header = [str(x).strip() for x in rows[0]]
    body = rows[1:] if len(rows) > 1 else []
    return pd.DataFrame(body, columns=header)


def _strip_excel_phonetic_artifacts(text: str) -> str:
    """Excel由来のふりがな混入っぽい末尾カタカナを保守的に除去する。"""
    s = "" if text is None else str(text)
    s = s.replace("\u3000", " ").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not s:
        return s

    # 句読点の直後にだけ付いた読み仮名を除去
    s = re.sub(r"([。．.!?！？]\s*)([ァ-ヶーｦ-ﾟ]{1,20})$", r"\1", s)

    # 漢字を含む語の末尾にだけ付いたカタカナ読みを除去
    m = re.match(r"^(.*[一-龥々〆ヵヶ])([ァ-ヶー]{2,20})$", s)
    if m:
        prefix, suffix = m.groups()
        if not re.search(r"[ァ-ヶー]$", prefix):
            s = prefix

    return s.strip()


def read_faq_uploaded_file(file_name: str, raw: bytes) -> pd.DataFrame:
    """FAQアップロード読込。
    まずCSV/XLSXを安全にDataFrame化し、列名を正規化する。
    文字化けやExcel由来のふりがな混入を抑えるため、読み込み後に文字列を明示的に整形する。
    """
    suffix = Path(file_name).suffix.lower()
    if suffix == '.csv':
        tmp = Path('/tmp/_faq_upload.csv')
        tmp.write_bytes(raw)
        df = read_csv_flexible(tmp)
    else:
        df = _read_xlsx_bytes(raw)

    df = normalize_faq_columns(df)
    for col in ["question", "answer", "category"]:
        if col in df.columns:
            df[col] = (
                df[col]
                .fillna("")
                .astype(str)
                .map(_strip_excel_phonetic_artifacts)
                .str.replace("\u3000", " ", regex=False)
                .str.replace("\r\n", "\n", regex=False)
                .str.replace("\r", "\n", regex=False)
                .str.strip()
            )
    return df


def save_faq_csv_full(faq_path: Path, df: pd.DataFrame) -> int:
    clean = normalize_faq_columns(df)
    faq_path.parent.mkdir(parents=True, exist_ok=True)
    clean.to_csv(faq_path, index=False, encoding='utf-8-sig', quoting=csv.QUOTE_ALL)
    if faq_path == FAQ_PATH:
        persist_faq_now()
    return len(clean)

# ===== PDF生成（ReportLab）===== 
REPORTLAB_AVAILABLE = False
try:
    from reportlab.pdfgen import canvas  # type: ignore
    from reportlab.lib.pagesizes import A4  # type: ignore
    from reportlab.lib.units import mm  # type: ignore
    from reportlab.lib.colors import HexColor  # type: ignore
    from reportlab.pdfbase import pdfmetrics  # type: ignore
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont  # type: ignore
    REPORTLAB_AVAILABLE = True
except ModuleNotFoundError:
    # Streamlit Cloudでは requirements.txt に reportlab を追加してください
    REPORTLAB_AVAILABLE = False


from pathlib import Path
from datetime import datetime, timedelta

import streamlit as st
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# sentence-transformers は任意。未導入でも既存機能だけで動くようにする
SENTENCE_TRANSFORMERS_AVAILABLE = False
try:
    from sentence_transformers import SentenceTransformer  # type: ignore
    SENTENCE_TRANSFORMERS_AVAILABLE = True
except Exception:
    SentenceTransformer = None  # type: ignore
    SENTENCE_TRANSFORMERS_AVAILABLE = False

from services.auth import check_password
from services.llm_router import chat as llm_chat

# ======================
# 基本設定
# ======================
ROOT_DIR = Path(".")
ROOT_FAQ_PATH = ROOT_DIR / "faq.csv"
DATA_DIR = ROOT_DIR / "runtime_data"
FAQ_PATH = DATA_DIR / "faq.csv"
LOG_DIR = DATA_DIR / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

# ======================
# 営業用UI設定（secrets / 環境変数で上書き可）
# ======================
def _get_setting(key: str, default: str = "") -> str:
    # secrets.toml → 環境変数 → デフォルト の順で参照
    try:
        v = st.secrets.get(key, None)  # type: ignore[attr-defined]
    except Exception:
        v = None
    if v is None:
        v = os.environ.get(key)
    return str(v) if v is not None else default

COMPANY_NAME = _get_setting("COMPANY_NAME", "株式会社〇〇（デモ）")
LOGO_PATH = _get_setting("LOGO_PATH", "assets/logo.png")
CONTACT_URL = _get_setting("CONTACT_URL", "")
CONTACT_EMAIL = _get_setting("CONTACT_EMAIL", "")

def build_contact_link() -> str:
    if CONTACT_URL:
        return CONTACT_URL
    if CONTACT_EMAIL:
        # 件名などは最低限。必要なら後で増やせます。
        return f"mailto:{CONTACT_EMAIL}?subject=情シス問い合わせAI%20導入相談"
    return ""


# ======================
# 永続化設定（v13: GitHub連携対応）
# ======================
PERSIST_MODE = _get_setting("PERSIST_MODE", "local").strip().lower()
GITHUB_TOKEN = _get_setting("GITHUB_TOKEN", "").strip()
GITHUB_REPO = _get_setting("GITHUB_REPO", "").strip()  # owner/repo
GITHUB_BRANCH = _get_setting("GITHUB_BRANCH", "main").strip() or "main"
GITHUB_BASE_PATH = _get_setting("GITHUB_BASE_PATH", "streamlit_data").strip().strip("/")


def _github_persistence_enabled() -> bool:
    """GitHub永続化の有効判定。NameError回避のため内部関数を実体にする。"""
    try:
        return str(PERSIST_MODE).strip().lower() == "github" and bool(str(GITHUB_TOKEN).strip() and str(GITHUB_REPO).strip())
    except Exception:
        return False


def github_persistence_enabled() -> bool:
    """後方互換用ラッパー。画面側からは常にこの名前で呼ぶ。"""
    return _github_persistence_enabled()


def persistence_status_text() -> str:
    if _github_persistence_enabled():
        return f"GitHub永続化: ON（{GITHUB_REPO}@{GITHUB_BRANCH} / {GITHUB_BASE_PATH}）"
    return "ローカル保存のみ（Streamlit Cloud の Reboot で消える可能性があります）"


def _remote_relpath(local_path: Path) -> str:
    try:
        rel = local_path.resolve().relative_to(DATA_DIR.resolve())
    except Exception:
        rel = Path(local_path.name)
    return rel.as_posix()


def _github_api_url(rel_path: str) -> str:
    rel_path = rel_path.strip("/")
    full_path = f"{GITHUB_BASE_PATH}/{rel_path}" if GITHUB_BASE_PATH else rel_path
    return f"https://api.github.com/repos/{GITHUB_REPO}/contents/{full_path}"


def _github_headers() -> dict:
    return {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


def github_download_file(rel_path: str, local_path: Path) -> bool:
    if not _github_persistence_enabled():
        return False
    try:
        res = requests.get(_github_api_url(rel_path), headers=_github_headers(), params={"ref": GITHUB_BRANCH}, timeout=20)
        if res.status_code != 200:
            return False
        data = res.json()
        content = data.get("content", "")
        encoding = data.get("encoding", "")
        if encoding != "base64" or not content:
            return False
        raw = base64.b64decode(content)
        local_path.parent.mkdir(parents=True, exist_ok=True)
        local_path.write_bytes(raw)
        return True
    except Exception:
        return False


def github_upload_file(local_path: Path, rel_path: str | None = None, commit_message: str | None = None) -> bool:
    if not github_persistence_enabled() or not local_path.exists():
        st.error("GitHub保存の前提条件を満たしていません。PERSIST_MODE / GITHUB_TOKEN / GITHUB_REPO / 対象ファイルを確認してください。")
        return False
    rel_path = rel_path or _remote_relpath(local_path)
    try:
        existing_sha = None
        get_res = requests.get(_github_api_url(rel_path), headers=_github_headers(), params={"ref": GITHUB_BRANCH}, timeout=20)
        if get_res.status_code == 200:
            existing_sha = get_res.json().get("sha")
        elif get_res.status_code not in (200, 404):
            st.error(f"GitHub API error (GET): {get_res.status_code}")
            st.code(get_res.text)
            st.caption(f"保存先: {GITHUB_REPO}@{GITHUB_BRANCH} / {GITHUB_BASE_PATH}/{rel_path}")
            return False

        raw = local_path.read_bytes()
        payload = {
            "message": commit_message or f"Update {rel_path} from Streamlit app",
            "content": base64.b64encode(raw).decode("ascii"),
            "branch": GITHUB_BRANCH,
        }
        if existing_sha:
            payload["sha"] = existing_sha
        put_res = requests.put(_github_api_url(rel_path), headers=_github_headers(), json=payload, timeout=25)
        if put_res.status_code not in (200, 201):
            st.error(f"GitHub API error (PUT): {put_res.status_code}")
            st.code(put_res.text)
            st.caption(f"保存先: {GITHUB_REPO}@{GITHUB_BRANCH} / {GITHUB_BASE_PATH}/{rel_path}")
            return False
        return True
    except Exception as e:
        st.error("GitHub保存エラー")
        st.exception(e)
        st.caption(f"保存先: {GITHUB_REPO}@{GITHUB_BRANCH} / {GITHUB_BASE_PATH}/{rel_path}")
        return False


def github_list_dir(rel_dir: str) -> list[str]:
    if not _github_persistence_enabled():
        return []
    try:
        res = requests.get(_github_api_url(rel_dir), headers=_github_headers(), params={"ref": GITHUB_BRANCH}, timeout=20)
        if res.status_code != 200:
            return []
        data = res.json()
        if not isinstance(data, list):
            return []
        return [str(item.get("path", "")) for item in data if item.get("type") == "file"]
    except Exception:
        return []


def bootstrap_persistent_storage():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    # 初回はリポジトリ同梱のfaq.csvをruntimeへコピー
    if not FAQ_PATH.exists() and ROOT_FAQ_PATH.exists():
        try:
            FAQ_PATH.write_bytes(ROOT_FAQ_PATH.read_bytes())
        except Exception:
            pass

    # GitHub永続化が有効なら、リモートを優先して取得
    if _github_persistence_enabled():
        github_download_file("faq.csv", FAQ_PATH)
        for remote_path in github_list_dir("logs"):
            if not remote_path.endswith('.csv'):
                continue
            name = Path(remote_path).name
            github_download_file(f"logs/{name}", LOG_DIR / name)


def persist_runtime_file(local_path: Path, label: str = "data") -> bool:
    if not local_path.exists():
        return False
    if not _github_persistence_enabled():
        return True
    rel_path = _remote_relpath(local_path)
    msg = f"Update {label}: {rel_path}"
    return github_upload_file(local_path, rel_path=rel_path, commit_message=msg)


def persist_faq_now() -> bool:
    return persist_runtime_file(FAQ_PATH, label="faq")


def persist_log_now(path: Path) -> bool:
    return persist_runtime_file(path, label="log")





DEFAULT_SEARCH_THRESHOLD = 0.42
DEFAULT_SUGGEST_THRESHOLD = 0.26

SEARCH_SETTINGS_PATH = DATA_DIR / "search_settings.json"


def default_search_settings() -> dict:
    return {
        "answer_threshold": DEFAULT_SEARCH_THRESHOLD,
        "suggest_threshold": DEFAULT_SUGGEST_THRESHOLD,
    }

def _sanitize_search_settings(data: dict | None) -> dict:
    base = default_search_settings()
    src = data or {}
    try:
        answer = float(src.get("answer_threshold", base["answer_threshold"]))
    except Exception:
        answer = base["answer_threshold"]
    try:
        suggest = float(src.get("suggest_threshold", base["suggest_threshold"]))
    except Exception:
        suggest = base["suggest_threshold"]
    answer = max(0.10, min(1.20, answer))
    suggest = max(0.05, min(1.20, suggest))
    if suggest > answer:
        suggest = max(0.05, round(answer - 0.05, 2))
    return {"answer_threshold": round(answer, 2), "suggest_threshold": round(suggest, 2)}

def load_search_settings() -> dict:
    if SEARCH_SETTINGS_PATH.exists():
        try:
            data = json.loads(SEARCH_SETTINGS_PATH.read_text(encoding="utf-8"))
            return _sanitize_search_settings(data if isinstance(data, dict) else {})
        except Exception:
            return default_search_settings()
    return default_search_settings()

def save_search_settings(answer_threshold: float, suggest_threshold: float) -> tuple[bool, dict]:
    settings = _sanitize_search_settings({
        "answer_threshold": answer_threshold,
        "suggest_threshold": suggest_threshold,
    })
    try:
        SEARCH_SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
        SEARCH_SETTINGS_PATH.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
        ok = persist_runtime_file(SEARCH_SETTINGS_PATH, label="search_settings")
        return ok, settings
    except Exception:
        return False, settings

bootstrap_persistent_storage()
if _github_persistence_enabled():
    github_download_file("search_settings.json", SEARCH_SETTINGS_PATH)
SEARCH_SETTINGS = load_search_settings()


def current_search_threshold() -> float:
    try:
        return float(st.session_state.get("search_threshold", SEARCH_SETTINGS.get("answer_threshold", DEFAULT_SEARCH_THRESHOLD)))
    except Exception:
        return DEFAULT_SEARCH_THRESHOLD

def current_suggest_threshold() -> float:
    try:
        answer = current_search_threshold()
        suggest = float(st.session_state.get("suggest_threshold", SEARCH_SETTINGS.get("suggest_threshold", DEFAULT_SUGGEST_THRESHOLD)))
        return min(suggest, max(0.05, answer - 0.05))
    except Exception:
        return DEFAULT_SUGGEST_THRESHOLD



def list_log_files():
    """logsフォルダ内のCSV（nohit_*.csv）を新しい順に返す"""
    try:
        files = sorted(LOG_DIR.glob("nohit_*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
        return files
    except Exception:
        return []


def make_logs_zip(files):
    """指定されたCSVをZIP化してbytesで返す"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in files:
            try:
                zf.write(p, arcname=p.name)
            except Exception:
                pass
    buf.seek(0)
    return buf.getvalue()


def normalize_public_base_url(url: str) -> str:
    s = str(url or "").strip()
    if not s:
        return ""
    if not re.match(r"^https?://", s, flags=re.IGNORECASE):
        s = "https://" + s
    return s.rstrip("/")


def build_render_keepalive_workflow_yaml(target_url: str, cron_expr: str = "*/10 * * * *") -> str:
    target = normalize_public_base_url(target_url)
    ping_url = target or "https://example.onrender.com"
    return f"""name: Render Keep Alive

on:
  schedule:
    - cron: '{cron_expr}'
  workflow_dispatch:

jobs:
  ping:
    runs-on: ubuntu-latest
    steps:
      - name: Wake Render app
        run: |
          echo "Pinging {ping_url}"
          curl -L --fail --silent --show-error --max-time 60 "{ping_url}" > /dev/null
"""


def build_keepalive_readme_text(target_url: str, cron_expr: str = "*/10 * * * *") -> str:
    target = normalize_public_base_url(target_url)
    return f"""Render無料プランの常時起動支援

対象URL: {target or 'https://あなたのRenderURL'}
実行間隔(cron): {cron_expr}

手順:
1. このZIPを展開します。
2. .github/workflows/render-keepalive.yml を、同じGitHubリポジトリに追加します。
3. GitHub Actions を有効にします。
4. 10分ごとに Render のURL へアクセスし、スリープ復帰待ちを減らします。

補足:
- app.py 単体では、サービス自体がスリープした後に自力で自分を起こすことはできません。
- そのため、外部から定期アクセスする GitHub Actions 方式を同梱しています。
- Render の利用条件や仕様変更により、期待どおり動作しない場合があります。
"""


def build_keepalive_zip_bytes(target_url: str, cron_expr: str = "*/10 * * * *") -> bytes:
    buf = io.BytesIO()
    workflow = build_render_keepalive_workflow_yaml(target_url, cron_expr=cron_expr)
    readme = build_keepalive_readme_text(target_url, cron_expr=cron_expr)
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(".github/workflows/render-keepalive.yml", workflow)
        zf.writestr("README_keepalive.txt", readme)
    buf.seek(0)
    return buf.getvalue()

# =========================
# 管理者向けPDF（操作説明書 / 提案資料）
# =========================
def _wrap_lines_for_pdf(text: str, font_name: str, font_size: int, max_width_pt: float):
    """日本語を含む文章を、指定幅に収まるようにざっくり折り返す（ReportLab用）"""
    if not REPORTLAB_AVAILABLE:
        return [text]
    # 改行は保持しつつ、各行を折り返す
    lines = []
    for raw in str(text).splitlines() or [""]:
        buf = ""
        for ch in raw:
            if ch == "\t":
                ch = "  "
            trial = buf + ch
            try:
                w = pdfmetrics.stringWidth(trial, font_name, font_size)
            except Exception:
                # 万一フォント計測に失敗したら文字数で折る
                w = len(trial) * font_size
            if w <= max_width_pt:
                buf = trial
            else:
                if buf:
                    lines.append(buf)
                buf = ch
        lines.append(buf)
    return lines


def _pdf_draw_paragraph(c, x, y, text, font_name, font_size, max_width_pt, leading=None):
    if leading is None:
        leading = font_size * 1.35
    c.setFont(font_name, font_size)
    for line in _wrap_lines_for_pdf(text, font_name, font_size, max_width_pt):
        c.drawString(x, y, line)
        y -= leading
    return y


def _pdf_draw_title(c, title: str, subtitle: str | None = None):
    w, h = A4
    c.setFont("HeiseiKakuGo-W5", 20)
    c.drawString(20 * mm, h - 25 * mm, title)
    if subtitle:
        c.setFont("HeiseiKakuGo-W5", 11)
        c.drawString(20 * mm, h - 33 * mm, subtitle)
    # line
    c.setLineWidth(1)
    c.line(20 * mm, h - 36 * mm, w - 20 * mm, h - 36 * mm)


def _pdf_set_stroke_fill(c, stroke="#0F172A", fill="#FFFFFF"):
    c.setStrokeColor(HexColor(stroke))
    c.setFillColor(HexColor(fill))


def _pdf_draw_box(c, x, y, w, h, title, subtitle=None, fill="#FFFFFF", stroke="#CBD5E1", title_color="#0F172A"):
    _pdf_set_stroke_fill(c, stroke=stroke, fill=fill)
    c.roundRect(x, y, w, h, 8, stroke=1, fill=1)
    c.setFillColor(HexColor(title_color))
    c.setFont("HeiseiKakuGo-W5", 10)
    lines = _wrap_lines_for_pdf(title, "HeiseiKakuGo-W5", 10, w - 12)
    yy = y + h - 14
    for ln in lines[:3]:
        c.drawString(x + 6, yy, ln)
        yy -= 12
    if subtitle:
        c.setFillColor(HexColor("#475569"))
        c.setFont("HeiseiKakuGo-W5", 8)
        for ln in _wrap_lines_for_pdf(subtitle, "HeiseiKakuGo-W5", 8, w - 12)[:3]:
            c.drawString(x + 6, yy, ln)
            yy -= 10


def _pdf_draw_arrow(c, x1, y1, x2, y2, color="#64748B"):
    import math
    c.setStrokeColor(HexColor(color))
    c.setLineWidth(1.2)
    c.line(x1, y1, x2, y2)
    ang = math.atan2(y2 - y1, x2 - x1)
    ah = 6
    a1 = ang + math.pi * 0.86
    a2 = ang - math.pi * 0.86
    c.line(x2, y2, x2 + ah * math.cos(a1), y2 + ah * math.sin(a1))
    c.line(x2, y2, x2 + ah * math.cos(a2), y2 + ah * math.sin(a2))


def _pdf_draw_section_band(c, x, y, w, label, fill="#E0F2FE", text_color="#075985"):
    _pdf_set_stroke_fill(c, stroke=fill, fill=fill)
    c.roundRect(x, y - 4, w, 14, 6, stroke=0, fill=1)
    c.setFillColor(HexColor(text_color))
    c.setFont("HeiseiKakuGo-W5", 10)
    c.drawString(x + 6, y, label)


def _pdf_draw_bullet_list(c, x, y, items, max_width_pt, font_size=11, bullet_color="#0EA5E9", text_color="#0F172A", gap_after=2):
    for item in items:
        c.setFillColor(HexColor(bullet_color))
        c.setFont("HeiseiKakuGo-W5", font_size)
        c.drawString(x, y, "•")
        c.setFillColor(HexColor(text_color))
        y = _pdf_draw_paragraph(c, x + 10, y, str(item), "HeiseiKakuGo-W5", font_size, max_width_pt - 10)
        y -= gap_after
    return y


def _pdf_draw_two_column_steps(c, x, y, col_w, left_title, left_items, right_title, right_items):
    _pdf_draw_section_band(c, x, y, col_w, left_title)
    _pdf_draw_section_band(c, x + col_w + 10 * mm, y, col_w, right_title, fill="#DCFCE7", text_color="#166534")
    y_body = y - 16
    y_left = _pdf_draw_bullet_list(c, x, y_body, left_items, col_w)
    y_right = _pdf_draw_bullet_list(c, x + col_w + 10 * mm, y_body, right_items, col_w, bullet_color="#22C55E")
    return min(y_left, y_right)


def _pdf_draw_flow(c, x0, y0):
    """PDFで崩れにくい、左基準の縦フロー図。本文はこの関数では描画しない。"""
    box_w = 95 * mm
    box_h = 18 * mm
    gap = 8 * mm

    steps = [
        ("① ユーザーが質問", "チャット / おすすめ質問から入力", "#F8FAFC", "#CBD5E1", "#0F172A"),
        ("② AIがFAQを検索", "登録済みのFAQから近い回答を探す", "#EFF6FF", "#60A5FA", "#1E3A8A"),
        ("③ 回答を表示", "参考FAQもあわせて表示", "#ECFDF5", "#4ADE80", "#166534"),
        ("④ 見つからない場合", "問い合わせテンプレートを表示", "#FEF3C7", "#F59E0B", "#92400E"),
        ("⑤ 管理者がログ確認", "不足FAQを追加して次回に備える", "#DCFCE7", "#22C55E", "#166534"),
    ]

    x = x0
    y = y0

    for idx, (title, subtitle, fill, stroke, title_color) in enumerate(steps):
        _pdf_draw_box(
            c, x, y, box_w, box_h, title, subtitle,
            fill=fill, stroke=stroke, title_color=title_color
        )

        if idx < len(steps) - 1:
            arrow_x = x + box_w / 2
            _pdf_draw_arrow(c, arrow_x, y, arrow_x, y - gap)

            if idx == 2:
                c.setFillColor(HexColor("#92400E"))
                c.setFont("HeiseiKakuGo-W5", 9)
                c.drawString(x, y - gap + 1.5 * mm, "解決しない場合は、問い合わせテンプレートへ進みます")

        y -= (box_h + gap)

    return y - 2 * mm

def _pdf_draw_growth_cycle(c, x0, y0):
    """FAQ育成サイクル図"""
    box_w = 38 * mm
    box_h = 15 * mm
    gap = 9 * mm
    coords = [
        (x0, y0, "① 該当なしを記録"),
        (x0 + box_w + gap, y0, "② ログを確認"),
        (x0 + box_w + gap, y0 - box_h - gap, "③ FAQを追加"),
        (x0, y0 - box_h - gap, "④ 次回から自動回答"),
    ]
    for x, y, label in coords:
        _pdf_draw_box(c, x, y, box_w, box_h, label, fill="#F8FAFC")
    _pdf_draw_arrow(c, x0 + box_w, y0 + box_h / 2, x0 + box_w + gap, y0 + box_h / 2)
    _pdf_draw_arrow(c, x0 + box_w + gap + box_w / 2, y0, x0 + box_w + gap + box_w / 2, y0 - gap)
    _pdf_draw_arrow(c, x0 + box_w + gap, y0 - box_h - gap + box_h / 2, x0 + box_w, y0 - box_h - gap + box_h / 2)
    _pdf_draw_arrow(c, x0 + box_w / 2, y0 - box_h - gap + box_h, x0 + box_w / 2, y0 - 2)
    return y0 - box_h - gap - 14 * mm


def _pdf_draw_value_cards(c, x, y, cards, total_width):
    """カード群を重なりなく描画する。上部ラベルは外に出し、カード内は見出し+説明だけにする。"""
    gap = 5 * mm
    label_gap = 3 * mm
    card_w = (total_width - gap * (len(cards) - 1)) / len(cards)
    card_h = 26 * mm
    label_h = 6 * mm
    side_pad = 6

    for idx, (title, value, note, fill, stroke) in enumerate(cards):
        cx = x + idx * (card_w + gap)
        label_y = y - label_h
        cy = label_y - label_gap - card_h

        _pdf_set_stroke_fill(c, stroke=stroke, fill=fill)
        c.roundRect(cx, label_y, card_w, label_h, 6, stroke=1, fill=1)
        c.setFillColor(HexColor("#334155"))
        c.setFont("HeiseiKakuGo-W5", 8)
        for i, ln in enumerate(_wrap_lines_for_pdf(title, "HeiseiKakuGo-W5", 8, card_w - side_pad * 2)[:1]):
            c.drawString(cx + side_pad, label_y + label_h - 10 - i * 8, ln)

        _pdf_set_stroke_fill(c, stroke=stroke, fill="#FFFFFF")
        c.roundRect(cx, cy, card_w, card_h, 8, stroke=1, fill=1)

        c.setFillColor(HexColor("#0F172A"))
        c.setFont("HeiseiKakuGo-W5", 14)
        value_y = cy + card_h - 16
        for ln in _wrap_lines_for_pdf(value, "HeiseiKakuGo-W5", 14, card_w - side_pad * 2)[:2]:
            c.drawString(cx + side_pad, value_y, ln)
            value_y -= 15

        c.setFillColor(HexColor("#475569"))
        c.setFont("HeiseiKakuGo-W5", 8)
        note_y = cy + 12
        for ln in _wrap_lines_for_pdf(note, "HeiseiKakuGo-W5", 8, card_w - side_pad * 2)[:2]:
            c.drawString(cx + side_pad, note_y, ln)
            note_y -= 9

    return cy - 6 * mm


def generate_ops_manual_pdf() -> bytes:
    """完全版の操作説明書PDF（誰でも理解できる説明 + 図解付き）"""
    if not REPORTLAB_AVAILABLE:
        return b""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))

    w, h = A4
    margin = 18 * mm
    maxw = w - margin * 2

    # Page 1: cover
    _pdf_draw_title(c, "操作説明書_情シス問い合わせAI", "社員向け / 管理者向け / 誰でもわかる完全版")
    y = h - 52 * mm
    _pdf_draw_section_band(c, margin, y, 74 * mm, "この資料でわかること")
    y -= 18
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "このAIで何ができるのか",
            "社員がどの順番で使えばよいのか",
            "回答が見つからない時にどう動けばよいのか",
            "管理者がFAQを育てて精度を上げる方法",
        ],
        maxw,
    )
    y -= 3
    _pdf_draw_section_band(c, margin, y, 90 * mm, "最初に知っておきたいこと", fill="#DCFCE7", text_color="#166534")
    y -= 18
    y = _pdf_draw_paragraph(
        c,
        margin,
        y,
        "このシステムは、社内のITに関するよくある質問へすぐに答えるための問い合わせAIです。\n"
        "まずAIに質問し、解決できない場合だけ情シス担当者へ問い合わせる運用にすると、対応時間を減らしながら回答品質をそろえられます。",
        "HeiseiKakuGo-W5",
        11,
        maxw,
    )
    c.setFont("HeiseiKakuGo-W5", 10)
    c.setFillColor(HexColor("#64748B"))
    c.drawString(margin, 18 * mm, f"生成日: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.showPage()

    # Page 2: what it does
    _pdf_draw_title(c, "1. このAIでできること", "まずは全体像をつかむ")
    y = h - 52 * mm
    cards = [
        ("すぐに答える", "FAQ検索", "登録済みの質問と回答を探します", "#EFF6FF", "#93C5FD"),
        ("根拠を見せる", "参考FAQ表示", "どのFAQを元にしたか確認できます", "#ECFEFF", "#67E8F9"),
        ("迷った時を助ける", "テンプレ表示", "必要情報をそろえて問い合わせできます", "#FEFCE8", "#FDE68A"),
    ]
    y = _pdf_draw_value_cards(c, margin, y, cards, maxw)
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "よくある問い合わせにすぐ回答します。",
            "AIの答えとあわせて、参考にしたFAQ候補も表示します。",
            "回答が見つからない場合は、問い合わせ時に必要な項目をテンプレートで案内します。",
            "管理者はFAQファイルの入れ替え、問い合わせログの確認、PDF資料のダウンロードができます。",
            "使われ方のログを見ながら、FAQを追加して精度を上げていけます。",
        ],
        maxw,
    )
    y -= 5
    _pdf_draw_section_band(c, margin, y, 84 * mm, "利用イメージ", fill="#F8FAFC", text_color="#334155")
    y -= 18
    y = _pdf_draw_paragraph(
        c,
        margin,
        y,
        "例: 社員が『Wi-Fiがつながらない』と入力すると、AIはFAQを探して最も近い回答を表示します。\n"
        "答えが見つからない時は、端末名・利用場所・発生時刻など、情シスが確認したい情報をそろえた問い合わせテンプレートを表示します。",
        "HeiseiKakuGo-W5",
        11,
        maxw,
    )
    c.showPage()

    # Page 3: employee flow
    _pdf_draw_title(c, "2. 社員の使い方", "まずはこの順番で使います")
    y = h - 52 * mm
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "画面の入力欄に困っている内容をそのまま入力します。",
            "表示された回答を読み、必要に応じて参考FAQも確認します。",
            "その場で解決できたら完了です。",
            "解決しない時は、問い合わせテンプレートに沿って情シスへ連絡します。",
        ],
        maxw,
    )
    y -= 8

    c.setFillColor(HexColor("#0F172A"))
    c.setFont("HeiseiKakuGo-W5", 11)
    c.drawString(margin, y, "問い合わせ対応の流れ")
    y -= 80
    y = _pdf_draw_flow(c, margin, y)
    y = _pdf_draw_paragraph(
        c,
        margin,
        y,
        "見つかった回答だけで解決できる質問は、情シスへ連絡せずにその場で自己解決できます。\n"
        "回答が見つからない質問はログに残るため、後からFAQへ追加して再発防止につなげられます。",
        "HeiseiKakuGo-W5",
        10,
        maxw,
    )
    c.showPage()

    # Page 4: admin steps
    _pdf_draw_title(c, "3. 管理者の使い方", "左メニューの管理者画面で行うこと")
    y = h - 52 * mm
    col_w = (maxw - 10 * mm) / 2
    y = _pdf_draw_two_column_steps(
        c,
        margin,
        y,
        col_w,
        "毎日または週次で確認すること",
        [
            "問い合わせログ状況を見て、該当なしの増減を確認する。",
            "必要に応じてログCSVをダウンロードする。",
            "利用状況や削減時間シミュレーションを確認する。",
        ],
        "FAQを改善する時に行うこと",
        [
            "FAQをExcelでダウンロードして現在内容を確認する。",
            "不足しているQ&Aを追加したExcelまたはCSVをアップロードする。",
            "反映後、必要に応じてキャッシュクリアや再確認を行う。",
        ],
    )
    y -= 4
    _pdf_draw_section_band(c, margin, y, 92 * mm, "管理者向けPDFでできること", fill="#FEF3C7", text_color="#92400E")
    y -= 18
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "操作説明書PDF: 社員や管理者へ使い方を説明する時に利用します。",
            "提案資料PDF: 導入効果や導入ステップを説明する営業資料として利用します。",
            "導入効果レポートPDF: 実際のログを元に削減時間や削減額の試算を共有できます。",
        ],
        maxw,
        bullet_color="#F59E0B",
    )
    c.showPage()

    # Page 5: FAQ growth cycle and rules
    _pdf_draw_title(c, "4. AIを育てる運用", "使うほど精度が上がる仕組み")
    y = h - 52 * mm
    c.setFillColor(HexColor("#0F172A"))
    c.setFont("HeiseiKakuGo-W5", 11)
    c.drawString(margin, y, "FAQ改善サイクル")
    y -= 10
    bottom = _pdf_draw_growth_cycle(c, margin, y - 28 * mm)
    y = bottom + 10 * mm
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "該当なしの質問をためるだけで終わらせず、週1回など決めて確認します。",
            "同じ内容が複数回出ているものは優先してFAQへ追加します。",
            "回答文は短く、社内で実際に使う手順や連絡先まで書くと使いやすくなります。",
            "個人情報・機密情報は入力しない運用ルールを明確にしてください。",
            "FAQ更新後は、必要に応じて反映確認を行ってから社内へ案内します。",
        ],
        maxw,
    )
    y -= 6
    _pdf_draw_section_band(c, margin, y, 70 * mm, "おすすめの社内周知文", fill="#E0F2FE", text_color="#075985")
    y -= 18
    y = _pdf_draw_paragraph(
        c,
        margin,
        y,
        "『まずは情シス問い合わせAIで確認してください。回答が見つからない場合だけ、表示されたテンプレートを添えて問い合わせしてください。』\n"
        "この一文を社内ポータルやTeams/Slackの案内に載せると、自己解決の定着に役立ちます。",
        "HeiseiKakuGo-W5",
        10,
        maxw,
    )

    c.save()
    buf.seek(0)
    return buf.getvalue()


def generate_sales_proposal_pdf() -> bytes:
    """コンサルレベルの営業提案資料PDF（図解・導入効果・提案ストーリー付き）"""
    if not REPORTLAB_AVAILABLE:
        return b""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))

    w, h = A4
    margin = 18 * mm
    maxw = w - margin * 2

    # Page 1: cover
    _pdf_draw_title(c, "提案資料_情シス問い合わせAI", "社内問い合わせを減らし、対応品質をそろえるための提案書")
    y = h - 54 * mm
    _pdf_draw_section_band(c, margin, y, 78 * mm, "提案の結論")
    y -= 18
    y = _pdf_draw_paragraph(
        c,
        margin,
        y,
        "情シス問い合わせAIを導入することで、よくある問い合わせを自己解決へ誘導し、\n"
        "情シス担当者は本当に人手が必要な問い合わせへ集中できるようになります。",
        "HeiseiKakuGo-W5",
        12,
        maxw,
    )
    y -= 6
    cards = [
        ("問い合わせ削減", "一次対応を自動化", "同じ質問への繰り返し対応を減らす", "#EFF6FF", "#93C5FD"),
        ("品質平準化", "回答をそろえる", "担当者ごとの差を減らす", "#F0FDF4", "#86EFAC"),
        ("ナレッジ蓄積", "FAQが育つ", "ログから不足FAQを追加できる", "#FEFCE8", "#FDE68A"),
    ]
    y = _pdf_draw_value_cards(c, margin, y, cards, maxw)
    c.setFont("HeiseiKakuGo-W5", 10)
    c.setFillColor(HexColor("#64748B"))
    c.drawString(margin, 18 * mm, f"生成日: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.showPage()

    # Page 2: issues and solution
    _pdf_draw_title(c, "1. 現状課題と解決方針", "よくある課題をどう解決するか")
    y = h - 52 * mm
    y = _pdf_draw_two_column_steps(
        c,
        margin,
        y,
        (maxw - 10 * mm) / 2,
        "現場で起きがちな課題",
        [
            "同じ問い合わせが繰り返し発生している。",
            "担当者によって回答内容やスピードがばらつく。",
            "問い合わせ文に必要情報が不足し、切り分けに時間がかかる。",
            "FAQが更新されず、知識が属人化する。",
        ],
        "本提案の解決方針",
        [
            "まずAIに聞く導線をつくり、よくある質問を自己解決へ導く。",
            "FAQを元にした回答で、誰でも同じ案内ができる状態をつくる。",
            "見つからない場合はテンプレートで必要情報をそろえる。",
            "該当なしログからFAQを追加し、継続的に改善する。",
        ],
    )
    y -= 4
    _pdf_draw_section_band(c, margin, y, 85 * mm, "導入後の期待効果", fill="#DCFCE7", text_color="#166534")
    y -= 18
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "一次対応の自動化により、情シス担当者の負荷を下げる。",
            "回答品質を標準化し、新人や兼任担当でも案内しやすくする。",
            "問い合わせログを改善材料に変え、FAQ資産を増やす。",
        ],
        maxw,
        bullet_color="#22C55E",
    )
    c.showPage()

    # Page 3: process diagram
    _pdf_draw_title(c, "2. システムの仕組み", "問い合わせから改善までを1つの流れにする")
    y = h - 52 * mm
    c.setFillColor(HexColor("#0F172A"))
    c.setFont("HeiseiKakuGo-W5", 11)
    c.drawString(margin, y, "運用フロー図")
    y -= 10
    y = _pdf_draw_flow(c, margin, y)
    y = _pdf_draw_paragraph(
        c,
        margin,
        y,
        "ポイントは、回答できた質問だけでなく、回答できなかった質問も価値あるデータとして残ることです。\n"
        "この仕組みによって、導入直後はFAQが少なくても、使うほど回答範囲を広げられます。",
        "HeiseiKakuGo-W5",
        11,
        maxw,
    )
    y -= 4
    c.setFillColor(HexColor("#0F172A"))
    c.setFont("HeiseiKakuGo-W5", 11)
    c.drawString(margin, y, "FAQ育成サイクル")
    _pdf_draw_growth_cycle(c, margin, y - 26 * mm)
    c.showPage()

    # Page 4: ROI and model case
    _pdf_draw_title(c, "3. 導入効果の考え方", "削減時間を数字で説明する")
    y = h - 52 * mm
    cards = [
        ("モデルケース", "100件/月", "月100件の問い合わせを想定", "#F8FAFC", "#CBD5E1"),
        ("平均対応時間", "5分/件", "情シスが1件対応する平均", "#F8FAFC", "#CBD5E1"),
        ("削減時間", "約8時間/月", "100件 x 5分 = 500分", "#ECFEFF", "#67E8F9"),
    ]
    y = _pdf_draw_value_cards(c, margin, y, cards, maxw)
    y = _pdf_draw_paragraph(
        c,
        margin,
        y,
        "例として、月100件・1件5分の問い合わせがある場合、単純計算で月500分の対応時間が発生しています。\n"
        "このうち多くをAIで自己解決へ回せれば、月約8時間、年間では約96時間の削減余地があります。",
        "HeiseiKakuGo-W5",
        11,
        maxw,
    )
    y -= 6
    _pdf_draw_section_band(c, margin, y, 78 * mm, "経営層への説明ポイント", fill="#FEF3C7", text_color="#92400E")
    y -= 18
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "削減時間 = 問い合わせ件数 x 1件あたり対応時間 x AIで自己解決できる割合",
            "人件費換算を入れると、投資対効果を説明しやすくなる",
            "数値効果に加えて、回答品質の標準化や問い合わせ品質向上も副次効果として大きい",
        ],
        maxw,
        bullet_color="#F59E0B",
    )
    c.showPage()

    # Page 5: implementation plan
    _pdf_draw_title(c, "4. 導入ステップ", "最短でデモから本運用まで進める")
    y = h - 52 * mm
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "Step 1 現状確認: よくある問い合わせ、対応ルール、入力してはいけない情報を確認する。",
            "Step 2 FAQ準備: まずは30〜100件程度のFAQをCSVまたはExcelで用意する。",
            "Step 3 デモ公開: Streamlit上で社内向けに試験公開し、使い方を周知する。",
            "Step 4 ログ改善: 該当なしログを確認し、足りないFAQを追加する。",
            "Step 5 横展開: 総務、人事、経理など他部門の問い合わせへ拡張する。",
        ],
        maxw,
    )
    y -= 8
    _pdf_draw_section_band(c, margin, y, 84 * mm, "初回提案時に確認したい項目", fill="#E0F2FE", text_color="#075985")
    y -= 18
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "月間の問い合わせ件数",
            "1件あたり平均対応時間",
            "よくある問い合わせ上位10件",
            "社内で利用する連絡手段（メール / Teams / Slack など）",
            "個人情報や機密情報の取り扱いルール",
        ],
        maxw,
        bullet_color="#0EA5E9",
    )
    c.showPage()

    # Page 6: proposal closing
    _pdf_draw_title(c, "5. ご提案のまとめ", "小さく始めて、着実に育てる")
    y = h - 52 * mm
    y = _pdf_draw_paragraph(
        c,
        margin,
        y,
        "情シス問い合わせAIは、大規模なシステム刷新ではなく、既存のFAQ資産を活用しながら小さく始められる改善策です。\n"
        "まずはよくある問い合わせから対象にし、回答できなかった質問をログから追加する運用にすることで、短期間でも効果を体感しやすい構成です。",
        "HeiseiKakuGo-W5",
        11,
        maxw,
    )
    y -= 8
    _pdf_draw_section_band(c, margin, y, 64 * mm, "次のアクション", fill="#DCFCE7", text_color="#166534")
    y -= 18
    y = _pdf_draw_bullet_list(
        c,
        margin,
        y,
        [
            "問い合わせ例を10件いただければ、デモFAQを作成できます。",
            "月間件数・平均対応時間・単価がわかれば、削減効果の試算ができます。",
            "社内向け説明用として、本資料と操作説明書PDFをそのまま活用できます。",
        ],
        maxw,
        bullet_color="#22C55E",
    )

    c.save()
    buf.seek(0)
    return buf.getvalue()



# ===== PDF互換ラッパー（旧v25参照が残っていても落ちないようにする） =====
def _pdf_draw_bullet_list_safe(c, x, y, items, max_width_pt, font_size=11, bullet_color="#0EA5E9", text_color="#0F172A", gap_after=2):
    """旧PDFコード互換。既存の _pdf_draw_bullet_list が使えるならそれを優先し、
    使えない場合のみ簡易描画でフォールバックする。"""
    fn = globals().get("_pdf_draw_bullet_list")
    if callable(fn):
        return fn(c, x, y, items, max_width_pt, font_size=font_size, bullet_color=bullet_color, text_color=text_color, gap_after=gap_after)

    for item in items:
        try:
            c.setFillColor(HexColor(bullet_color))
        except Exception:
            pass
        try:
            c.setFont("HeiseiKakuGo-W5", font_size)
        except Exception:
            pass
        try:
            c.drawString(x, y, "•")
        except Exception:
            pass
        try:
            c.setFillColor(HexColor(text_color))
        except Exception:
            pass

        para = globals().get("_pdf_draw_paragraph")
        if callable(para):
            y = para(c, x + 10, y, str(item), "HeiseiKakuGo-W5", font_size, max_width_pt - 10)
        else:
            try:
                c.drawString(x + 10, y, str(item))
            except Exception:
                pass
            y -= font_size * 1.35
        y -= gap_after
    return y


def generate_sales_proposal_pdf_v25() -> bytes:
    """旧UI互換。v25名で呼ばれても現行の提案資料PDF生成へ委譲する。"""
    return generate_sales_proposal_pdf()

def render_match_bar(score: float):
    """一致度（0-1）をバーで表示"""
    try:
        v = float(score)
    except Exception:
        v = 0.0
    v = max(0.0, min(1.0, v))
    st.progress(v, text=f"一致度：{int(v*100)}%")


def count_nohit_logs(days: int = 7):
    """該当なしログ件数を集計（今日 / 過去N日 / 累計）
    文字コードやCSV崩れに強い集計にする。
    """
    files = list_log_files()
    if not files:
        return 0, 0, 0

    today_str = datetime.now().strftime("%Y%m%d")
    today_count = 0
    total_count = 0
    recent_count = 0

    today = datetime.now().date()
    recent_days = {(today - timedelta(days=i)).strftime("%Y%m%d") for i in range(days)}

    for p in files:
        name = p.name
        m = re.match(r"nohit_(\d{8})\.csv$", name)
        day = m.group(1) if m else ""
        try:
            df_log = read_csv_flexible(Path(p))
            cnt = int(len(df_log)) if df_log is not None else 0
        except Exception:
            cnt = 0

        total_count += cnt
        if day == today_str:
            today_count += cnt
        if day in recent_days:
            recent_count += cnt

    return today_count, recent_count, total_count


def read_interactions(days: int = 7) -> pd.DataFrame:
    """直近days日分のinteractionsログを結合して返す（無ければ空DF）"""
    frames = []
    for i in range(days):
        d = (datetime.now() - timedelta(days=i)).strftime("%Y%m%d")
        p = LOG_DIR / f"interactions_{d}.csv"
        if p.exists():
            try:
                frames.append(pd.read_csv(p, encoding="utf-8"))
            except Exception:
                try:
                    frames.append(pd.read_csv(p, encoding="utf-8", engine="python", on_bad_lines="skip"))
                except Exception:
                    pass
    if not frames:
        return pd.DataFrame(columns=["timestamp", "question", "matched", "best_score", "category"])

    df_all = pd.concat(frames, ignore_index=True)

    # 型整形
    if "matched" in df_all.columns:
        df_all["matched"] = pd.to_numeric(df_all["matched"], errors="coerce").fillna(0).astype(int)
    else:
        df_all["matched"] = 0
    if "best_score" in df_all.columns:
        df_all["best_score"] = pd.to_numeric(df_all["best_score"], errors="coerce").fillna(0.0)
    else:
        df_all["best_score"] = 0.0
    if "category" not in df_all.columns:
        df_all["category"] = ""

    return df_all

def format_minutes_to_hours(minutes: float) -> str:
    """分→表示用（xx分 / x.x時間）"""
    try:
        m = float(minutes)
    except Exception:
        m = 0.0
    h = m / 60.0
    if h < 1:
        return f"{int(round(m))}分"
    return f"{h:.1f}時間"
def register_jp_font():
    if not REPORTLAB_AVAILABLE:
        return "Helvetica"

    """ReportLabで日本語を表示できるフォントを登録"""
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
        return "HeiseiKakuGo-W5"
    except Exception:
        return "Helvetica"


def generate_effect_report_pdf(
    df: pd.DataFrame,
    avg_min: float,
    deflect: float,
    hourly_cost_yen: int,
    title: str = "導入効果レポート（情シス問い合わせAI）",
) -> bytes:
    """導入効果レポートPDFを生成してbytesで返す"""
    if not REPORTLAB_AVAILABLE:
        raise ModuleNotFoundError("reportlab is not installed")
    buf = io.BytesIO()
    font = register_jp_font()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    # ヘッダー
    c.setFont(font, 16)
    c.drawString(20 * mm, height - 20 * mm, title)
    c.setFont(font, 10)
    c.drawString(20 * mm, height - 27 * mm, f"作成日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # KPI計算
    total = int(len(df))
    matched = int(df["matched"].sum()) if total and "matched" in df.columns else 0
    auto_rate = (matched / total * 100.0) if total else 0.0
    saved_min = matched * float(avg_min) * float(deflect)
    saved_hours = saved_min / 60.0
    saved_yen = int(round(saved_hours * int(hourly_cost_yen))) if hourly_cost_yen else 0

    # 本文
    y = height - 45 * mm
    c.setFont(font, 12)
    c.drawString(20 * mm, y, "サマリー（今月）")
    y -= 8 * mm

    c.setFont(font, 11)
    lines = [
        f"・問い合わせ件数：{total} 件",
        f"・自動対応率：{auto_rate:.1f} %",
        f"・削減時間（推定）：{saved_hours:.1f} 時間（{int(round(saved_min))} 分）",
        f"・想定人件費削減：{saved_yen:,} 円（{hourly_cost_yen:,} 円/時間で試算）",
        f"・前提：1件あたり平均対応時間 {avg_min:.0f} 分、AIで解決できる割合 {deflect*100:.0f} %",
    ]
    for line in lines:
        c.drawString(22 * mm, y, line)
        y -= 7 * mm

    y -= 5 * mm
    c.setFont(font, 12)
    c.drawString(20 * mm, y, "補足")
    y -= 8 * mm
    c.setFont(font, 10)
    notes = [
        "・本レポートは、アプリが自動記録する利用ログ（interactions）から集計しています。",
        "・自動対応はFAQヒット（matched=1）を基準に計算しています。",
        "・削減時間／削減額は推定値です（実運用に合わせて係数調整できます）。",
    ]
    for line in notes:
        c.drawString(22 * mm, y, line)
        y -= 6 * mm

    c.showPage()
    c.save()
    return buf.getvalue()



TOP_K = 3
st.set_page_config(page_title="情シス問い合わせAI", layout="wide")


# ===== プロっぽい見た目（CSS）=====
st.markdown(
    """
<style>
/* タイトル上部の余白を確保 */
.block-container {
    padding-top: 3rem !important;
}

/* h1の高さを確保 */
h1 {
    padding-top: 0.5rem;
    line-height: 1.3 !important;
}

/* スマホ対応 */
@media (max-width: 768px) {
    .block-container {
        padding-top: 2.5rem !important;
    }
}

.block-container {padding-top: 2.0rem; padding-bottom: 10rem; max-width: 1100px;}
.hero {padding: 18px 20px; border-radius: 14px; background: linear-gradient(135deg, #0ea5e9 0%, #22c55e 100%); color: white; margin-bottom: 18px;}
.hero h1 {font-size: 34px; margin: 0 0 6px 0;}
.hero p {margin: 0; font-size: 15px; opacity: 0.95;}
.badges {margin-top: 12px; display:flex; gap:8px; flex-wrap:wrap;}
.badge {background: rgba(255,255,255,0.18); border: 1px solid rgba(255,255,255,0.25); padding: 6px 10px; border-radius: 999px; font-size: 12px;}
.card {background: #ffffff; border: 1px solid #e5e7eb; border-radius: 14px; padding: 14px 14px; box-shadow: 0 6px 20px rgba(0,0,0,0.06);}
.card h3 {margin: 0 0 8px 0; font-size: 16px;}
.small {font-size: 12px; color:#6b7280;}
.refbox {border-left: 4px solid #0ea5e9; background: #f8fafc; padding: 10px 12px; border-radius: 10px;}
.answerbox {border-left: 4px solid #22c55e; background: #f0fdf4; padding: 12px 14px; border-radius: 12px; line-height: 1.6;}

.kpi-grid {display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 12px; margin: 14px 0 18px 0;}
@media (max-width: 1100px){ .kpi-grid {grid-template-columns: repeat(2, minmax(0, 1fr));} }
.kpi {background:#ffffff; border:1px solid #e5e7eb; border-radius:16px; padding:14px 14px; box-shadow: 0 8px 26px rgba(0,0,0,0.06);}
.kpi .label {font-size:12px; color:#6b7280; margin-bottom:6px;}
.kpi .value {font-size:28px; font-weight:800; letter-spacing: -0.02em; margin:0;}
.kpi .sub {font-size:12px; color:#6b7280; margin-top:6px;}
.section-title {font-size:18px; font-weight:800; margin: 8px 0 10px 0;}
.cta-row {display:flex; gap:10px; flex-wrap:wrap; margin-top:12px;}
.cta {background: rgba(255,255,255,0.18); border: 1px solid rgba(255,255,255,0.25); padding: 10px 12px; border-radius: 12px; font-size: 13px;}

</style>
""",
    unsafe_allow_html=True,
)

# ===== 会社名 / ロゴ（左上）=====
contact_link = build_contact_link()
logo_path_obj = Path(LOGO_PATH)
col_logo, col_name, col_btn = st.columns([1, 6, 3])
with col_logo:
    if LOGO_PATH and logo_path_obj.exists():
        st.image(str(logo_path_obj), width=52)
    else:
        st.markdown("### 🏢")
with col_name:
    st.markdown(f"### {COMPANY_NAME}")
    st.caption("情シス問い合わせAI（営業デモ）")
with col_btn:
    if contact_link:
        st.link_button("📩 導入相談", contact_link, width="stretch")
    else:
        st.button("📩 導入相談（リンク未設定）", disabled=True, width="stretch")

# ===== ヒーローヘッダー =====
st.markdown(
    """
<div class="hero">
  <h1>情シス問い合わせAI</h1>
  <p>FAQ根拠付きで回答し、問い合わせ対応を削減する社内ヘルプデスクAI（RAG + LLM）</p>

  <div class="cta-row">
    <span class="cta">🎯 導入効果：問い合わせ削減 / 品質平準化 / ナレッジ蓄積</span>
    <span class="cta">🧩 既存FAQ（CSV）で即導入</span>
    <span class="cta">📄 効果レポートPDF出力</span>
  </div>

  <div class="badges">
    <span class="badge">✅ FAQ参照（根拠表示）</span>
    <span class="badge">⚡ Groq高速推論</span>
    <span class="badge">📝 ログ / 該当なし蓄積</span>
    <span class="badge">🔐 管理者でFAQ育成</span>
  </div>
</div>
""",
    unsafe_allow_html=True,
)

# ==== サイドバー ========


# ===== KPI（直近7日）=====
try:
    _df7 = read_interactions(days=7)
    if _df7 is not None and len(_df7) > 0:
        _total7 = int(len(_df7))
        _matched7 = int(_df7["matched"].sum()) if "matched" in _df7.columns else 0
        _rate7 = (_matched7 / _total7 * 100.0) if _total7 else 0.0
        _today_prefix = datetime.now().strftime("%Y-%m-%d")
        _today = _df7[_df7["timestamp"].astype(str).str.startswith(_today_prefix)]
        _total_today = int(len(_today))

        # 営業用：削減時間（推定）KPI（サイドバー入力と連動）
        _avg_min_kpi = float(st.session_state.get("avg_min", 5))
        _deflect_kpi = float(st.session_state.get("deflect", 0.7))
        _hourly_cost_kpi = int(st.session_state.get("hourly_cost", 4000))
        _saved_min7 = _matched7 * _avg_min_kpi * _deflect_kpi
        _saved_h7 = _saved_min7 / 60.0
        _saved_yen7 = int(round(_saved_h7 * _hourly_cost_kpi)) if _hourly_cost_kpi else 0

        # 営業用：KPIカード（中央に大きく）
        st.markdown('<div class="section-title">📊 直近の利用状況</div>', unsafe_allow_html=True)
        st.markdown(
            f'''
<div class="kpi-grid">
  <div class="kpi">
    <div class="label">直近7日 問い合わせ</div>
    <div class="value">{_total7}</div>
    <div class="sub">ログベース</div>
  </div>
  <div class="kpi">
    <div class="label">直近7日 自動対応率</div>
    <div class="value">{_rate7:.1f}%</div>
    <div class="sub">FAQヒット率</div>
  </div>
  <div class="kpi">
    <div class="label">直近7日 自動対応件数</div>
    <div class="value">{_matched7}</div>
    <div class="sub">自己解決に寄与</div>
  </div>
  <div class="kpi">
    <div class="label">推定削減（直近7日）</div>
    <div class="value">{_saved_h7:.1f}h</div>
    <div class="sub">約{_saved_yen7:,}円（{_hourly_cost_kpi:,}円/時間）</div>
  </div>
  <div class="kpi">
    <div class="label">今日の問い合わせ</div>
    <div class="value">{_total_today}</div>
    <div class="sub">当日分</div>
  </div>
</div>
''',
            unsafe_allow_html=True,
        )
    else:
        st.caption("（利用ログがまだありません。質問するとKPIが表示されます）")
except Exception:
    pass


with st.sidebar:
    st.markdown("### 📌 このAIでできること")
    st.markdown(
    """
    このAIは、社内のIT問い合わせを自己解決につなげるための **情シス問い合わせ支援AI** です。

    主な機能

    ・FAQデータを検索し、最も近い回答を自動表示  
    ・RAG検索により、表現が少し違う質問でも近いFAQを提示  
    ・回答の根拠となるFAQ候補と一致度を表示  
    ・該当するFAQがない場合は問い合わせテンプレートを提示  
    ・問い合わせログを自動記録し、未整備FAQを可視化  

    管理者機能

    ・FAQを **Excelでダウンロード / アップロード / 更新反映**  
    ・問い合わせログの確認  
    ・削減時間シミュレーション  
    ・導入効果レポートPDFの出力  
    ・操作説明書 / 提案資料PDFのダウンロード
    """
    )


    st.markdown("### 📈 想定効果（例）")
    st.markdown(
    """
    このAIを導入すると、次のような効果が期待できます。

    ・よくある問い合わせを自己解決できるようになる  
    ・情シス担当者の対応時間を削減できる  
    ・回答内容のばらつきを減らし、対応品質を安定化できる  
    ・新人担当者でも一定品質の対応が可能になる  
    ・問い合わせログをもとにFAQを継続改善できる  

    例（100人規模の企業）

    ・月100件の問い合わせ  
    ・1件5分対応  

    → 月 **約500分（約8時間）削減**  
    → 年間 **約96時間削減**
    """
    )


    st.markdown("### 🧭 使い方")
    st.markdown(
    """
    ① 質問を入力します  
    例  
    ・Wi-Fiがつながらない  
    ・PCが起動しない  
    ・ソフトをインストールしたい  

    ② AIがFAQを検索します  

    ③ 回答と参考FAQが表示されます  

    ④ 該当FAQがない場合  
    問い合わせテンプレートを使って情シスへ連絡できます  

    ⑤ 管理者はFAQを更新して  
    AIの回答精度を継続的に改善できます
    """
    )
    # ======================
    # ログ（該当なし）状況＆ダウンロード
    # ======================
    st.markdown("### 📊 問い合わせログ状況（該当なし）")
    t_cnt, w_cnt, total_cnt = count_nohit_logs(days=7)
    cA, cB, cC = st.columns(3)
    cA.metric("今日", t_cnt)
    cB.metric("過去7日", w_cnt)
    cC.metric("累計", total_cnt)


    # ======================
    # 削減時間シミュレーター（営業用）
    # ======================
    st.markdown("### ⏱ 削減時間シミュレーター")
    avg_min = st.slider("1件あたりの平均対応時間（分）", 1, 20, int(st.session_state.get("avg_min", 5)), help="情シスが1件対応する平均時間の目安", key="avg_min")
    deflect_pct = st.slider("AIで解決できる割合（推定）", 30, 100, int(st.session_state.get("deflect_pct", 70)), help="AI回答で自己解決に至る割合の推定", key="deflect_pct")
    deflect = deflect_pct / 100.0
    st.session_state["deflect"] = deflect

    df_int = read_interactions(days=7)
    if df_int is None or len(df_int) == 0:
        st.caption("まだ利用ログがありません（質問すると自動で蓄積します）。")
    else:
        matched_7d = int(df_int["matched"].sum()) if "matched" in df_int.columns else 0
        total_7d = int(len(df_int))
        nohit_7d = total_7d - matched_7d
        saved_min_7d = matched_7d * float(avg_min) * float(deflect)
        st.metric("推定削減（過去7日）", format_minutes_to_hours(saved_min_7d))
        st.caption(f"内訳：自動対応 {matched_7d} 件 / 該当なし {nohit_7d} 件（合計 {total_7d} 件）")

        # 今日分（timestamp先頭が YYYY-MM-DD の想定）
        try:
            today_prefix = datetime.now().strftime("%Y-%m-%d")
            df_today = df_int[df_int["timestamp"].astype(str).str.startswith(today_prefix)]
            matched_today = int(df_today["matched"].sum()) if len(df_today) else 0
            saved_min_today = matched_today * float(avg_min) * float(deflect)
            st.metric("推定削減（今日）", format_minutes_to_hours(saved_min_today))
        except Exception:
            pass

    
    # ======================
    # 見える化（グラフ）
    # ======================
    if df_int is not None and len(df_int) > 0:
        st.markdown("### 📈 見える化（直近7日）")

        df_plot = df_int.copy()
        df_plot["date"] = pd.to_datetime(df_plot["timestamp"], errors="coerce").dt.date
        daily = (
            df_plot.groupby("date", dropna=True)
            .agg(total=("question", "count"), matched=("matched", "sum"))
            .reset_index()
            .sort_values("date")
        )
        daily["auto_rate"] = (daily["matched"] / daily["total"]).replace([pd.NA, float("inf")], 0.0) * 100.0
        daily["saved_min"] = daily["matched"] * float(avg_min) * float(deflect)
        daily["saved_min_cum"] = daily["saved_min"].cumsum()

        # 1) 問い合わせ件数
        st.caption("📈 7日間の問い合わせ件数推移")
        st.line_chart(daily.set_index("date")[["total"]])

        # 2) 自動対応率
        st.caption("🧠 AI自動対応率の推移（FAQヒット率）")
        st.line_chart(daily.set_index("date")[["auto_rate"]])

        # 3) 削減時間（累計）
        st.caption("⏱ 削減時間の累計（推定）")
        st.line_chart(daily.set_index("date")[["saved_min_cum"]])

    # ======================
    # 効果レポートPDF出力（今月）
    # ======================
    st.markdown("### 📄 効果レポート（PDF）")

    if not REPORTLAB_AVAILABLE:
        st.warning("PDF出力には 'reportlab' が必要です。Streamlit Cloud の requirements.txt に 'reportlab' を追加して再デプロイしてください。")
        st.code("reportlab", language="text")
    else:
        hourly_cost = st.number_input("想定人件費（円/時間）", min_value=0, max_value=20000, value=int(st.session_state.get("hourly_cost", 4000)), step=500, key="hourly_cost")
        # st.session_state["hourly_cost"] = int(hourly_cost)

        # 今月のログを集計（最大60日読み込み→今月分だけ抽出）
        df_month_all = read_interactions(days=60)
        if df_month_all is None or len(df_month_all) == 0:
            st.caption("今月の利用ログがまだありません。質問すると自動で蓄積します。")
        else:
            try:
                ts = pd.to_datetime(df_month_all["timestamp"], errors="coerce")
                month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                df_month = df_month_all[ts >= month_start]
            except Exception:
                df_month = df_month_all

            try:
                pdf_bytes = generate_effect_report_pdf(
                    df=df_month,
                    avg_min=float(avg_min),
                    deflect=float(deflect),
                    hourly_cost_yen=int(hourly_cost),
                )
                st.download_button(
                    "📄 今月の導入効果レポートをダウンロード",
                    data=pdf_bytes,
                    file_name=f"effect_report_{datetime.now().strftime('%Y%m')}.pdf",
                    mime="application/pdf",
                    width="stretch",
                )
            except Exception as e:
                st.error(f"PDF生成でエラー: {e}")
    st.markdown("### 📥 ログ（該当なし）ダウンロード")
    log_files = list_log_files()
    if not log_files:
        st.caption("まだログはありません。")
    else:
        latest = log_files[0]
        try:
            latest_bytes = latest.read_bytes()
        except Exception:
            latest_bytes = b""
        st.download_button(
            "⬇ 最新ログCSVをダウンロード",
            data=latest_bytes,
            file_name=latest.name,
            mime="text/csv",
            width="stretch",
        )

        zip_bytes = make_logs_zip(log_files)
        st.download_button(
            "⬇ ログをZIPでまとめてDL",
            data=zip_bytes,
            file_name="nohit_logs.zip",
            mime="application/zip",
            width="stretch",
        )

        with st.expander("ログ一覧を見る"):
            for p in log_files[:20]:
                st.write(f"• {p.name}")


# ======================
# FAQロード（落ちやすい箇所を全てガード）
# ======================

FULLWIDTH_TRANS = str.maketrans({
    "０": "0", "１": "1", "２": "2", "３": "3", "４": "4",
    "５": "5", "６": "6", "７": "7", "８": "8", "９": "9",
    "ａ": "a", "ｂ": "b", "ｃ": "c", "ｄ": "d", "ｅ": "e", "ｆ": "f", "ｇ": "g",
    "ｈ": "h", "ｉ": "i", "ｊ": "j", "ｋ": "k", "ｌ": "l", "ｍ": "m", "ｎ": "n",
    "ｏ": "o", "ｐ": "p", "ｑ": "q", "ｒ": "r", "ｓ": "s", "ｔ": "t", "ｕ": "u",
    "ｖ": "v", "ｗ": "w", "ｘ": "x", "ｙ": "y", "ｚ": "z",
    "Ａ": "a", "Ｂ": "b", "Ｃ": "c", "Ｄ": "d", "Ｅ": "e", "Ｆ": "f", "Ｇ": "g",
    "Ｈ": "h", "Ｉ": "i", "Ｊ": "j", "Ｋ": "k", "Ｌ": "l", "Ｍ": "m", "Ｎ": "n",
    "Ｏ": "o", "Ｐ": "p", "Ｑ": "q", "Ｒ": "r", "Ｓ": "s", "Ｔ": "t", "Ｕ": "u",
    "Ｖ": "v", "Ｗ": "w", "Ｘ": "x", "Ｙ": "y", "Ｚ": "z",
})

CANONICAL_PATTERNS = [
    (r"デスクトップパソコン|デスクトップｐｃ|desktop\s*pc", "デスクトップpc"),
    (r"ノートパソコン|ラップトップ", "ノートpc"),
    (r"パーソナルコンピュータ|パソコン|ピーシー|ｐｃ|pc端末", "pc"),
    (r"コンピューター", "コンピュータ"),
    (r"無線lan|wi-?fi|wifi|ワイファイ", "wifi"),
    (r"ｖｐｎ|ぶいぴーえぬ|vpn接続", "vpn"),
    (r"サインイン", "ログイン"),
    (r"サインアウト", "ログアウト"),
    (r"パスコード|passcode", "パスワード"),
    (r"pw", "パスワード"),
    (r"立ち上がらない|起ち上がらない|立ちあがらない|起ちあがらない", "起動しない"),
    (r"電源がつかない|電源が付かない|電源がはいらない", "電源が入らない"),
    (r"ログイン出来ない|ログインできません", "ログインできない"),
    (r"接続できない|接続できません|接続出来ない|接続出来ません|接続しない|接続されない|つながりません|繋がりません|繋がらない|繋げない|つなげない", "つながらない"),
    (r"利用できない|使用できない", "使えない"),
    (r"開けない", "起動しない"),
    (r"印字できない|プリントできない", "印刷できない"),
    (r"メール送れない", "メールが送信できない"),
    (r"メール受け取れない", "メールが受信できない"),
    (r"認証に失敗|認証エラー", "認証できない"),
    (r"ロックされた|凍結された", "ロックされた"),
]

CONCEPT_ALIASES = {
    "vpn": ["vpn", "リモートアクセス", "社外接続"],
    "network": ["ネットワーク", "wifi", "lan", "通信", "internet", "インターネット"],
    "login": ["ログイン", "サインイン", "認証", "アカウント"],
    "password": ["パスワード", "password", "pw"],
    "boot": ["起動", "立ち上が", "立上", "電源", "シャットダウン", "再起動"],
    "mail": ["メール", "outlook", "受信", "送信"],
    "print": ["印刷", "プリンタ", "printer", "print"],
    "lock": ["ロック", "凍結", "無効", "停止"],
    "error": ["エラー", "失敗", "不具合", "異常", "障害"],
    "cannot": ["できない", "できません", "使えない", "つながらない", "入らない", "起動しない"],
}


def normalize_search_text(text: str) -> str:
    """FAQ検索用の正規化。表記ゆれ・同義表現を寄せて意味検索を強化する。"""
    s = str(text or "").strip().lower()
    if not s:
        return ""

    s = s.translate(FULLWIDTH_TRANS)
    for pattern, repl in CANONICAL_PATTERNS:
        s = re.sub(pattern, repl, s)

    s = re.sub(r"([^a-z0-9])pc([^a-z0-9])", r"\1 pc \2", f" {s} ")
    s = re.sub(r"[\/／・,、。．・:：;；\-ー_（）()\[\]{}『』「」\"'`]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def extract_search_tokens(text: str) -> set[str]:
    """日本語FAQ検索向けの軽量トークン抽出。"""
    s = normalize_search_text(text)
    if not s:
        return set()

    tokens = set()
    for part in s.split():
        part = part.strip()
        if part:
            tokens.add(part)

    for tok in re.findall(r"[a-z0-9]+|[\u3040-\u30ff\u4e00-\u9fff]{2,}", s):
        tokens.add(tok)

    split_hints = ["できない", "つながらない", "起動しない", "ログイン", "パスワード", "電源", "印刷", "メール", "アカウント", "vpn", "wifi"]
    for tok in list(tokens):
        for hint in split_hints:
            if hint in tok and tok != hint:
                tokens.add(hint)
                remain = tok.replace(hint, " ").strip()
                if len(remain) >= 2:
                    tokens.add(remain)

    return {t for t in tokens if t}


def extract_concepts(text: str) -> set[str]:
    s = normalize_search_text(text)
    found = set()
    for concept, aliases in CONCEPT_ALIASES.items():
        if any(alias in s for alias in aliases):
            found.add(concept)
    return found


# 文字n-gramも混ぜて、日本語の部分一致に強くする
WORD_VECTORIZER = TfidfVectorizer(ngram_range=(1, 2))
CHAR_VECTORIZER = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4))
SENTENCE_MODEL_NAME = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"

def _load_sentence_transformer_model():
    if not SENTENCE_TRANSFORMERS_AVAILABLE or SentenceTransformer is None:
        return None
    cache_key = "_st_sentence_transformer_model"
    try:
        model = st.session_state.get(cache_key)
        if model is not None:
            return model
    except Exception:
        model = None
    try:
        model = SentenceTransformer(SENTENCE_MODEL_NAME)
        try:
            st.session_state[cache_key] = model
        except Exception:
            pass
        return model
    except Exception:
        return None


def _build_sentence_embeddings(model, texts: list[str]):
    if model is None or not texts:
        return None
    try:
        return model.encode(texts, normalize_embeddings=True, show_progress_bar=False)
    except Exception:
        return None


def _search_with_sentence_transformers(query_norm: str, faq_embeddings) -> list[float] | None:
    if not query_norm or faq_embeddings is None:
        return None
    model = _load_sentence_transformer_model()
    if model is None:
        return None
    try:
        q_emb = model.encode([query_norm], normalize_embeddings=True, show_progress_bar=False)
        sims_sem = cosine_similarity(q_emb, faq_embeddings).flatten()
        return sims_sem.tolist()
    except Exception:
        return None


def load_faq_index(faq_path: Path):
    if not faq_path.exists():
        empty = pd.DataFrame(columns=["question", "answer", "category"])
        return empty, None, None, None, None, None

    try:
        df = normalize_faq_columns(read_csv_flexible(faq_path))
    except Exception:
        empty = pd.DataFrame(columns=["question", "answer", "category"])
        return empty, None, None, None, None, None

    df["question"] = df["question"].fillna("").astype(str)
    df["answer"] = df["answer"].fillna("").astype(str)
    df["category"] = df["category"].fillna("").astype(str)

    if len(df) == 0:
        return df, None, None, None, None, None

    df["question_norm"] = df["question"].apply(normalize_search_text)
    df["answer_norm"] = df["answer"].apply(normalize_search_text)
    df["qa_text"] = (df["question"] + " / " + df["answer"]).astype(str)
    df["qa_text_norm"] = (df["question_norm"] + " / " + df["answer_norm"]).astype(str)
    df["search_tokens"] = df["qa_text_norm"].apply(extract_search_tokens)
    df["search_concepts"] = df["qa_text_norm"].apply(extract_concepts)

    try:
        word_vectorizer = WORD_VECTORIZER
        char_vectorizer = CHAR_VECTORIZER
        X_word = word_vectorizer.fit_transform(df["qa_text_norm"])
        X_char = char_vectorizer.fit_transform(df["qa_text_norm"])
    except Exception:
        return df, None, None, None, None, None

    faq_embeddings = None
    if SENTENCE_TRANSFORMERS_AVAILABLE:
        model = _load_sentence_transformer_model()
        faq_embeddings = _build_sentence_embeddings(model, df["qa_text_norm"].tolist())

    return df, word_vectorizer, X_word, char_vectorizer, X_char, faq_embeddings


df, vectorizer, X, char_vectorizer, X_char, faq_embeddings = load_faq_index(FAQ_PATH)

if df is None or len(df) == 0 or vectorizer is None or X is None or char_vectorizer is None or X_char is None:
    st.warning("faq.csv が未配置/空/不正のため、FAQ検索は無効です。まず faq.csv を配置してください。")


def retrieve_faq(query: str):
    if not query:
        return []
    if vectorizer is None or X is None or char_vectorizer is None or X_char is None or df is None or len(df) == 0:
        return []
    try:
        query_norm = normalize_search_text(query)
        qv_word = vectorizer.transform([query_norm])
        qv_char = char_vectorizer.transform([query_norm])
        sims_word = cosine_similarity(qv_word, X).flatten()
        sims_char = cosine_similarity(qv_char, X_char).flatten()
        if sims_word.size == 0 or sims_char.size == 0:
            return []

        # 単語一致と文字部分一致を合成して、表記ゆれに強くする
        sims = (sims_word * 0.52) + (sims_char * 0.48)

        # sentence-transformers による意味検索を追加（未導入時は自動スキップ）
        sims_sem = _search_with_sentence_transformers(query_norm, faq_embeddings)
        if sims_sem is not None and len(sims_sem) == len(sims):
            sims = sims + (pd.Series(sims_sem).fillna(0.0).to_numpy() * 0.35)

        # 完全一致に近い正規化結果は少し加点
        exact_bonus = (df["question_norm"] == query_norm).astype(float).to_numpy() * 0.22
        contains_bonus = df["question_norm"].apply(
            lambda x: 0.10 if query_norm and (query_norm in x or x in query_norm) else 0.0
        ).to_numpy()

        # トークン重複率で加点
        q_tokens = extract_search_tokens(query_norm)
        token_bonus = df["search_tokens"].apply(
            lambda toks: (0.18 * len(q_tokens & set(toks)) / max(1, len(q_tokens))) if q_tokens else 0.0
        ).to_numpy()

        # 概念一致で加点（vpn / login / boot など）
        q_concepts = extract_concepts(query_norm)
        concept_bonus = df["search_concepts"].apply(
            lambda cs: (0.22 * len(q_concepts & set(cs)) / max(1, len(q_concepts))) if q_concepts else 0.0
        ).to_numpy()

        prefix_bonus = df["question_norm"].apply(
            lambda x: 0.06 if query_norm and str(x).startswith(query_norm[: min(8, len(query_norm))]) else 0.0
        ).to_numpy()

        sims = sims + exact_bonus + contains_bonus + token_bonus + concept_bonus + prefix_bonus

        idxs = sims.argsort()[::-1][:3]
        return [(df.iloc[i], float(sims[i])) for i in idxs if float(sims[i]) > 0]
    except Exception:
        return []


def build_prompt(user_q: str, hits):
    context_parts = []
    for i, (row, score) in enumerate(hits, 1):
        q = str(row.get("question", ""))
        a = str(row.get("answer", ""))
        context_parts.append(f"\n[FAQ{i}]\nQ:{q}\nA:{a}\n")
    context = "".join(context_parts)

    return f"""
あなたは社内の情シス担当です。
必ず日本語のみで回答してください。
丁寧で簡潔に、手順は箇条書きで書いてください。

参照FAQ:
{context}

質問:
{user_q}
"""


def nohit_template():
    return """
FAQに該当がありませんでした。

情シスへお問い合わせの際は、以下の情報を添えてください：

・何ができないか（具体的な操作）
・エラー画面のスクリーンショット
・発生時刻
・利用場所（社内 / 社外）
・ネットワーク（Wi-Fi / VPN）
・端末（Windows / Mac）
・影響範囲（自分のみ / 他の人も）

※これらを共有いただくと対応が早くなります。
※このAIは御社の運用に合わせてカスタマイズ可能です。
""".strip()



def _ensure_nohit_schema(path: Path):
    """既存nohit CSVが旧形式（timestamp,questionのみ）でも、新スキーマに移行する。"""
    cols = ["timestamp", "question", "device", "location", "network", "error_text", "impact", "channel"]
    if not path.exists():
        return cols

    try:
        # 先頭行だけ見る（軽量）
        with path.open("r", encoding="utf-8", errors="ignore") as f:
            header = f.readline().strip()
        header_cols = [h.strip() for h in header.split(",")] if header else []
    except Exception:
        header_cols = []

    # すでに新スキーマなら何もしない
    if set(cols).issubset(set(header_cols)):
        return header_cols

    # 移行：既存を読み取り→新ヘッダで書き直し
    try:
        old_df = read_csv_flexible(path)
        if old_df is None:
            old_df = pd.DataFrame()
    except Exception:
        old_df = pd.DataFrame()

    if len(old_df) == 0:
        # 空なら新ヘッダで作り直す
        with path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(cols)
        return cols

    qcol = pick_question_column(old_df.columns) or ("question" if "question" in old_df.columns else None)
    tcol = "timestamp" if "timestamp" in old_df.columns else None

    rows = []
    for _, r in old_df.iterrows():
        ts = str(r.get(tcol, "")).strip() if tcol else ""
        q = str(r.get(qcol, "")).strip() if qcol else ""
        if not q:
            continue
        rows.append([ts, q, "", "", "", "", "", ""])

    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        w.writerows(rows)

    return cols


def log_nohit(question: str, extra: dict | None = None) -> str:
    """該当なしログを追記して、記録したtimestamp（秒）を返す。"""
    if not question:
        return ""
    extra = extra or {}
    day = datetime.now().strftime("%Y%m%d")
    path = LOG_DIR / f"nohit_{day}.csv"
    cols = _ensure_nohit_schema(path)

    ts = datetime.now().isoformat(timespec="seconds")
    row = {
        "timestamp": ts,
        "question": question,
        "device": extra.get("device", ""),
        "location": extra.get("location", ""),
        "network": extra.get("network", ""),
        "error_text": extra.get("error_text", ""),
        "impact": extra.get("impact", ""),
        "channel": extra.get("channel", "web"),
    }

    try:
        is_new = not path.exists()
        with path.open("a", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            if is_new:
                w.writerow(cols)
            w.writerow([row.get(c, "") for c in cols])
        persist_log_now(path)
    except Exception:
        pass
    return ts


def update_nohit_record(day: str, timestamp: str, question: str, extra: dict) -> bool:
    """同じday/timestamp/question の行があれば更新。無ければ追記。"""
    if not day or not timestamp or not question:
        return False
    path = LOG_DIR / f"nohit_{day}.csv"
    cols = _ensure_nohit_schema(path)

    try:
        df_log = read_csv_flexible(path)
        if df_log is None:
            df_log = pd.DataFrame(columns=cols)
    except Exception:
        df_log = pd.DataFrame(columns=cols)

    # 必須列を揃える
    for c in cols:
        if c not in df_log.columns:
            df_log[c] = ""

    # 既存行を更新（最初の一致）
    mask = (df_log["timestamp"].astype(str) == str(timestamp)) & (df_log["question"].astype(str) == str(question))
    idxs = df_log.index[mask].tolist()
    if idxs:
        i = idxs[0]
        for k, v in (extra or {}).items():
            if k in df_log.columns:
                df_log.at[i, k] = v
        df_log.at[i, "channel"] = extra.get("channel", df_log.at[i, "channel"] or "web")
    else:
        # 無ければ追記
        row = {c: "" for c in cols}
        row["timestamp"] = timestamp
        row["question"] = question
        for k, v in (extra or {}).items():
            if k in row:
                row[k] = v
        if not row.get("channel"):
            row["channel"] = "web"
        df_log = pd.concat([df_log, pd.DataFrame([row])], ignore_index=True)

    # UTF-8で書き戻す（Excel対応ならutf-8-sigでもOK）
    try:
        with path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(cols)
            for _, r in df_log[cols].iterrows():
                w.writerow([str(r.get(c, "")) for c in cols])
        persist_log_now(path)
        return True
    except Exception:
        return False


def seed_nohit_questions(n: int = 20) -> int:
    """本番前のデモ用：情シス定番のnohit質問を今日のログに追加する。"""
    seeds = [
        "VPNにつながらない", "Outlookの送受信ができない", "Teamsにログインできない", "パスワードを忘れた",
        "アカウントがロックされた", "共有フォルダにアクセスできない", "プリンタが印刷できない", "Wi-Fiが頻繁に切れる",
        "PCが重い", "PCが固まる", "Excelが起動しない", "Excelがフリーズする", "OneDriveが同期しない",
        "メール添付ファイルが開けない", "二段階認証が通らない", "カメラが映らない", "マイクが認識されない",
        "ソフトのインストール申請方法が分からない", "Windows更新が終わらない", "画面が真っ黒になる",
    ]
    added = 0
    for q in seeds[:n]:
        ts = log_nohit(q, {"channel": "seed"})
        if ts:
            added += 1
    return added



def log_interaction(question: str, matched: bool, best_score: float, category: str):
    """全ての質問をログ化（削減時間の見える化用）: logs/interactions_YYYYMMDD.csv"""
    if not question:
        return
    day = datetime.now().strftime("%Y%m%d")
    path = LOG_DIR / f"interactions_{day}.csv"
    try:
        is_new = not path.exists()
        with path.open("a", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            if is_new:
                w.writerow(["timestamp", "question", "matched", "best_score", "category"])
            w.writerow([datetime.now().isoformat(timespec="seconds"), question, int(bool(matched)), float(best_score), category or ""])
        persist_log_now(path)
    except Exception:
        pass




import json

def normalize_question(q: str) -> str:
    q = (q or "").strip().lower()
    q = re.sub(r"\s+", " ", q)
    # 日本語も含めて記号類をざっくり除去
    q = re.sub(r"[\u3000\s\t\r\n]+", " ", q)
    q = re.sub(r"[!！?？。、,.，:：;；\-_=+~`'\"()（）\[\]{}<>＜＞/\\|@#%^&*]", "", q)
    return q.strip()



def load_nohit_questions_from_logs(files, max_questions: int = 100) -> list[str]:
    """nohit_*.csv から質問を収集（新しいログから優先）。文字コード/カラム揺れに強く読む。"""
    questions: list[str] = []
    seen: set[str] = set()
    for p in files:
        try:
            _df = read_csv_flexible(Path(p))
            if _df is None or len(_df) == 0:
                continue

            qcol = pick_question_column(_df.columns)
            if not qcol:
                continue

            for q in _df[qcol].fillna("").astype(str).tolist():
                nq = normalize_question(q)
                if not nq:
                    continue
                if nq in seen:
                    continue
                seen.add(nq)
                questions.append(q.strip())
                if len(questions) >= max_questions:
                    return questions
        except Exception:
            continue
    return questions



def generate_faq_candidates(nohit_questions: list[str], n_items: int = 8) -> pd.DataFrame:
    """該当なしログからFAQ案を生成してDataFrameで返す（category/question/answer）。"""
    if not nohit_questions:
        return pd.DataFrame(columns=["category", "question", "answer"])

    # 入力が長すぎると落ちるので上限
    max_in = min(len(nohit_questions), 80)
    sample = nohit_questions[:max_in]
    examples = "\n".join([f"- {q}" for q in sample])

    prompt = f"""あなたは社内情シスのベテラン担当です。
以下は『FAQに該当なし』として蓄積された、社員からの問い合わせ例です。

【目的】
この問い合わせ例から、社内で使えるFAQ（Q&A）を {n_items} 件作成してください。

【要件】
- 日本語のみ
- 1件ごとに category / question / answer を作る
- answer は手順を箇条書きで（3〜7行）
- 個人情報や会社固有の秘密情報は作らない
- できるだけ汎用的（どの会社でも通用）に
- 出力は必ずJSONのみ（前後に説明文を入れない）
- コードブロック ``` は使わない

【出力JSON形式】
[
  {{"category":"VPN", "question":"...", "answer":"- ...\n- ..."}},
  ...
]

【問い合わせ例】
{examples}
"""

    out = llm_chat(
        [
            {"role": "system", "content": "あなたは情シスのFAQ作成者です。出力はJSONのみ。"},
            {"role": "user", "content": prompt},
        ]
    )

    out_text = out if isinstance(out, str) else str(out)

    json_text = extract_json_array(out_text) or out_text.strip()

    try:
        data = json.loads(json_text)
        if not isinstance(data, list):
            raise ValueError("JSON is not a list")
    except Exception:
        return pd.DataFrame(columns=["category", "question", "answer"])

    rows = []
    for item in data:
        if not isinstance(item, dict):
            continue
        cat = str(item.get("category", "")).strip()
        q = str(item.get("question", "")).strip()
        a = str(item.get("answer", "")).strip()
        if not q or not a:
            continue
        rows.append({"category": cat, "question": q, "answer": a})

    return pd.DataFrame(rows, columns=["category", "question", "answer"])

def append_faq_csv(faq_path: Path, new_df: pd.DataFrame) -> int:
    """faq.csv に追記。重複（question）をざっくり除外して追記件数を返す。"""
    if new_df is None or len(new_df) == 0:
        return 0

    # 必須列を揃える
    for col in ["question", "answer", "category"]:
        if col not in new_df.columns:
            new_df[col] = ""

    new_df = new_df[["question", "answer", "category"]].copy()
    new_df["question"] = new_df["question"].fillna("").astype(str).str.strip()
    new_df["answer"] = new_df["answer"].fillna("").astype(str).str.strip()
    new_df["category"] = new_df["category"].fillna("").astype(str).str.strip()
    new_df = new_df[(new_df["question"] != "") & (new_df["answer"] != "")]
    if len(new_df) == 0:
        return 0

    # 既存読み込み
    if faq_path.exists():
        try:
            exist = normalize_faq_columns(read_csv_flexible(faq_path))
        except Exception:
            exist = pd.DataFrame(columns=["question", "answer", "category"])
    else:
        exist = pd.DataFrame(columns=["question", "answer", "category"])

    exist_q = set(normalize_question(x) for x in exist.get("question", pd.Series(dtype=str)).fillna("").astype(str).tolist())

    rows = []
    for _, r in new_df.iterrows():
        nq = normalize_question(str(r.get("question", "")))
        if not nq:
            continue
        if nq in exist_q:
            continue
        exist_q.add(nq)
        rows.append([r["question"], r["answer"], r.get("category", "")])

    if not rows:
        return 0

    is_new = not faq_path.exists()
    with faq_path.open("a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if is_new:
            w.writerow(["question", "answer", "category"])
        w.writerows(rows)

    persist_faq_now()
    return len(rows)



def generate_slack_bot_zip_bytes():
    """Slack Bot 完全版コード一式をZIPで返す。既存アプリ本体とは分離し、Render等に別サービスとして配置する想定。"""
    import textwrap

    render_base = "https://your-render-url.onrender.com"

    slack_bot_py = textwrap.dedent("""    import os
    import hmac
    import hashlib
    import time
    import json
    from typing import Any

    import requests
    from flask import Flask, request, jsonify

    app = Flask(__name__)

    HELP_DESK_ASK_URL = os.getenv("HELPDESK_ASK_URL", "https://your-streamlit-or-api-url/ask")
    SLACK_SIGNING_SECRET = os.getenv("SLACK_SIGNING_SECRET", "")
    REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "30"))


    def verify_slack_request(req) -> bool:
        if not SLACK_SIGNING_SECRET:
            return True

        timestamp = req.headers.get("X-Slack-Request-Timestamp", "")
        slack_signature = req.headers.get("X-Slack-Signature", "")
        if not timestamp or not slack_signature:
            return False

        try:
            if abs(time.time() - int(timestamp)) > 60 * 5:
                return False
        except Exception:
            return False

        body = req.get_data(as_text=True)
        basestring = f"v0:{timestamp}:{body}"
        my_signature = "v0=" + hmac.new(
            SLACK_SIGNING_SECRET.encode("utf-8"),
            basestring.encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()

        return hmac.compare_digest(my_signature, slack_signature)


    def ask_helpdesk(question: str, user_name: str = "", channel_name: str = "") -> str:
        payload: dict[str, Any] = {
            "question": question,
            "source": "slack",
            "user_name": user_name,
            "channel_name": channel_name,
        }

        r = requests.post(
            HELP_DESK_ASK_URL,
            json=payload,
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()

        data = r.json()
        answer = data.get("answer") or data.get("message") or "回答を取得できませんでした。"
        return str(answer)


    @app.get("/")
    def health() -> tuple[str, int]:
        return "ok", 200


    @app.post("/slack/command")
    def slack_command():
        if not verify_slack_request(request):
            return jsonify({"text": "署名検証に失敗しました。"}), 403

        question = (request.form.get("text") or "").strip()
        user_name = request.form.get("user_name", "")
        channel_name = request.form.get("channel_name", "")

        if not question:
            return jsonify({
                "response_type": "ephemeral",
                "text": "質問文を入れてください。例: /helpdesk VPNがつながらない",
            })

        try:
            answer = ask_helpdesk(question, user_name=user_name, channel_name=channel_name)
            return jsonify({
                "response_type": "in_channel",
                "text": f"*質問:* {question}\n*回答:* {answer}",
            })
        except Exception as e:
            return jsonify({
                "response_type": "ephemeral",
                "text": f"問い合わせAIへの接続に失敗しました: {e}",
            }), 500


    @app.post("/slack/events")
    def slack_events():
        if not verify_slack_request(request):
            return jsonify({"error": "invalid signature"}), 403

        data = request.get_json(silent=True) or {}

        if data.get("type") == "url_verification":
            return jsonify({"challenge": data.get("challenge", "")})

        event = data.get("event", {})
        if event.get("type") == "app_mention":
            text = event.get("text", "")
            return jsonify({"ok": True, "note": f"mention received: {text}"})

        return jsonify({"ok": True})


    if __name__ == "__main__":
        port = int(os.getenv("PORT", "3000"))
        app.run(host="0.0.0.0", port=port)
    """).strip() + "\n"

    requirements_txt = textwrap.dedent("""    flask==3.0.3
    requests==2.32.3
    gunicorn==22.0.0
    """).strip() + "\n"

    render_yaml = textwrap.dedent("""    services:
      - type: web
        name: slack-helpdesk-bot
        env: python
        plan: free
        buildCommand: pip install -r requirements.txt
        startCommand: gunicorn slack_bot:app
        autoDeploy: true
    """).strip() + "\n"

    env_example = textwrap.dedent(f"""    HELPDESK_ASK_URL={render_base}/ask
    SLACK_SIGNING_SECRET=your_signing_secret
    REQUEST_TIMEOUT=30
    """).strip() + "\n"

    readme_md = textwrap.dedent("""    # Slack Helpdesk Bot 完全版

    このZIPは、既存の Streamlit アプリ本体とは別サービスとして Render に配置する想定です。

    ## 含まれるファイル
    - `slack_bot.py` : Slack Slash Command / Events 受信用 Flask アプリ
    - `requirements.txt` : 必要ライブラリ
    - `render.yaml` : Render デプロイ設定例
    - `.env.example` : 環境変数の雛形

    ## Slack 側設定
    1. Slack API で App を作成
    2. Slash Commands に `/helpdesk` を追加
    3. Request URL に `https://あなたのRenderURL/slack/command` を設定
    4. Event Subscriptions を使う場合は `https://あなたのRenderURL/slack/events` を設定
    5. Signing Secret を Render の環境変数 `SLACK_SIGNING_SECRET` に設定

    ## Render 側設定
    1. このZIPを GitHub リポジトリに配置
    2. Render で New + → Web Service
    3. Build Command: `pip install -r requirements.txt`
    4. Start Command: `gunicorn slack_bot:app`
    5. `HELPDESK_ASK_URL` に既存問い合わせAIの API URL を設定

    ## 重要
    既存の Streamlit アプリに `/ask` API が無い場合は、別途 API 追加が必要です。
    まずは Slack Bot コード一式を先に配布し、後から API 側を接続しても構いません。
    """).strip() + "\n"

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("slack_bot.py", slack_bot_py)
        zf.writestr("requirements.txt", requirements_txt)
        zf.writestr("render.yaml", render_yaml)
        zf.writestr(".env.example", env_example)
        zf.writestr("README.md", readme_md)

    buf.seek(0)
    return buf.getvalue()


# ======================
# 管理者ログイン
# ======================
if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

with st.sidebar:
    st.markdown("## 🛠 管理者")
    if not st.session_state.is_admin:
        pwd = st.text_input("管理者パスワード", type="password")
        if st.button("ログイン"):
            if check_password(pwd):
                st.session_state.is_admin = True
                st.success("ログイン成功")
                st.rerun()
            else:
                st.error("パスワードが違います")
    else:
        st.success("ログイン中")
        if st.button("ログアウト"):
            st.session_state.is_admin = False
            st.rerun()

        with st.expander("💾 永続化ステータス（v13）", expanded=True):
            st.caption(persistence_status_text())
            st.code("""
# Streamlit Cloud secrets.toml の例
PERSIST_MODE = "github"
GITHUB_REPO = "owner/repo"
GITHUB_BRANCH = "main"
GITHUB_BASE_PATH = "streamlit_data"
GITHUB_TOKEN = "ghp_xxx"
""".strip(), language="toml")
            col_sync1, col_sync2 = st.columns(2)
            with col_sync1:
                if st.button("📥 GitHubからFAQ再読込", width="stretch", disabled=not _github_persistence_enabled()):
                    ok = github_download_file("faq.csv", FAQ_PATH)
                    if ok:
                        try:
                            load_faq_index.clear()
                        except Exception:
                            pass
                        st.success("GitHub上の faq.csv を再読込しました。")
                        st.rerun()
                    else:
                        st.warning("GitHubからFAQを取得できませんでした。設定を確認してください。")
            with col_sync2:
                if st.button("📤 FAQをGitHubへ保存", width="stretch", disabled=not _github_persistence_enabled()):
                    ok = persist_faq_now()
                    if ok:
                        st.success("faq.csv を GitHub に保存しました。")
                    else:
                        st.warning("GitHubへの保存に失敗しました。設定を確認してください。")

        with st.expander("🎯 検索精度設定（v14）", expanded=True):
            st.caption("しきい値を管理者が調整できます。高いほど厳密一致寄り、低いほど似た質問も拾いやすくなります。")
            st.info(f"現在値：自動回答 {current_search_threshold():.2f} / 候補表示 {current_suggest_threshold():.2f}")

            admin_answer_threshold = st.slider(
                "自動回答しきい値",
                min_value=0.10,
                max_value=1.20,
                value=float(st.session_state.get("search_threshold", DEFAULT_SEARCH_THRESHOLD)),
                step=0.01,
                key="admin_answer_threshold_slider",
                help="この値以上なら通常回答します。",
            )
            max_suggest_value = max(0.05, round(admin_answer_threshold - 0.05, 2))
            suggest_default = min(float(st.session_state.get("suggest_threshold", DEFAULT_SUGGEST_THRESHOLD)), max_suggest_value)
            admin_suggest_threshold = st.slider(
                "候補表示しきい値",
                min_value=0.05,
                max_value=max_suggest_value,
                value=suggest_default,
                step=0.01,
                key="admin_suggest_threshold_slider",
                help="この値以上かつ自動回答しきい値未満なら『近いFAQ候補』として表示します。",
            )

            st.markdown("""- 自動回答以上: 通常回答
- 候補表示以上: 近いFAQ候補を表示
- 候補表示未満: 該当なしとして追加情報フォームへ""")

            col_th1, col_th2 = st.columns(2)
            with col_th1:
                if st.button("💾 検索設定を保存", width="stretch"):
                    ok, settings = save_search_settings(admin_answer_threshold, admin_suggest_threshold)
                    st.session_state.search_threshold = settings["answer_threshold"]
                    st.session_state.suggest_threshold = settings["suggest_threshold"]
                    if ok:
                        st.success(f"保存しました。自動回答={settings['answer_threshold']:.2f} / 候補表示={settings['suggest_threshold']:.2f}")
                    else:
                        st.warning("ローカル保存またはGitHub保存に失敗した可能性があります。設定値自体はこのセッションに反映しています。")
                    st.rerun()
            with col_th2:
                if st.button("↩ 初期値に戻す", width="stretch"):
                    ok, settings = save_search_settings(DEFAULT_SEARCH_THRESHOLD, DEFAULT_SUGGEST_THRESHOLD)
                    st.session_state.search_threshold = settings["answer_threshold"]
                    st.session_state.suggest_threshold = settings["suggest_threshold"]
                    if ok:
                        st.success("検索設定を初期値に戻しました。")
                    else:
                        st.warning("初期値に戻しましたが、外部保存に失敗した可能性があります。")
                    st.rerun()

        with st.expander("📂 FAQ管理（Excelダウンロード / アップロード）", expanded=True):
            st.caption("管理者は FAQ を Excel(.xlsx) で一括入出力できます。500件以上でもまとめて置き換え可能です。推奨列名は『質問 / 回答 / カテゴリ』です。")

            if st.session_state.get("faq_replace_result"):
                st.success(st.session_state["faq_replace_result"])
                st.session_state.pop("faq_replace_result", None)

            current_faq_df = normalize_faq_columns(read_csv_flexible(FAQ_PATH)) if FAQ_PATH.exists() else pd.DataFrame(columns=["question", "answer", "category"])
            excel_bytes = faq_df_to_excel_bytes(current_faq_df)
            st.download_button(
                "⬇ 現在のFAQをExcelでダウンロード",
                data=excel_bytes,
                file_name="faq.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
            st.caption(f"現在登録中のFAQ件数: {len(current_faq_df)} 件")

            uploaded_faq = st.file_uploader(
                "FAQファイルをアップロード",
                type=["xlsx", "xls", "csv"],
                key="faq_excel_uploader_admin",
                help="Excel(.xlsx) 推奨。質問 / 回答 / カテゴリ、または question / answer / category に対応。",
            )

            if uploaded_faq is not None:
                try:
                    incoming_df = read_faq_uploaded_file(uploaded_faq.name, uploaded_faq.getvalue())
                    st.success(f"アップロード確認OK: {len(incoming_df)} 件のFAQを検出しました。")
                    preview_df = incoming_df.rename(columns={"question": "質問", "answer": "回答", "category": "カテゴリ"})
                    st.dataframe(preview_df.head(20), width="stretch", height=420)
                    if len(incoming_df) > 20:
                        st.caption(f"先頭20件を表示中です。保存対象は全 {len(incoming_df)} 件です。")

                    if st.button("📥 この内容でFAQを反映する", type="primary", key="replace_faq_excel_admin", width="stretch"):
                        with st.spinner("FAQを保存しています..."):
                            saved = save_faq_csv_full(FAQ_PATH, incoming_df)
                            reloaded_df = normalize_faq_columns(read_csv_flexible(FAQ_PATH)) if FAQ_PATH.exists() else pd.DataFrame(columns=["question", "answer", "category"])

                            try:
                                load_faq_index.clear()
                            except Exception:
                                pass
                            try:
                                st.cache_resource.clear()
                            except Exception:
                                pass
                            try:
                                st.cache_data.clear()
                            except Exception:
                                pass

                            if int(saved) != int(len(reloaded_df)):
                                st.error(f"保存件数と再読込件数が一致しません。保存: {saved} 件 / 再読込: {len(reloaded_df)} 件")
                            else:
                                msg = f"FAQを {saved} 件反映しました。現在登録中のFAQ件数も {len(reloaded_df)} 件です。"
                                st.session_state["faq_replace_result"] = msg
                                st.success(msg)
                                st.info("FAQの反映が完了しました。再読み込みは不要です。GitHub永続化ONなら自動で外部保存されます。")
                                current_faq_df = reloaded_df
                except Exception as e:
                    st.error(f"FAQファイルの取込でエラー: {e}")

        # ===== FAQ自動生成（該当なしログ → FAQ案）=====
        
        # =========================
        # 管理者向け資料（PDF）ダウンロード
        # =========================
        with st.expander("📘 管理者向け資料（PDF）", expanded=False):
            if not REPORTLAB_AVAILABLE:
                st.warning("PDF出力には reportlab が必要です。requirements.txt に reportlab を追加してください。")
            else:
                col_a, col_b = st.columns(2)
                with col_a:
                    ops_pdf = generate_ops_manual_pdf()
                    st.download_button(
                        "📄 操作説明書PDFをダウンロード",
                        data=ops_pdf,
                        file_name="操作説明書_情シス問い合わせAI.pdf",
                        mime="application/pdf",
                        width="stretch",
                    )
                with col_b:
                    proposal_pdf = generate_sales_proposal_pdf()
                    st.download_button(
                        "📑 提案資料PDFをダウンロード",
                        data=proposal_pdf,
                        file_name="提案資料_情シス問い合わせAI.pdf",
                        mime="application/pdf",
                        width="stretch",
                    )
                st.caption("※ どちらもアプリの現状に合わせて自動生成されます（必要に応じて文面はカスタマイズ可能）。")

        with st.expander("⏰ Render無料プラン常時起動支援", expanded=False):
            st.caption("Render無料プランのスリープを減らすため、GitHub Actionsから一定間隔でRender URLへアクセスする設定を生成します。")
            st.warning("app.py単体では、サービスが完全にスリープした後に自力で自分自身を起こすことはできません。常時起動に近づけるには、外部からの定期アクセスが必要です。")

            default_keepalive_url = normalize_public_base_url(os.environ.get("RENDER_EXTERNAL_URL", ""))
            keepalive_url = st.text_input(
                "Renderの公開URL",
                value=default_keepalive_url,
                placeholder="https://あなたのRenderURL.onrender.com",
                help="Renderで公開しているこのアプリのURLを入力してください。例: https://helpdesk-ai-xxxx.onrender.com",
                key="keepalive_render_url",
            )
            cron_options = {
                "5分ごと": "*/5 * * * *",
                "10分ごと": "*/10 * * * *",
                "14分ごと": "*/14 * * * *",
            }
            picked_label = st.selectbox("GitHub Actionsの実行間隔", list(cron_options.keys()), index=1, key="keepalive_cron_label")
            cron_expr = cron_options[picked_label]

            normalized_url = normalize_public_base_url(keepalive_url)
            if normalized_url:
                st.code(normalized_url, language="text")
                st.caption("このURLへGitHub Actionsから定期アクセスします。")
            else:
                st.info("まずは Render の公開URL を入力してください。")

            workflow_yaml = build_render_keepalive_workflow_yaml(normalized_url, cron_expr=cron_expr)
            zip_bytes = build_keepalive_zip_bytes(normalized_url, cron_expr=cron_expr)

            col_keep1, col_keep2 = st.columns(2)
            with col_keep1:
                st.download_button(
                    "⬇ GitHub Actions設定ZIPをダウンロード",
                    data=zip_bytes,
                    file_name="render_keepalive_actions.zip",
                    mime="application/zip",
                    width="stretch",
                )
            with col_keep2:
                st.download_button(
                    "⬇ render-keepalive.yml をダウンロード",
                    data=workflow_yaml.encode("utf-8"),
                    file_name="render-keepalive.yml",
                    mime="text/yaml",
                    width="stretch",
                )

            st.markdown("**導入手順**")
            st.write("1. ダウンロードしたZIPを展開し、`.github/workflows/render-keepalive.yml` をGitHubリポジトリへ追加します。")
            st.write("2. GitHubへpushすると、Actionsが定期実行されます。")
            st.write("3. Render無料プランのスリープ復帰待ちを減らせます。")

            with st.expander("生成される GitHub Actions YAML を見る", expanded=False):
                st.code(workflow_yaml, language="yaml")

        st.markdown("---")
        with st.expander("🧠 FAQ自動生成（該当なしログ → FAQ案）", expanded=False):
            st.caption("『該当なし』ログからFAQを自動生成し、faq.csvへ追記できます。")

            log_files = list_log_files()
            if not log_files:
                st.info("まだ nohit_*.csv がありません。まず質問して『該当なし』を発生させてください。")
            else:
                labels = [f.name for f in log_files[:15]]
                pick = st.selectbox("参照するログファイル", labels, index=0)
                picked_path = next((p for p in log_files if p.name == pick), log_files[0])

                max_q = st.slider("生成に使う質問数（重複除外後）", 10, 200, 60, step=10)
                n_items = st.slider("生成するFAQ件数", 3, 20, 8)

                col_seed1, col_seed2 = st.columns([2, 3])
                with col_seed1:
                    if st.button("🧪 デモ用に定番質問を追加（20件）"):
                        added = seed_nohit_questions(20)
                        st.success(f"nohitログに {added} 件追加しました。")
                        st.rerun()
                with col_seed2:
                    st.caption("※ 本番前にFAQ生成を試すためのテストデータです（channel=seedで記録）。")

                if st.button("🤖 FAQ案を自動生成", type="primary"):
                    with st.spinner("FAQ案を生成中..."):
                        qs = load_nohit_questions_from_logs([picked_path], max_questions=max_q)

                        # 生成前に「有効質問数」を可視化（原因切り分け）
                        st.info(f"ログから抽出できた有効質問数（重複除外後）：{len(qs)} 件")
                        if len(qs) < 5:
                            st.session_state.generated_faq_df = pd.DataFrame(columns=["category", "question", "answer"])
                            st.warning("有効な質問が少なすぎてFAQを生成できません。ログのCSV形式（カラム名/文字コード/区切り）を確認してください。")
                        else:
                            try:
                                gen_df = generate_faq_candidates(qs, n_items=n_items)
                            except Exception:
                                gen_df = pd.DataFrame(columns=["category", "question", "answer"])
                            st.session_state.generated_faq_df = gen_df

                gen_df = st.session_state.get("generated_faq_df")
                if isinstance(gen_df, pd.DataFrame) and len(gen_df) > 0:
                    st.markdown("### ✅ 生成結果（編集して保存できます）")
                    edited = st.data_editor(
                        gen_df,
                        num_rows="dynamic",
                        width="stretch",
                        key="faq_editor",
                    )

                    col_a, col_b = st.columns(2)
                    with col_a:
                        if st.button("💾 faq.csv に追記"):
                            added = append_faq_csv(FAQ_PATH, edited.rename(columns={"category": "category"}))
                            if added > 0:
                                st.success(f"faq.csv に {added} 件追記しました。")
                                # 反映のため再読み込み
                                st.session_state.generated_faq_df = pd.DataFrame()
                                st.rerun()
                            else:
                                st.warning("追記できる新規FAQがありません（重複/空欄の可能性）。")

                    with col_b:
                        if st.button("🧹 生成結果をクリア"):
                            st.session_state.generated_faq_df = pd.DataFrame()
                            st.rerun()
                elif isinstance(gen_df, pd.DataFrame) and len(gen_df) == 0 and st.session_state.get("generated_faq_df") is not None:
                    st.warning("FAQ案が生成できませんでした。ログの内容が少ないか、出力形式が崩れています。")
# ======================
# セッション初期化
# ======================
if "used_hits" not in st.session_state:
    st.session_state.used_hits = []

if "messages" not in st.session_state:
    st.session_state.messages = []

if "pending_q" not in st.session_state:
    st.session_state.pending_q = ""

if "search_threshold" not in st.session_state:
    st.session_state.search_threshold = float(SEARCH_SETTINGS.get("answer_threshold", DEFAULT_SEARCH_THRESHOLD))

if "suggest_threshold" not in st.session_state:
    st.session_state.suggest_threshold = float(SEARCH_SETTINGS.get("suggest_threshold", DEFAULT_SUGGEST_THRESHOLD))


# ======================
# チャット履歴表示
# ======================
for m in st.session_state.messages:
    with st.chat_message(m.get("role", "assistant")):
        st.markdown(m.get("content", ""))


# ======================
# 「参照FAQ」表示
# ======================
with st.expander("参照したFAQ（根拠）を見る"):
    used_hits = st.session_state.used_hits
    if not used_hits:
        st.markdown('<div class="refbox">該当なし（スコアが低いため問い合わせ誘導）</div>', unsafe_allow_html=True)
    else:
        for i, (row, score) in enumerate(used_hits, 1):
            render_match_bar(score)
            q_html = str(row.get("question", "")).replace("\n", "<br>")
            a_html = str(row.get("answer", "")).replace("\n", "<br>")
            cat = str(row.get("category", ""))
            match_pct = int(max(0.0, min(1.0, float(score))) * 100)
            st.markdown(
                f"""
<div class="refbox">
<b>FAQ{i}</b>（一致度：{match_pct}% / category={cat}）<br>
<b>Q:</b> {q_html}<br>
<b>A:</b> {a_html}
</div>
""",
                unsafe_allow_html=True,
            )


# ======================
# おすすめ質問ボタン（3つ）
# ======================
st.markdown("### 💡 おすすめ質問（クリックで送信）")
c1, c2, c3 = st.columns(3)

if c1.button("🔐 パスワードを忘れた"):
    st.session_state.pending_q = "パスワードを忘れました"
    st.rerun()

if c2.button("🧩 アカウントがロックされた"):
    st.session_state.pending_q = "アカウントがロックされました"
    st.rerun()

if c3.button("🌐 VPNに接続できない"):
    st.session_state.pending_q = "VPNに接続できません"
    st.rerun()



def build_suggest_answer(user_q: str, hits) -> str:
    if not hits:
        return nohit_template()
    row, score = hits[0]
    q = str(row.get("question", "")).strip()
    a = str(row.get("answer", "")).strip()
    cat = str(row.get("category", "")).strip()
    parts = [
        "入力内容に近いFAQ候補が見つかりました。完全一致ではありませんが、まずはこちらを確認してください。",
    ]
    if q:
        parts.append(f"【候補FAQ】{q}")
    if cat:
        parts.append(f"【カテゴリ】{cat}")
    if a:
        parts.append(f"【回答】\n{a}")
    parts.append("解決しない場合は、下の『追加情報を記録』から状況を補足してください。")
    return "\n\n".join(parts)


def render_nohit_extra_form(info: dict | None = None, expanded: bool = True):
    """『該当なし』直後に表示する追加情報フォーム（端末/利用場所/ネットワーク等）。"""
    info = info or (st.session_state.get("pending_nohit", {}) or {})
    with st.expander("📝 追加情報を記録（任意）", expanded=expanded):
        st.caption("該当なしのときだけ、状況を少しだけ補足するとFAQが育ちやすくなります。")
        with st.form("nohit_extra_form", clear_on_submit=False):
            c1, c2, c3 = st.columns(3)
            with c1:
                device = st.selectbox(
                    "端末",
                    ["", "Windows", "Mac", "iPhone/iPad", "Android", "不明"],
                    index=0,
                    key="nohit_device",
                )
            with c2:
                location = st.selectbox(
                    "利用場所",
                    ["", "社内", "社外", "不明"],
                    index=0,
                    key="nohit_location",
                )
            with c3:
                network = st.selectbox(
                    "ネットワーク",
                    ["", "Wi-Fi", "有線", "VPN", "モバイル回線", "不明"],
                    index=0,
                    key="nohit_network",
                )

            impact = st.selectbox(
                "影響範囲",
                ["", "自分のみ", "他の人も", "不明"],
                index=0,
                key="nohit_impact",
            )
            error_text = st.text_area(
                "エラー内容（任意）",
                placeholder="例：0x80190001 / '資格情報が無効です' など",
                key="nohit_error_text",
            )

            submitted = st.form_submit_button("✅ この内容で記録")
            if submitted:
                ok = update_nohit_record(
                    day=str(info.get("day", "")),
                    timestamp=str(info.get("timestamp", "")),
                    question=str(info.get("question", "")),
                    extra={
                        "device": device,
                        "location": location,
                        "network": network,
                        "impact": impact,
                        "error_text": error_text,
                        "channel": "web",
                    },
                )
                if ok:
                    st.success("追加情報をログに保存しました。ありがとうございます！")
                    # 次回以降はフォームを閉じる（保持は消す）
                    st.session_state["pending_nohit_active"] = False
                else:
                    st.warning("保存に失敗しました（もう一度お試しください）。")


# ======================
# 入力 → 検索 → 回答
# ======================
# ===== 該当なし（nohit）の追加情報フォーム =====
if st.session_state.get("pending_nohit_active"):
    render_nohit_extra_form(expanded=True)

# 先に chat_input を描画（画面下に固定されます）
chat_typed = st.chat_input("質問を入力してください")

# 入力が無いときは pending_q（おすすめボタン等）を使う
user_q = (chat_typed or st.session_state.get("pending_q", ""))

# 「おすすめ質問」など pending_q から来たかどうか
used_pending = (not chat_typed) and bool(st.session_state.get("pending_q", ""))

if user_q:
    st.session_state.pending_q = ""

    st.session_state.messages.append({"role": "user", "content": user_q})
    with st.chat_message("user"):
        st.markdown(user_q)

    hits = retrieve_faq(user_q)
    best_score = hits[0][1] if hits else 0.0

    if hits:
        render_match_bar(best_score)

    answer_threshold = current_search_threshold()
    suggest_threshold = current_suggest_threshold()

    if best_score < suggest_threshold:
        used_hits = []
        answer = nohit_template()
        ts_nohit = log_nohit(user_q)
        st.session_state["last_nohit"] = {"day": datetime.now().strftime("%Y%m%d"), "timestamp": ts_nohit, "question": user_q}
        was_nohit = True
        was_suggest = False
    elif best_score < answer_threshold:
        used_hits = hits[:1]
        answer = build_suggest_answer(user_q, used_hits)
        was_nohit = False
        was_suggest = True
    else:
        used_hits = hits
        was_nohit = False
        was_suggest = False
        faq_answer = ""
        try:
            faq_answer = str(hits[0][0].get("answer", "")).strip()
        except Exception:
            faq_answer = ""
        FAQ_DIRECT_SCORE = max(answer_threshold, 0.35)
        if faq_answer and best_score >= FAQ_DIRECT_SCORE:
            answer = faq_answer
        else:
            prompt = build_prompt(user_q, hits)
            try:
                answer = llm_chat(
                    [
                        {"role": "system", "content": "あなたは情シス担当です。FAQの内容を優先し、必ず日本語で簡潔に回答してください。"},
                        {"role": "user", "content": prompt},
                    ]
                )
                if not str(answer).strip() and faq_answer:
                    answer = faq_answer
            except Exception:
                answer = faq_answer if faq_answer else "現在AIの回答機能でエラーが発生しています。しばらくしてから再度お試しください。"

    st.session_state.used_hits = used_hits

    # 利用ログ（削減時間の見える化用）
    top_cat = ""
    if used_hits:
        try:
            top_cat = str(used_hits[0][0].get("category", ""))
        except Exception:
            top_cat = ""
    log_interaction(user_q, matched=(best_score >= answer_threshold), best_score=best_score, category=top_cat)

    with st.chat_message("assistant"):
        answer_html = str(answer).replace("\n", "<br>")
        st.markdown(f'<div class="answerbox">{answer_html}</div>', unsafe_allow_html=True)

        if 'was_nohit' in locals() and was_nohit:
            # 「該当なし」のとき、追加情報フォームを"次のrerunでも"表示できるように保持
            st.session_state["pending_nohit_active"] = True
            st.session_state["pending_nohit"] = st.session_state.get("last_nohit", {})
            st.info("該当なしログに追加しました。必要なら下の『追加情報を記録』で状況を補足できます。")
        elif 'was_suggest' in locals() and was_suggest:
            st.info(f"近いFAQ候補を表示しています（スコア: {best_score:.2f} / 自動回答しきい値: {answer_threshold:.2f}）。")
            st.caption("管理者はサイドバーの『検索精度設定』から判定基準を調整できます。")

    st.session_state.messages.append({"role": "assistant", "content": str(answer)})

    # nohit はフォーム表示のために再描画
    if "was_nohit" in locals() and was_nohit:
        st.rerun()

    # おすすめ質問ボタンから自動送信した場合は、もう一度 rerun して入力欄を確実に表示
    if used_pending:
        st.rerun()